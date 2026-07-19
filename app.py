import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import datetime
import io
import math
import hashlib
import secrets as pysecrets
import base64
import json
from datetime import date
import time as _time
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from sqlalchemy import create_engine, text, inspect, event
from sqlalchemy.exc import OperationalError, InterfaceError as SAInterfaceError, DBAPIError

# ═══════════════════════════════════════════════
# DATABASE SYSTEM — CLOUD (pg8000 driver — pure-Python, avoids psycopg2
# install issues on some Streamlit Cloud environments)
# ═══════════════════════════════════════════════
# BUGFIX: this file previously had TWO conflicting get_conn()/_get_engine()
# definitions (looked like a merge accident — one used psycopg2 pool
# settings, one used pg8000 with dead/unreachable code after an early
# `return`). Python silently used whichever was defined LAST, leaving stale
# dead code sitting in the file. Cleaned up into one consistent version here.
@st.cache_resource(show_spinner=False)
def _get_engine():
    db_url = st.secrets.get("DB_URL", "sqlite:///textile_inventory.db")

    # Auto-convert to the pg8000 driver regardless of what prefix is in
    # secrets.toml, since pg8000 is pure-Python and avoids psycopg2's native
    # build/install failures on some Streamlit Cloud environments.
    if "postgresql+psycopg2://" in db_url:
        db_url = db_url.replace("postgresql+psycopg2://", "postgresql+pg8000://")
    elif db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql+pg8000://", 1)
    elif db_url.startswith("postgresql://") and "+pg8000" not in db_url:
        db_url = db_url.replace("postgresql://", "postgresql+pg8000://", 1)

    # CRITICAL FIX (production-critical 20-minute stalls): NOTHING in the
    # previous config bounded how long a single connection attempt or query
    # could hang. pg8000's socket has no default timeout, and Postgres has
    # no default statement_timeout — so if the network path to the DB is
    # degraded (not fully down, just slow/lossy), a connect or a query can
    # sit there for minutes with zero ceiling. These caps make every attempt
    # fail FAST (single-digit seconds) instead of hanging.
    #
    # BUGFIX (this crashed every single connection attempt): the first cut
    # of this fix passed options="-c statement_timeout=..." in connect_args.
    # That's a psycopg2-ism — pg8000's connect() has no "options" parameter
    # at all, so EVERY connection attempt raised
    #   TypeError: connect() got an unexpected keyword argument 'options'
    # before it ever touched the network. Removed. The socket timeout below
    # (which pg8000 does support) stays; statement_timeout is now set the
    # driver-agnostic way, via an on-connect hook right after this function.
    connect_args = {}
    if db_url.startswith("postgresql+pg8000"):
        connect_args = {"timeout": 8}  # socket connect timeout, seconds

    engine = create_engine(
        db_url,
        pool_pre_ping=True,
        pool_size=5,
        max_overflow=10,
        pool_recycle=300,
        connect_args=connect_args,
    )

    if db_url.startswith("postgresql+pg8000"):
        @event.listens_for(engine, "connect")
        def _set_statement_timeout(dbapi_conn, conn_record):
            # Runs once per new physical connection (not per checkout), right
            # after pg8000 connects — caps any single query at 15s server-side
            # so a bad query can't hang a page indefinitely.
            cur = dbapi_conn.cursor()
            try:
                cur.execute("SET statement_timeout = 15000")
            finally:
                cur.close()

    return engine

# RELIABILITY FIX: crash seen in production —
#   pg8000.core.InterfaceError: network error
# fired from inside scalar()/q(), which crashed the whole page. This is a
# connection dropped mid-session by the DB provider (idle-connection kill,
# brief network blip) — NOT a SQL error, and NOT something pool_pre_ping
# catches, since pre_ping only checks a connection when it's checked OUT of
# the pool, not while a query is actively running on it. The fix below
# detects this class of error and retries exactly once on a fresh
# connection, so a one-off network hiccup no longer takes the app down.
def _is_transient_db_error(exc):
    msg = str(exc).lower()
    return any(s in msg for s in (
        "network error", "connection", "closed", "reset", "broken pipe",
        "timeout", "eof", "server closed", "could not receive data",
    ))

def _qmark_to_named(sql, params):
    """Converts sqlite-style '?' placeholders + a positional params list into
    SQLAlchemy named-bind SQL + a params dict, so old call-sites work as-is."""
    out, pdict, i = [], {}, 0
    for ch in sql:
        if ch == "?":
            pname = f"p{i}"
            out.append(f":{pname}")
            pdict[pname] = params[i]
            i += 1
        else:
            out.append(ch)
    return "".join(out), pdict

# PRODUCTION-CRITICAL AUDIT TOOLING: records every DB call slower than
# SLOW_QUERY_THRESHOLD_S. Printed to stdout (visible in Streamlit Cloud's
# "Manage app" → logs, same place the original network-error traceback was
# recorded) AND kept in an in-memory ring buffer surfaced to Admin/CEO in
# the app itself under 👤 User Management → 🩺 Performance Diagnostics —
# so the exact slow function/query is visible without needing log access.
SLOW_QUERY_THRESHOLD_S = 2.0
_SLOW_QUERY_LOG = []  # process-global ring buffer, capped below

def _log_slow_query(sql, elapsed, error=None, retry=False):
    if elapsed < SLOW_QUERY_THRESHOLD_S and not error:
        return
    entry = {
        "ts": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "elapsed_s": round(elapsed, 2),
        "sql": " ".join(sql.split())[:200],
        "error": error,
        "retry": retry,
    }
    tag = "ERROR" if error else ("RETRY" if retry else "SLOW")
    print(f"[{tag} QUERY] {entry['elapsed_s']}s :: {entry['sql']}" + (f" :: {error}" if error else ""))
    _SLOW_QUERY_LOG.append(entry)
    del _SLOW_QUERY_LOG[:-100]  # keep only the most recent 100

class _CompatConn:
    """Thin wrapper so `conn = get_conn(); conn.execute(sql, [params]); 
    conn.commit(); conn.close()` (written for sqlite3) keeps working unchanged
    against a SQLAlchemy engine/connection for Postgres or MySQL."""
    def __init__(self, sa_conn):
        self._conn = sa_conn
    def execute(self, sql, params=None):
        named_sql, pdict = _qmark_to_named(sql, params or [])
        t0 = _time.monotonic()
        try:
            result = self._conn.execute(text(named_sql), pdict)
            _log_slow_query(sql, _time.monotonic() - t0)
            return result
        except (OperationalError, SAInterfaceError, DBAPIError) as e:
            _log_slow_query(sql, _time.monotonic() - t0, error=str(e))
            if not _is_transient_db_error(e):
                raise
            # Connection died mid-session (see _is_transient_db_error note
            # above). Dispose the pool so no other stale connections get
            # handed out either, open a fresh one, and retry this exact
            # statement once before giving up.
            try:
                self._conn.close()
            except Exception:
                pass
            _get_engine().dispose()
            self._conn = _get_engine().connect()
            t1 = _time.monotonic()
            result = self._conn.execute(text(named_sql), pdict)
            _log_slow_query(sql, _time.monotonic() - t1, retry=True)
            return result
    def commit(self):
        self._conn.commit()
    def close(self):
        self._conn.close()

@st.cache_resource(show_spinner=False)
def _init_schema():
    engine = _get_engine()
    dialect = engine.dialect.name  # 'postgresql', 'mysql', or 'sqlite' (local dry-test only)
    if dialect == "postgresql":
        pk = "SERIAL PRIMARY KEY"
    elif dialect == "mysql":
        pk = "INT AUTO_INCREMENT PRIMARY KEY"
    else:
        pk = "INTEGER PRIMARY KEY AUTOINCREMENT"

    # PHASE 1 — tables. Its own transaction.
    with engine.begin() as conn:
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS sheet_orders (
                id {pk},
                call_off_no TEXT, po_no TEXT, sale_contract TEXT,
                brand TEXT, article TEXT, category TEXT, order_qty REAL)"""))
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS inventory (
                id {pk},
                call_off_no TEXT, contract_no TEXT, dc_no TEXT, po_no TEXT,
                article TEXT, category TEXT, qty REAL,
                entry_date TEXT, remark TEXT, company_token TEXT DEFAULT '',
                style_type TEXT DEFAULT '—')"""))
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS bilty (
                id {pk},
                call_off_no TEXT, contract_no TEXT, article TEXT, category TEXT,
                qty REAL, cartons INTEGER, transport_mode TEXT,
                bilty_date TEXT, created_at TEXT)"""))
        # NEW: multi-user login + role-based access control
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS app_users (
                id {pk},
                username TEXT UNIQUE, password_hash TEXT, role TEXT,
                full_name TEXT, created_at TEXT)"""))

    # PHASE 2 — column migration. Its own transaction, only runs at all if
    # something is actually missing (cuts a round-trip on every normal boot).
    insp = inspect(engine)
    inv_cols = [c["name"] for c in insp.get_columns("inventory")]
    missing_cols = [c for c in ["entry_date", "remark", "company_token", "contract_no", "style_type"] if c not in inv_cols]
    if missing_cols:
        with engine.begin() as conn:
            for col in missing_cols:
                conn.execute(text(f"ALTER TABLE inventory ADD COLUMN {col} TEXT DEFAULT ''"))

    # NEW ADDITION: item_description column on inventory — stores the exact
    # accessory wording as it appears on the buyer's Purchase Order PDF
    # (e.g. "INLAY CARD FITTED 36.5X40.5 CM- DIXX JERSEY"), captured once per
    # DC line at entry time. Powers the Ditto DC Excel export below.
    missing_desc_cols = [c for c in ["item_description"] if c not in inv_cols]
    if missing_desc_cols:
        with engine.begin() as conn:
            for col in missing_desc_cols:
                conn.execute(text(f"ALTER TABLE inventory ADD COLUMN {col} TEXT DEFAULT ''"))

    # NEW ADDITION: destination column on inventory — the "SOHRAB/HSU"-style
    # destination field shown on the DC layout (top-right, replacing the old
    # plain Brand slot).
    missing_dest_cols = [c for c in ["destination"] if c not in inv_cols]
    if missing_dest_cols:
        with engine.begin() as conn:
            for col in missing_dest_cols:
                conn.execute(text(f"ALTER TABLE inventory ADD COLUMN {col} TEXT DEFAULT ''"))

    # NEW ADDITION: last_seen column on app_users, powers the Active Users
    # display. Same isolated-migration pattern as above.
    au_cols = [c["name"] for c in insp.get_columns("app_users")]
    missing_au_cols = [c for c in ["last_seen"] if c not in au_cols]
    if missing_au_cols:
        with engine.begin() as conn:
            for col in missing_au_cols:
                conn.execute(text(f"ALTER TABLE app_users ADD COLUMN {col} TEXT DEFAULT ''"))

    # NEW ADDITION: variant column on sheet_orders — powers the "290 Call-Off"
    # dual-pack (Jersey/Molton) detection. Blank for every existing row
    # (old data/behavior completely unaffected); only used going forward
    # when the new Packaging Detail parser tags a Jersey/Molton variant.
    so_cols = [c["name"] for c in insp.get_columns("sheet_orders")]
    missing_so_cols = [c for c in ["variant"] if c not in so_cols]
    if missing_so_cols:
        with engine.begin() as conn:
            for col in missing_so_cols:
                conn.execute(text(f"ALTER TABLE sheet_orders ADD COLUMN {col} TEXT DEFAULT ''"))

    # NEW ADDITION: company_type column on inventory — which brand's letterhead
    # (Vertex Packaging vs VOGUE Printers) prints on this DC's PDF. Defaults
    # to Vertex for every existing row so old DCs keep printing exactly as
    # before; only new entries get to pick VOGUE.
    missing_company_cols = [c for c in ["company_type"] if c not in inv_cols]
    if missing_company_cols:
        with engine.begin() as conn:
            for col in missing_company_cols:
                conn.execute(text(f"ALTER TABLE inventory ADD COLUMN {col} TEXT DEFAULT 'Vertex Packaging'"))

    # PHASE 2b — NEW ADDITION: rider_expenses table, powers the "💸 Daily
    # Expenses / Staff Expense" tab. bill_image stores the receipt photo
    # (camera capture or gallery upload) as raw bytes — dialect-aware type
    # since "BLOB" isn't valid on Postgres (needs BYTEA) or ideal on MySQL
    # (LONGBLOB, since a phone photo can exceed MySQL's plain BLOB 64KB cap).
    if dialect == "postgresql":
        blob_type = "BYTEA"
    elif dialect == "mysql":
        blob_type = "LONGBLOB"
    else:
        blob_type = "BLOB"
    with engine.begin() as conn:
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS rider_expenses (
                id {pk},
                date TEXT, user_id TEXT, entered_by TEXT, category TEXT,
                amount REAL, reference_no TEXT, remarks TEXT,
                bill_image {blob_type})"""))
    # Column migration for rider_expenses, same isolated-transaction pattern
    # used for inventory above — safe to re-run, only fires if missing.
    re_cols = [c["name"] for c in insp.get_columns("rider_expenses")]
    missing_re_cols = [c for c in ["bill_image"] if c not in re_cols]
    if missing_re_cols:
        with engine.begin() as conn:
            for col in missing_re_cols:
                conn.execute(text(f"ALTER TABLE rider_expenses ADD COLUMN {col} {blob_type}"))

    # PHASE 3 — indexes. BUGFIX: these used to run in the SAME transaction as
    # everything else, wrapped in a per-statement try/except. On PostgreSQL,
    # one failed statement (e.g. index already exists) poisons the WHOLE
    # transaction — every statement after it (including admin-seeding below)
    # then also fails, even though the Python try/except silently swallowed
    # it. Since _init_schema() then raised, st.cache_resource never cached a
    # successful run, so this entire setup was silently re-running on every
    # single page load. Fixed by:
    #  - Using "CREATE INDEX IF NOT EXISTS" directly on Postgres/SQLite (one
    #    idempotent round-trip each, no separate existence-check needed).
    #  - Giving each statement its OWN transaction on MySQL (which lacks
    #    IF NOT EXISTS for indexes), so one failure can't cascade.
    index_defs = [
        ("idx_inv_article_category", "inventory(article, category)"),
        ("idx_inv_calloff",          "inventory(call_off_no)"),
        ("idx_inv_contract",         "inventory(contract_no)"),
        ("idx_inv_po",               "inventory(po_no)"),
        ("idx_inv_dc",               "inventory(dc_no)"),
        ("idx_inv_token",            "inventory(company_token)"),
        ("idx_inv_entrydate",        "inventory(entry_date)"),
        ("idx_so_calloff",           "sheet_orders(call_off_no)"),
        ("idx_so_article_category",  "sheet_orders(article, category)"),
        ("idx_so_contract",          "sheet_orders(sale_contract)"),
        ("idx_so_po",                "sheet_orders(po_no)"),
        ("idx_so_dc_article_cat",    "sheet_orders(call_off_no, sale_contract, article, category)"),
        ("idx_inv_dc_lines",         "inventory(dc_no, id)"),
        ("idx_inv_dc_article_cat",   "inventory(call_off_no, contract_no, article, category)"),
        ("idx_bilty_calloff_art",    "bilty(call_off_no, article, category)"),
        ("idx_bilty_contract",       "bilty(contract_no)"),
        ("idx_rexp_user_date",       "rider_expenses(user_id, date)"),
        ("idx_rexp_category",        "rider_expenses(category)"),
    ]
    if dialect == "mysql":
        for name, target in index_defs:
            try:
                with engine.begin() as conn:
                    conn.execute(text(f"CREATE INDEX {name} ON {target}"))
            except Exception:
                pass  # already exists — isolated transaction, doesn't affect the others
    else:
        with engine.begin() as conn:
            for name, target in index_defs:
                conn.execute(text(f"CREATE INDEX IF NOT EXISTS {name} ON {target}"))

    # PHASE 4 — default admin seed. Its own transaction, guaranteed to run
    # regardless of what happened with indexes above.
    with engine.begin() as conn:
        user_count = conn.execute(text("SELECT COUNT(*) FROM app_users")).scalar()
        if not user_count:
            conn.execute(
                text("INSERT INTO app_users (username,password_hash,role,full_name,created_at) "
                     "VALUES (:u,:p,:r,:f,:c)"),
                {"u": "admin", "p": _hash_password("admin123"), "r": "Admin",
                 "f": "Default Admin", "c": str(datetime.datetime.now())})

    # PHASE 5 — NEW ADDITION: site_stats table for the Visit Counter. Own
    # transaction, seeds the 'total_visits' row once if missing.
    with engine.begin() as conn:
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS site_stats (
                id {pk},
                stat_key TEXT UNIQUE, stat_value INTEGER DEFAULT 0)"""))
        vc = conn.execute(text("SELECT COUNT(*) FROM site_stats WHERE stat_key='total_visits'")).scalar()
        if not vc:
            conn.execute(text("INSERT INTO site_stats (stat_key, stat_value) VALUES ('total_visits', 0)"))

    # PHASE 6 — NEW ADDITION: item_desc_cache — remembers the exact PO
    # wording an operator types for a given Article+Category, so it's
    # auto-suggested next time instead of retyping. Powers the Ditto DC
    # Excel export's "Item Code, Description, Brand" column.
    with engine.begin() as conn:
        conn.execute(text(f"""
            CREATE TABLE IF NOT EXISTS item_desc_cache (
                id {pk},
                article TEXT, category TEXT, description TEXT)"""))
    return True

def get_conn():
    _init_schema()  # cached no-op after the first call this session
    engine = _get_engine()
    return _CompatConn(engine.connect())

def _hash_password(password, salt=None):
    salt = salt or pysecrets.token_hex(16)
    h = hashlib.sha256((salt + password).encode()).hexdigest()
    return f"{salt}${h}"

def _verify_password(password, stored_hash):
    try:
        salt, h = stored_hash.split("$", 1)
        return hashlib.sha256((salt + password).encode()).hexdigest() == h
    except Exception:
        return False


# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────
ITEM_TYPES = [
    "Inlay Card / Bandrolle", 
    "Tag Card / Barcode Sticker", 
    "Barcode Item", 
    "Safety", 
    "Washing Paper", 
    "Transparent Sticker",
    "Eco Friendly"
]

STYLES_INLAY = ["Normal", "Topper", "Split"]

# NEW ADDITION: multi-company DC letterhead support. Backend logic/formulas
# stay identical for both brands — only the printed PDF header/logo changes.
COMPANY_TYPES = ["Vertex Packaging", "VOGUE Printers"]
COMPANY_LETTERHEADS = {
    "Vertex Packaging": {
        "name": "VERTEX",
        "address": "24-Abbot Road, Opposite Metropole Cinema, Lahore, Pakistan",
        "phone": "PH: +92 42 36283733&nbsp;&nbsp;&nbsp;&nbsp;Mob: +92 300 4747660",
        "email": "Email: vertex.printerlhr@gmail.com",
    },
    "VOGUE Printers": {
        "name": "VOGUE PRINTERS",
        "address": "SE-26-R-24, adjacent to Naghma Cinema, 10 Abbott Road, Lahore.",
        "phone": "PH: +92 42 36364858&nbsp;&nbsp;&nbsp;&nbsp;Mob/WhatsApp: +92 322 9959069",
        "email": "Email: vogue.printerspak@gmail.com",
    },
}

# NEW ADDITION: the actual VOGUE Printers logo (client-supplied image),
# embedded as base64 so the app stays a single self-contained file — no
# separate asset upload/path to manage on deploy. Decoded on demand inside
# the PDF generator via _vogue_logo_flowable().
VOGUE_LOGO_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAUgAAABbCAYAAAARIRyyAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAgY0hSTQAAeiYAAICEAAD6AAAAgOgAAHUw"
    "AADqYAAAOpgAABdwnLpRPAAAAAlwSFlzAAAXEQAAFxEByibzPwAAjilJREFUeF7tfQWgXcW19rUYHhLcoZQWaUspWihUqFHBJQlxT5DgkGDBNUgCcXd3CCSE"
    "uLs7cb837sn6v2+tNfvse3PlJOSl73+9Gybn3H22zKyZ+Wb5pKQUHoUUKKRAIQUKKZA3BUR2yiEROYjyv/vIXsNDh/aLHDqQRJV5zT5t3x6UQ8K/98v+Q4dE"
    "n3Fwe5LP4c37cCuecuAg7t0jhw7iqXwcPvAkFFCSf5OgBRy8/iDrv38bym58x60oh/CUA3j2wYOsG35P4lkHcSOrchBtOoC64X/+oe3a5e0WPFNkrz4Ol2nh"
    "ZQdwBncojdi+Q3gv23VI6YNr7CIth/CeQ/g7/yrxJlx8cC8K+4iFNNtpRc/jFGmGS1n00ErxJXiH0tJea+eVWgWRNKnf+fQDaKuOA322fZD2+/zdRqRAJ7aH"
    "bQB9tO6kU7jXL8yNIlG9cQ/qfxAFPWtjRDubHcbnOEH1UfG/vd18r7af/eRUUJr5X7wfncRH4kr91Gcmc/Bamw4YGhgb+/bqeLG7/X28Rk/gHVp71tcGT/RO"
    "nsc7OWYO6oDJ4wjvI7G1+qCH097eaKd34Z89oM9+0PoA6a3v1AHobfNJl0wbk7lGpxpmwkGMz/37ZA8qEiHmwt37pO/KTdJndaYMXJclfddmSu+1WdIPn/3x"
    "OQBl4HEufG/fdaiH1od1QVm3Rb8PWrlR5m7bpQBX0FQ12pCoGKAcA3qHzXb+rZ19EOCUDBH9GkItu2yXQhnHig0YTjlOAH9wwU/U6hOEOIEMCDlfDqCuB1An"
    "q2eyB1u2w1vIRxHIAEqY0IcwinU+6WJirY4OncQOjN6ugxz1DrYEjLxoEwBsOx6bteegLj48l9fByaPPxUA8CJqx6AKD90Xv0PpYHe0/P7Svj81hAMmJznc5"
    "ObQeWPDQF4dAJ8V3wx27JJ/38/fd+Gfzrv2yA5+508DBEEC7H/Tm27V32VRdUEELH53aVVEvBZAkQHt9lQyxse/jyEl5RETi6xNLj6GX1Z8E8IVAF9rwzgDK"
    "tngqoCltHMTib8+zywwGw0jU+7ko7+fiQwJiruK7LvbRKDh2/Z8bgViXnajRbrSD3cFFIgLILxatlXMajZRz2sySs9pMkTPazJAz28yVs1vPkNJtZ0qpdtPx"
    "OUvOaDv3uJVS7WbJqR2nS8n2M/DOmXJGu5lyevs5Uqr9PDm70XBpNGlRos8KHBK2+pDEBmgGmLqI6Yp3AMRxbqXAZ9kFJGg0JJQLIHdAAtt7kjoIXHgQn6Nc"
    "BZ+rE9M4W54n5CX9PB/au3HP3n2sEwccH2gAxOfZw1hH4xv1hHIyhgRcNg4ELsobQeBbsWW7jPlhm7SbuEae7zFJyjUdJvd99r082Hik/P39b+TONwfLPY2+"
    "l/s/HykPff6dPNZxnHwyeq0MXLJTlmzfkw3qtX76Zs4wtPAQh2ac8Q4TzqauTacjoUPB1GfbFZQDzii6kF6ondKNLIVxbdovKKt27ZHRK7dL1+mZ8vZXC6Va"
    "yxHyIGhw7+ej5Z9o++/fGCj/+GiYPACalG3ynTzRcaw0HrMaC/pOWbhjr9FfD3JNB5w7tleQg+fiGugQmLqwVOhFgYNyaoSqK32c49fLkh8wUY1M+vHW6ksx"
    "PrhQ+ChxKFQgj68rUb2OYJSGl3JccSwox3mAC7lxk5RgZO8WfSLHP68J9AhgngscF9zpBVzBZxMHdGziBRFAfrAsU05ssUjSe2dJSq9NktJnu6T03CWpPbfg"
    "+1ZJ6YfzffjbtkTpje+h5HWevx/tPX3w7r6ZeG+mpPbCu/msvrtQl72S1nqhvDljpQNkkqMhrHY6KRzQYqJc66EzpWqTb6RGyzEoo/MsNZuPlbpNJ0mdL8dI"
    "rSZfy3sdR0nmdus2DvC9gTNJprtchA2iOafIfpU77eY+ExbKo19+g4k4WmrmUyfWt2aL0VKt+Tip3HKy1GjyrUxbslafYTUjGBK6jS0iMJmA6bRzpiBe5a34"
    "Y+LKLGn09XQp9/m3cvXj7aV0lY5SvGoPSSnXWVLKdsEnCj/L4xxLue4o3SSlDH4v00nSK/WQ06t0kmuf6ix1WoyULmMWyA/byD0FiZ2ghCEJgLQlIoJsqzlH"
    "qS47scmbDF0LuIatjgQ0RRkbHCoexvhlTuDp63dL25FzpF7b0XLz813k7Cpt5YTKaGeFnmgjSlmUR739gQ5lQJMyHVDaSZGKHeWUah3lF8/2liotx0v7CUtl"
    "7e6IX9ZeAK9k/eRjNCwIARQiOULHrYnbcVppe3SxTaDj4OkrpVKTYfmP5xajpE7zkVK35TipjrFTpclIebPTKNkNlAo0MlmL5CEwEsgMIA3IXUWFb3vw7nf7"
    "z5AKXwyTmq1G4b0jvfB7otTWd47CeOVYnSKVvhwlg0bP0j4gMO1SGuA7QPLd3lOkbLMJUr3lWKnd/HvMA3xvMVFq4Rk8Vw0lv7ma7G+cW481+14ebzZcqmEe"
    "1fp0iMxZsCYBkJ8u2SQntlxgINgf4NhnNwByLwq/7wQwAZwIWL2PYwGwpvXcCZDebkBNoO27T1IGHJSUdkvkjZmrfBokBlsyc0chQUU7inm2LnIiVP10oKT8"
    "+1MMfEzw8pz0eZSKmAwV+mBS4PcHPpNfPt5RFm+zOnByBRBKZkGNuCKtFOGRekjjM1au3yPX120pKY80xvs65V8n1vVRTNqK/STl3qbyx9e7yPpdFNttNVSx"
    "/xCEhgOcXMaRRcK2i7NKFpTFW/dJ65FL5b53+spltVoDAFrg/QC8aoMkpfrXKKBT9f6SWgPvqjFAUmqi8LMGfq85COcHoPSXNPyeVrWPpCuYoH6PtJfij7aR"
    "G57tJO/1nSiLN5M3sHceoihHPbDriRN8G2tJKDdxNMmlMM9hEBeTbZlwbIyJruyTpQDxdmOWyt0ffS3nPdZNipQDHR5pIymVeklK1f6gBWiAksr2ghYpNUMB"
    "Tar3U9qk1ewrGdV7S0aV7pJRuaukVuyqNDjh0dZyywtd5OP+U2RFFgU5581Uj+hwFPSbOqZiulhFcdPpxgEyAKmqehxsK345FOOzFcYN3pvneEa/lufY4jUo"
    "9zWVexsNcT29AbapZbhImQrIpA2rE4VR6uH5w9KsfXJGXS6afF5HH7P4zrF7WMF7dZ5hbNz7mbT/HgCJg2N1n9LgoMzL2i9n18R1D+K6R/GMsu1tMSqHPuDi"
    "w8X50fzals8czpUeeH55PBf0OKl8S5m0DIxiOL5csl5ObT0TALQBZTMAaZtkdN8hxbob95baa7OkgbNM7Y3f8F0Lv4cSzuU8H7/+SO7BfRk9tkixbjslo2eW"
    "pPdah8/1qBe42T67JKMtOcgVPmGOACDDDHOdoXENJno+1WkyBnBbneQpNTAB8io1ARK1vpWUOoMBSO3l+hcGyRLYWOxwXVqSM9k4KS7HqiHHsOCE4TARebXD"
    "ZEkvg06rY8CTb51Q19TqQzCBB8iZlVrKyPlrrDo03OhAdvUBRW2dYICH/cYJhGP2+m3yas9xcs3TfSSD7yWXWBkDkgBYi3XAOxT8+gL8euETiwTP1wKtaoIW"
    "UbFrDTQdQHlNLV6D3/DMtLKt5ddYWJp8NU02h+7bD4A8AMMUOQn2iXMpPit10hztEQwHZnRCOQAa4F2ke3gqaTR99TZ5rctoueG5HlK0XDtMFkxO0oCLAetf"
    "6ytvB9sFAKzRXYpUx4JZE7SoyWtQFCz5O+kSFg+jg9KuKiZ4RU76dnLV872l2bB5ahDQ0UN9MRcxiJsU9Q+xjqqHi3HQOnZN9A2cZehjM+RgcYWF49qn0X+V"
    "etvClc94TmO9tb6o3yMt5JPhS7wyHCeUOKhyoI7e3+b6adJrnwI5AVKk79TVUqIq2lU9tmj6c8PzjR7hfahXtf5yfp2uMntNlo5QBX1lEA5J/6lL5RRILCnV"
    "v8JiM0DSq4NuNTD3MP/SdOzxOfnM0/zmcK6/cXxiXlcYINfX/0q2oPkRQDZeul5ObAOA7AeA7EeA3KoAldELf/feJOk9M6Vojyw7d7xK922S3oMc7GaA9hqA"
    "5BpwlPgOTrJomwXyxvQVRzRlAudoI9E5Fz1pU+S5ntMNIJV4sYGd83vNb9BpWJ1rYrJU6A6A/E4WbwkPTSiek5nMtt6zEhwU5PiMe5y4PFPOr4HVrCo6rc4w"
    "B8h86qRAhHqV6SpPtJvor6awb3xS4MOoIZWDu1THGY5F4OSe6TJGzqvbXlLLtpWMshCNq3LwEfwIdmhnjcEA4EGSBu5ROcdafVEABAqQscJ7QnGOMqU26Fkb"
    "9avN6zkI2R7QD5O3eJnmct8H38jklUpAWCxh5EKhbjgh2vEXa8HRgmTgHCMOklzKXi5GRofJa7bI4+3HyEU1OwC8wa082hPtRF2Vrqw3C0E+FNIH3DEWCpa0"
    "GgAiAKaCpF7P+3zRiICVIIpJrvTCRGSBJJLxaEe5v9EwmbHaaQAO7SBBkga2A6gj6spaRoYU5XYppwSAtDZwQbGFVmTg7PVSsgrGTzXUKb+xrAse2zlMUqsO"
    "kTMqd5TvlmywoREkC11Q6fkQdLY2eXTRdQDlOH6+63iMPyys+j7QIK+iYwhtr/k9OMLe8pd3h8lWnYK0WCdUHA26jQMH3FTHjC3KpLEtsrY4+xgsoH0FtT/6"
    "HeM7pdZQzIGeUgeqEB4RQDZatkGKt50PcKTOD1waOMiUXtQBEjAzAZAUd7eBc4S4e5xKCt6T0ht16Y969NuoHGwaOUjoItPbzZPXwUEeGU9h4kkQa9VAo4YJ"
    "G1Sv9J8mKQ+1BRAMBhAADCBSRkXBwf+GSJXGgV4bq2XVzvLTur1k2mqKi/asSGy2YVbgYeIuB5+1hgO9PHSIKeRg6uATYm0qBpW+P16PUD8/l1Kxp1zxXE+Z"
    "n2liGwevrvmmLDI3Fk4q1+Bn7T0gn4N7ubQexBRM0hSIwwS/DE4qcqO1hqON/CQo9NVBmQbx0UCDK7kNXJ4PhSt7vPB3AwtOGAIEBzafBzDnal0Fz4PYeeVT"
    "3aTfnFWqINfeUJeahHHGVrQIIgqkac4L4gBp301gX7l9rzToO1XOe5w6U9AboJ1K8VnbhzoC5LUNAfiUIwoLArlJXod+Udr01s8EPcJ9zllygagN0RvXFKk2"
    "ANfj+bVBY3JB5VrKL5/sKn2mrw1LpYGkcpAGkJGKISdAutuLukqpO4zIy72mSsrDHWxR4zjJc+ywLwA8dfFZsYvc0rC3rCH77kdYpIKkZTPFlnVdrlz83oSX"
    "//l9tAVir41Vti/3omMAdEipDTqWayP1Oo+OVB00xvC5WzA5//weFtLy6BcyLBg/qcqRg+a6YFOqoZoH74jP03zbmsf88XtSaoIOtXtLMai0uo4xLjoBkEs3"
    "SYnWixV8KF6n9+An9JHkKPtnAZi2AyQJTgDQ41QovqehmO5zqxTtjjp0h5GG+tCOc+X1mcutG5NGSRNEOO0SFjsMAa7SON79agoAEromipLaIbmXdHR8uooJ"
    "1Ef1kkvAHU1cThmbAElrbQLsCp7JwWCSuLL/7A1yWmXojrBipmJiUo+VFtWFKz6KimuhfhgsAK5i0BU2BuDxMOujgaKBi7tr+GtGLVknd73TX9IxMVMq8z1D"
    "dKDxuWkAtHR+kgOIxEau4H0h5mCCq/4xgB7pQCAgUMaL6eFUF6efBhwpCjasN56vYGNtSCnXQS6s1lo6TFzsfUpwoGgXM9ooMCSpu8iV8CZekx78t/PUtXJz"
    "g36gQXPQALqw2l9JKkToNIBXuqosUG9XbZhqgW1ne0AL/WQ/UNoIuleOG6cL6aZjJNCHIGlcKOlXBKKlAgjuTyN3zeugQz67SnvpMWmpGbHU9YUrWwKWrDt5"
    "zhY/hU43ZHHB5xjci0X/gUbfgTvrrvXLbyzbb1AhUFyF3rBWu++dpw7MhKuMOLRRjIc3gFTAhrsWj0lrtknpWtRTQ/WizyTt8ioYCxwzoOHJldtJnxmrbczS"
    "59H7Zs7abXJmTaguKmNMVeNzSGdyulyUWZxmYV7oXHDRnnOD56P5kfdcjtNGJRzMuYtrtJM5KzdrnSKA/GzpRinRBgBJ7hBAmAZwVMtxH1q0t+I7C8GRIq4X"
    "fg8lnONn/Hz8+py/FXBPKjnGXhvVMJQGwC7SHfWiyA2ATG0/X3WQRtlcZ0OOk7q8Glfi4ptCE52iXaxt+u006N5gkFCdWxjc/sm//VwquYpq7CR0WpU+cikG"
    "/eQVNDCYP6OKtQV7UnvV1anGBjqOnZgL//oIXBuMIql4blEMuHSKFgAtDjjlxrQQdHxA1AbwVOgo//pwkGS5zlGNHDaOUZ+9eD4XAXNl+nDIAjm/OhaCcuAa"
    "a3BFBhdEjsZVC2nQraZilU+F6KgTl+fBVRtnzQlOcCC3hFKdgEfx2bkqrugAgHQaLngfxEkdqARGFblMTaBiuoMHBykXnDRwsedWbyf9FtjgVIdy9A19JZUv"
    "N2zL1uHhVPZP086ZjszhkIYFQxBZtmUvxOlJUrwCOWeASHVwfzqhCIAUm8kFc1GiSOdqBOVeONFsoWDhd9JJgQ/gqvpJioDKdbpuTMeScTvGRXMxC3ozvrM3"
    "3ofnUG9XG/1QuZtcXK25dIEFWtd9N8gEFkv9DnShMJcyc9sy8SCoTeZmbpfL6qJvyZ0rKPvYjY1ho30Y09RT9pfi5VtLu/HLfN7wHca96jDy6WPCiPk/RvXD"
    "uc4TVoMDxTv1HZQaWOLvDedsgdSxUKW//OyxPrI4y3SY6qurBhqRLmOXSAY4WuozuShzDKluE4W00r91DLnKgxwp3ml95+/mOQfpw2iQC010oYYXwl/fHSq0"
    "b7LtMYDcICVg+FCxVl1qAIz8BCgqNxmBI88dp6LcIwGS+lCCNsGXoL1dMtoskbenBSt2MggZADI+v+j5T/Cw+4dOWwaXFHQyVvd8dSg6WahHQUeA87gcADVl"
    "hVlpOJV1vU/S5MrBzkiB4CbdffwPcmIFA2mulEUg7upgiICGnW6DI40raS2AKdxuSlZpJd8sNt3RHnAQGo3DhUAHuHEgS2ExrQaXoRRaZCtRxwaOibpAVXwT"
    "bINIyAltk98GJq3XmPwskV6O9eoJ7raHgqRyWqofck5LucawqhdATx3sAB1OZgzQq58bIrPWeizHfvhHOhgkXFgS/Z0bQKpWFyI6DQh7aeCgMcojO75ftFlu"
    "exXvga41pQq5GPalGTHM8m4qAlMNBF2icSBqlQ/AHk0wW6SMUyOnE0Rs7yc31ASOJt9xpfRCfWCZveSx9jJyqfWn8lU6fE32CcCU6Fkf006W7rM3S4kK0AXq"
    "IlXQWCZwo61Ve8uF1ZvJ1PWmB7V1yIIXgoHIK6F/R37jblB8otVIM2gpLQt4p9IV73y0m9z/4UQ42LNF7gXs0T9120A/qQt4AfpM5d4xLnWBJmfeF8VpD85c"
    "JZUwnyNPgxz1c8DWPi/bUZ7o4jp8gPV/H0BGvc8v5qhg0TgiI2aukDOqQUFfIEAGVp4A2VMurdPXOcgjB0i1ULqYsnHXPvkdJy/ETV3NgpFDOUXjxtIh1hcl"
    "KOlE5WDE4MBAerLTWIdB6u4YzkcHDIK1WcQnLN8lN9THPeCQU6vDikwwUM4RIKkiCwHPLNPkpMi1ZsBqqNbCwGnowOaEiq/g5LhgzcWgTCfdlOMyzkGfE1b/"
    "ZCYNri+K56Q92Foe/Xy47NRuOYAJBJ/NyL8vicWQ8w36VYroypE489l27EI5t1YHdeXIgAEsHRZoFft1cXB9ouoUY0W55gFo3wAVizNUNCZYspBbJkc9EG0n"
    "p0Ordg/0HUTDWjTGGNdMNyDSR1UWBRV9NvoaVv4/vN4TrlqmP1I+LloNgmYwCLvGUQZJql6nCZJKFzQCRUHv08WXoNxJ7nr/K9mMxSRi1BnZwminAJiuBw6q"
    "G7Pz7ZMs6Cxvrw8utCLUFGqlLgggTe2Qin5+u5+595jDkB0bEZL0u9ex2NIVpwCApDSjorUCpInf7A8dh6q64OLuXhWR9JUQxRPWdUpJA6V45U7SbeYmqwga"
    "+F8JkIkpZvodOtjymAbL8QWPYXCTs8i3k4OeCddV6SkX1wVArjQO0gLYSNxoVHm35/HhjC1//WDwZAxsrvxf6SqoIjQ7ntwJuQHoSNKhLypWrZuKEur+U7G3"
    "/OKp3vBdNCC02Djq7ijyW52Gzt8oV9TDs8rRZQfPglVZOR4YI9Kpc9NJSb0QuUroIwm+NNLQ0kixkbqZupgAdWiYIojSkMM6EUhokaZoacYHFT1VdxkGoYNm"
    "gZPGVvui4ADSqvSAv2RH6TTWVCgM4dPJE4VJ5qSl6xIC0YkX7Np9CV3Zu/1nyqmVwJFA30ojQSoALAIy5ZyDYYYGMbfYq/IfCwjbqVy2qSM4CdOhYklX5b6p"
    "DMyQwOIitqoZsHhAf2zqBKdHQXTQ6/g+vAuLWcPek31c0bE/iLoUSW3sUtdsekiTEoind76NOsBx3cTQgjhILrx41yOt5LkeE1ySD3pHs5JH/qKRvlP1FpF+"
    "e/ySjXJOdagrqpmoXmChWgIc61lVu8g3c9wdDaqUYDwcheeV5PNgNCzoWUXoZ4oF37h9SjjsL/aDi/oK2FRphDmbx3jkdVUHycWPdZG57ptK/e9/N0CqPpLO"
    "4gaQs1ZtkYueALGPECAvenyATFoVnJ6zh4/lj44JEF26eZdc8QyU3AAHusBkwL2HnJla+9RaFzgzcjDg7FQ/Bs6mXHtpNnxhBI7mw0sjBLWNB2XkvPXyk8ew"
    "snPCYIKnB0duGgjAoQa3HbPUBv1QsBaapVoHqSveddDSL7Aiffmo+8Q9dJyHI3RKVQIsV2IHFLd+Rzqn/CYP24hCcT2NbjHl+8vtL/eXdUBHFfF0wUm4j1uD"
    "w+oSB0hLlrDXY8+346dXuo2RE8tBbVGZBii4ccQni3K6JlaTAySoZeCzaLU+UgyTOE0trgHwqUJwR3nSAWoKunmppbUiJI/KnNDmvpJa/VvlqGlgU4CsTR/K"
    "JMRP52Z5P/XbF9ZuDXUDF7tEuB3df8jXqe5c1TNh8RCZumqjXFiXfRG4/CQAEgvyiXAJ6jplmZM1iLthnTcu0oDYAkTjev8vhiHApAwWH3o4BP1iAX3Nhf1X"
    "9XrJup2ui4otfk2HLQMHDU5fgS3/+rO/zCeS9MWiDmOjSV7sM9dRqvonLGB5feKa8r3kPhi3dmqWFypZ/6s4yAQYRZKKZ1EJhtFlWXvl8mdBVFjOQsfkvgo7"
    "N0ARClbsM6CDHL0sy6esu9NGckp2iMzmsJxYhuW1bhMAYm0wkdDZdOup9o0pm2sDdGo5Z6YKf6yQ1K1QXIBLxV3wIdwe3kXXPryOWlUeI+ZvMHDkJMZgKYrJ"
    "nlHVuSU3ppAzVVBzEcVcKAxIM+ALmcZCt59HAQIPtJTTqsGt6dk+cu0LPeTG+t3lxhe7yQ0vdJern+oupaiegJuUAgXAgDpFOvdaxE1BnIW5C5kIzzp8C6NB"
    "O+k4dqm2xXz8LAooDo5c4KyQmzIrNa8gbJB/frwDdGMPfmnRL+T+wBHS5cXcRDihsBABxIpWg2M4VBdFaYRBezNQ33SK0fgtnYtRNdIR7SvTWk6q2VV++nw/"
    "+TXafvMLHeW2lzoiOqiLXI7FNYORGIzugMitE1V1xxDD1SUqCfHTnczT6WpELhKeCQ2gPgkAaYY3W4QtC5VFsQTEaj1uKdx74JkAGprltyC645oqA+XyxzrK"
    "IkQPGTAY6CaWIxvTEUCqkch6gVdW+BIW8wp04mb7kuhrLhS4vnxjtsvfpIufGYSqfwl3ocpkFJKovwIjFy18UsVTBX1VGXXh/SxV8D06x9/yKrju/hbyTr9p"
    "PryYzeq/FCAt+N4GlSr+fc6t3XFQfv48iIzVzRTv6HD1+8sxyNQPkiAJUa0OOhEDsi8SZ/AIScPiK6zPaP2g3yWL6h19kE0G93lJbXKPmFjwSVMrnRpEqF+h"
    "vjCIvkPwN6N48E6I4GdU7iBDF2y0xzP1WoiSwZ+j5m+SK6guYGgcrNK0yqq7kIfFqUEi6AfVcdnFOjw/nWIKJzejMB5uJ+dXbScPfDBU3u83Q76as05mbT4g"
    "yyDFr0ITVmNcQ70pczIPSt9Za+QZ+N9dDcBMKdsc4IoQO6VVwQPdxFGK8EHUpY9cV3nww6815ZX2VzYOMsE9amo2IEfMUK0keQd+rWn0SlAd1Ldu/eQC4aKb"
    "+jl+5RxjL3xiISKnRxrT/5Sf5JThOH5O7Y7gLr6VDwZMU7FwXuZ+WY3VaCUwZTk+l4Jhn7Rml3Qct1AqIYFFqRowhDGED36kFMmVDkmAh/UL9broK4IsJvnV"
    "j7WTeZt3muqGmIimG5RRL8nsQ4looOptJ1oIn+qnkwBkXlOhh9z3/kCVOVQ3Af01qavQq9MkBpAaEpoA5OVwu7jiCXCP5Fgj63JBXCvqhjj9z779wQBSvUiM"
    "k1yHSlxbj6qQZPWZlIL6yilVO8kFCIu9uE4LuYSlVkt4l7TSz4trt0DhdxR+hu/6u/19UY3mcmWt5vLtbKh1uMhq2rP/QoC0qBKfcIpYCaYEoaRyTX0HSAdF"
    "BckcAEkfuTQqwMlBko0v00r6RACZd6hhFOYWM8yQ46veaqyKKORW1E9O9YF8NsVgui6AC1OFP0KsyInQCPJwW6nX3nRGqoti9IlrjOZCH3rtM3gWo0Ho28dn"
    "qbhCcdkU2MGnzzgpggI5U3COGGzFKgGoH2mJyJouUqfTJBkF7niPu8gYGhuyB6W6nUscP2zbIy/3mCilK4Mj5uqdhKhE53SKthq+qWI9QAwcwQVV28vUlZn+"
    "8MA9hjqYaB1yVtJROvg4txk2R04jODJ+vDbBjoBBazv1VSjqAsJFCHpY9+000QxcZW0AJ+sNffA1z/WWl3vOkEmYuW4vyVYX0oB9GK8Zv49cniV/waKS8nAb"
    "FdsjY1cBHB0XrSIQ71PBtaqxAZxk8XKtpPmYZUb1/XBfQl87r2d9747kmTj5qxcp+qPuqhstiHv038u0kUZfzfGeJWBZ/+o73Pc0csqIMvlYiwfP3yKn1ASY"
    "qbeFeQIU+F70aylE7IwIETtc4JyCQ7EAl6rWDvVPUp+JcVIcse1Pdhgvy7buQE6EHTIfZZ5+7oy+2992Pnzn3/FzC7fvkm0ayUUGxtLw/RfpIMPUNq9Dc5Xw"
    "4isisy/+8mVMHLD/gXvMFSDpm0V2vg51SxiIsDj2nGR6wKDQzgka8b9tzNkgHDx3vZSuasaDDAAgDQRqAYUxhGCpFmZyseRUAYwayQL3iF881UPmb2bcAV9q"
    "cbk81uzYI3e9TtDGMzHRybkUgfJc20EuUX31TBdmUTEUPflsWmRRh/JtpdSjzaUmInkmrtjiNnA+mXHLTCQKXSsGtHJrGg9o/ndKUc09af6WPFqMXSZnV4fR"
    "iXHHBXCR6hJEQFCRifozgATE+6Kw6Dcbbe5c2eNoYhwkfqMI7mYq+WbuWrmwKkT9SuDKCbbKHdOIRLck+m8G4xf1j1w8SAcuHABHcpvQ615YrZW81H2crNqi"
    "fFXiUN8pNp7SB/NZ0vHGQiOVE1LndjOYrARaP4DEDykPQ8QPdC4QIKm3DB4K34AmDB/tIGW/+MYXQ3KMlhjXjgQ3NwZuQWfWAljB/1G51dykn5zvBxifWrGd"
    "DJlnxhL2ZVyJYVE88QxL9tbgGvdmn5k61lJrQccd8+/MFyThe/r717+SjXvtTUxUEnjg9/ryeQiS0Nj2gjhR/I5F9YRHvpB+M9Y7PWwUROTJ3nsF/sWRa/6z"
    "5pP5XwmQlvrLh0E0vg4BAg7KHe9iQFPE1oD7oHTOvhrTs5/6KeNC8PloW+k5ebENsMO6xoaXiUOx/kEH7MSsugeuFZwARZRjwoSgAzqfC9chAxV35Aawpdfs"
    "pqJyCYTmNR8ZnOSpIrBG0ExUs9UIcH9NnfuEWxAGbTEAZNHqmDjq0MzQQXJp5qoSMtGk0cm2LMLdnmgnzWA9dug10VUdtcG50GKqRi26zhAQbfIEcDKQoIhm"
    "Mc78s/OYRbBWOicZWXLdgBCclYMbk3K4tJqTrtCXkiuBnrV8s5FKuOyJex0gXe8YGNxFm3YgSQMWB/jkMZuOOSwbZ1MUdOBiYckbqH+kNZscJQ0x5NbQz4ik"
    "uuvNATIagRPh4LMtRth8zS1ck3TgZCRg0ZndukFJoCZ0M2asQFqzf36AdzHzjHLS7h9K2itosq1WErpaTzzCUEyCNrIA3fZqH8lUQIH4G1PPmDnBBtbnQ6ag"
    "D7Eg0TuBz1QnfpMWIif9iOZ8N3WB3aFD7ic/qLHEsN/g3RvjS75DmZHEx/FejItHkAOTvqtp0JWrE3ceYr35gbJOeC8AtQ78JpVWKsEZy8J0aeU+wfilA79K"
    "PQVxo7Q8D5RLnuwjczQzFOtMgxb7xJzZTWdNf07+xpexbzh+Q+E55Vg0MIHqipD0+b8SIKNRn8eXGq3GmcOrhpHRaGCruXW8uQtYCB06T2OMMcjgXNpt8lIf"
    "OwZW5tjLr+x6hjfST83Oq94TR/uJPyCaA0p9gFjRaj1NtKI7CSytFuPrk4ZcEDibdE7ksm3k3+8MlB2+RwHBSpX1OD7+fpUUZd0BAuqn5z5iWl91e2CdCRbk"
    "1PA81b+xDRiMSO/0+1fg4rAsrMSJiRDmRIJX8d+CdSv8GdE0LnCKNOw8UtKRJUYNQaoP/dacealvhShp9SFoExgMvBW0yN2hPb9r+LXs0PnL5wZ/P+PVlcSe"
    "HosLzgNI1JvyCBJOaEIE7x8+W52GQ+QL383fuYDAI4BcV9WuoB2chJELNIt+RdaLyh0GzlUB0moQM2BEPAfOGveoZ7SjTfiesm6HXPk40p1V4rhKhGCGUMz4"
    "p8V5kz7m9KyWb7h1XYUUcfM2mDU77GbA+W7bcxjXU775GFWrEFTV5Ur9Ud2FS8HY1DN0cTLwBPhCF1jzk36Jdmn8uxlMYsypY4jrPH38zs7cASNgJ0mFUUTV"
    "I9Dfhrlh0VM2Z9RZ2z0F+HeJCm2lgxvftA0+XOhec2EtiNfwhkhOh4rnl+kG3fAoBTXWWV32omxHzrBEhr1YXzFKSN3hzGJvVXCVTWzo/9dxkNEczuPLE20w"
    "yDBo1PoYJikGluqDlPMIYXIumvIclPgBIG0Vc+2cLe9KeHIdGham5w5BFN4nt7wEIwoU5DQG0Cm7iKZzIscIg4JOXgId9Y+czBjMlfvJWRD9hi20CAt16fEo"
    "hKlrtsIfk5ZW6rssllpF8xA9EFZjDVk0wwkt2jqwwQH/9r1BssCTXHglCyJV0r8vxOp+ZT0M+kqMd+YkIgAArNQXkUYpA2zVuYKDjjub0yJ53UuDZO0WS8Jg"
    "BI0IG3F1/KXFiAVSFFww/Ts1dE/FZj4v7vbhemP2o8ZC00LdCzka20qDXtOESYJ1sihnwz1k+N4AiQYcCe6E1TFpxKYZwdviljQOyLkXjoaG/WfDmg7REcYM"
    "tajnVdS5GzSpwjRrPh5Aq5Llv5TvF6w0EtCFybke3a4Cx2pkK7/yaXBeDC+kYUh1ggRIOrBzEeKiwxBScnEQh3VxH4JMQp2F+loeZiohQFrEdVziscWR5xj+"
    "aQtIDzhUn1ihtaTDF5iqHC6AGr/u3LHpI9EHnjuU4Zz8fk71NojYIaXtNWE/oh4zN8pJVDdBFRLmWkFcZOojneSVnlP9WYieUobQ5qDBHSseuGGL17FtUbj8"
    "OefoS2GASXuYHYUAGacGvr/WZw5iggGQyn1gYKkPHAGSnc1JlwdATnEO0gOgXQpxNt+mGHWcYdQ1/naeZJCr8hC9iCt1R2MO6mLVKVJbPsZUiltlu8mzncZH"
    "8GCbciGDJD7vp2hCQ4+u1rg+hCjmNDJxkEKfWQSGg+IELOjbbn7rK5kHC76OV+oTw7jJQZtk/owbosL1fPITnadCXIZekIBc4xtMWupEPZ45iNgEbgKkGo6c"
    "44VP6k8fh751VWaYT4FvNI7cZ/GcjTvkiicpWncy/Vvwe3ORL0o0QVD07C0KFoxBB91e7THFjS1MCEe9KlUF5KIUKbMXf6vLCsp9qeaL16lF3ZP7ajYe02HO"
    "wF5Kv4QrTUlEapwJJ+hE6YLvoXSV0kgQW7JWRzkbkT5ng3MsCc6rNIwg5z/0vgyetsC5RXKpntPSF+PhC5DPFYkuUqlv9vyPGo3FOHf0tYnzjBYyTlqlH8RC"
    "n1KpvYyHQYmHGWJCm6OhGs5aajNNl2c9+0znSehTRCQBFNMJ7Bq5YuqTRPG/gxQGkf7Od7+STTA26YNoEPGB8gIigCwfK/WPPg7yE7Phi1oStPtqkYVHMvNR"
    "FHvvTIl5P9DyrsipVSf8R2K09huvya7hDmO3ECADJfzz7YHzJQ2gYRl9nNtgOqsCALId/M94qH5Oleg2dYN6w8Qzm1IrsnbLr5+hozWTQZDLiXM7HBjM4NMD"
    "INbNxHwmMYBT9lVP9UE8te/JSP2fB/Z3m7RKitNRmZwTgdG5R+Ugcw4wFV9pCIBYBDeaq5/qJePXuv6GljuE6O2ni0MOuhzJnwEk4/cMhNvRaZXh+qLWSRpN"
    "CIKeZUUnr5XsAIlrAGCX1O4qM1z0j6yp+nCjJ8/VaDPePQE4IRO+nrbwePYdtVYzOsaNYdT7YouI6kj9b/wpVCFIVHsAxiaWQ9g4KlqNchAk8LGcVuZ7yr4m"
    "qDLRrXFfuq8KDFuHcG4PfpqCvWyGLsyS4Yu35lmG4bdvlmyVYYu3yfBFO+Q7lOEAgBEwPK3bvltjy7FVnXGs5F4Jwjhe6zkeYNURnHkPlR4sqoT9z3BPAiTB"
    "i2obc8hX+iOD+C2vfy1r97jW0QMnLObdyatP9+S8jAWnIQ512AO2709vYuwixR5DIxUgXR2l4ZWhkOYhoTKlo4faYT+jyT6+CJAGj3Tl+te7qB+369DrC7LC"
    "43dw2Fc93UGWa/2pU+SWIi5Wu+7UuF6COn1oE70V3PFCtyZiy7OP9EKAzDHzP0U25XTkqDOxwCeuirimNA5pr2xCu9hdtrN8Mcys2MFx2aybQRoJo826QxOB"
    "Mq2a+sh5phu1JHOFp0iYyJZjoidE8ArtpMX3C+wdmLiWgeiQ7iR46+vMHWkDK10twNktmNlWdDWUoC3Q85Ss2Fn6zVpnFOAgIiCon52v7jlocyR/JkDS2rwM"
    "LhRXPk5jjakUVPQCdx4lcXCXqQRAmr6U+tkLanSWKYttf52QTtXEI5vY3yN7eimm2qpMWhmXZMlEQvIJclXkSnje3YhIa2RM/8MbX8l6Zf/IZXDzMC4Otj1r"
    "0BmH5L0hw7ktfQl9VQIgcU4T3ZqVWac+dVzcIjimrz0SOma7ls9U6zmzbrsIj+dzF8W7PqS1uxMWVfrMEqyoqjFDoyXfIOi4D62ex9/w363eerRzcGa84KLL"
    "bX2zAyRtzB6T7fkm5yzfKBfV4aJMLwsYEDWZLd9rnF9CFwnjny7KjHD5Sk6o1Em6Tqb/Iw/VEem32Ws2wW8R3GMVA/eC3YVwHbw5qrebdBTkPHz5j8TxHE8r"
    "BMgcBGk57gdYAulc7PkQldPxDMb5AGTT7xb5k2zy8F/dnU2jHTyNPs5NW7NdzsFWAykwyFCvSU6Ofo7mamPhaNQPWXEdGjiDu97th6zLVNhwANuWqTzajsBG"
    "awRH952zSZGw/mUXd8xqmlobAAlweKHnLF/JOTncA1nHzo8HyARZlRKqBnjwE0ZcwJobQr8IkMF3LjeAVJ3VQDj5do0AkoaPKOkvLY6YYJU/AZ24D4oatwiA"
    "xpHTSGX6N1qrTUemnBMt1gDq8yDKjv7BrNWHuAMk+wm8YHBbVq7fxWVVkaAp3BrBepdX2W4/waxhXIxF+/Be/mr8kYl3SR28zrfCiDNx0X7dBBSKpzq2tOYy"
    "a9NOueRptA9O5ZZsxDloBatgxebiYX8rHUCbE2GYaz5irlcr6FljxhmvsmnsLE1gaEenUYuxPQPcl+gzirh9TY2nHCR1zAnLuWUCpwGHvqV95VI4lS9QfbIF"
    "TISdPDtN+EGKILjAuFw+ryArNvoR4vrDLabJgGnrZOCEZdIfz+g1cbX0mpQofSatlH4Tl8PLZJXm/+wxZbn0g0G156Tl0nXiShk0foFs3uJhwrn0USFA5hi1"
    "nbAPRpEyTT2axZXMBC0mx9WksN75OTjI3ADSFkcE4e+Hyt45ilrNkdwBuRtpzc3AZC2CicpJnEjTb8BoDt4Y0AiVOrNSc/ku2mOGQGbTbh1Ei9sb0kjUDc+y"
    "HIZxgMwJjpFbScUO8tv6AyFacRqbiSFh91C+N6m5nNxFBpD8twbS2Os+N2GTK4rYEHktG4v5vWUTsanTQtjnpdizZKYDGUGMHG5QAoxbvlnOqtxC0mjQoCqC"
    "3KPnomSEhQKkeiPwHXSh4qID41fZL6DctxA+ZWSU8SPUQRzGCd2KlJX2vWF4ESmjIY/aGrqOUN2RiAKJA2HQdR2u2Qqwl8enkcv6ww9dcrXNDrkqD4bkv4It"
    "aFeDm6Ll3rhkzcykekbzdzU3H343YNTziCO/tGZrmb3J7L8J+7wrMWLobJq7AJB2ZZ3WUyCiw+LMyCNNbGIuRRpi6omT1fvD49uVg0SEz93v9XaLMwEy4bta"
    "qzUt8FzkGDbIhCAFAaQxAfRCKFals5So0k6KoxTF96KISAuFv52AJCUnwCc2HTrLItgz54RKHaQY8m4WgUvUdTW/kIXQXxtbc/hRCJA5aNJr5kIYT5polpbI"
    "EspJRX85zfUXfPiyi9i5AyQ5Clo9jfTfLViHbVMZ3cHJipC+EDWh1mYThzRxgnKO9HsEgMLJ+KmOo6POM47UALLHzNVIKkp9Ka2xHFA+SF3vmN2a6FwlritS"
    "oaV0wa6FPA5ip0Ns2OnPtzijpLmdXAbU4adspvHf6s0wCcjthpRobG9uABl+p4EMCRuueKK7LPT9WswQAk4PgEFoYpo3GgrUihpoqKoPU4sk0rVhItdBJBI4"
    "7QxMmF8800mWb7VM8urSqZt3GRDwP8KGAuR+qh7MvUavTarN/1MXmbHI1dwGknjV890JVnAjogEMQGTZ5n3R1THFvncPjJDAAc77f39vAMTz+GLIceX6PDYh"
    "/hMx2cXhTbjk2pdAY/or0tMjbCinHKRZzZVrpHGTRiLQXIEP0VmNvp5hxHFnd/YhNtGUm18mR0ivBsbGG7AnVSh5UCLS7SxYSIPg0kbpj3MJNNAckcw8RQDm"
    "gkF9Z3e5B94blAzMch/MRYn++/8GINN/dMLc5AbtwAUr5KQKzVQcoEJfubgjBkhfj6CgV2soDuqJHmqEZ8GPUbNP685sllsw7DrHyczUTaZHxLuRUflX9TrI"
    "nJBxGdzMPgVbW9PLN4bYwu1YVSwnKLjONJsV0XWp+hsV8x3gV9hPNnigbQDIEKhoxoZIfkuOaAVcFZiRms3AQXL/aB3AnFhMDEEao251jPtNuPn4QIcR4Hpw"
    "u+t133Hz+5N9tAwfkDnYavenT+M+bh1KbkKdkZ1bVJC0BUe5KI0iGiJFkVyk6EPNsP/QbAM80HMvwRYRQEyerKCgGgYzq1GXzBC+tt/OkbewR/PriBx5q+8s"
    "ebfvDGxdO0Pexve3+s2WdxGn/mb/uXDnWShv9JuPPIcz5b1+U3F+mn7nNW8h9+Fb+j3v8iaue6M/rus/GWWaPuvNPgvk3a5jkI5vvfkroo6GVQclC9X8Y0O0"
    "kclEVB2EMYG2W/x+nAuzzO6WAR60wf47L/Q1GoQF0XwCzPMz8g8IIKnksLE8celmOZ2RXnBVYk5Ni21nAIJJAel4tznpkxNk1Bl/HwzOjRE75qrEYazZ6HCM"
    "X7QameTBPYITTcd8K8qEKkkCpGZ314XROGR6MGjoqOqfzdeVXh302qB/KRN4FGEeUNIB+7m/9pXp9S0e/PDlLwLILxA5UIwZxftlWTZx3T1wi+0oiIze3Jc6"
    "FTsbHq/9aPiedNShCHZSTOmzDdssbLU9u7VOWyW97Vx5K9py4diJhDPWZslPn2QmEHNgtkzZFB3Y4Ta46LdnfmV06uVK1FWaf2dGGhOzWMiJWVgYj87Qj5xU"
    "EX5wKu4Ei6LrCz30j+GLKm5rNmsMqIdbSGv49mn/6XP3RDqbueuhd8J+zZqtpA4cg1Wc4gAlMNDfzbcFUGA0fzgOoOIII2z3/WzTwdNqrYkOTP9EAdMMC3mp"
    "rL2J/hFhidfQ0CVWlA422fiGWk3HYmKSE6eeiQOWnEZwyXHjkrrlkK50bkfboIi/E9FNwQtSBUDfIqPluOVw7u6MiBuK10wPx1hyxrMbV6N7vlA3pnpHpIbD"
    "Z0aFnnJZ7U4ybYPvR02AVCstrc/UHxo2GrdkPOXo5VvlwiptJPWhllIU8dlF+E4kpC0Ob4cSENPoS1gE+SvTdK9lTD44t3OnwqLwrSyCrV0zIP6mw/0oHY7o"
    "6UzWi79D4d+h2HUQBcENZpRvj7a1xHXoY9Cs9P2NZOiMxUppLpIhoG46dNoXPQFOjklJsm31YPpXpS+5ShVHEUlEegDYSkDk7D11mYODzR8DRXPATzhQxS5x"
    "brPxN7MkFVmNGIVl2xtw7FnSZdvADHp1xry7rtm2Mugtv6nfR5bD/9dnSbRFxCdD8DzQlvVknoMMzQNJ8TzmzxqpCIJumUmaLTWd+Shb4mcm+GByY2ZDYklj"
    "0dycRh8yISWqd1CAPAGO+/3nWERaXggSAeSHSBab1mWZpHwFEBwMIBqwXVIG7sR3ANNX+OR3nhsEsDpehe/rxzqgfL0DdQh1Qh06z5dXsAueTejsE/fH/LVs"
    "XZZc+SSD7w1YVNRWMOTqbJPZ9mbhORpVuIVBV2n2na9Eii020Cwk75Bs2HVIbn8Fg4cGCia5VRCIO0QTcEP2bfzGa+B8fjciZraBXSCXE0AsxHC3HoHFDCGH"
    "wb8tXQ1JtBAT0LmaI00Xt0KgOKE5JSnC9Jarn+kimxDEbzhGfz+iQaBY0MQYl1rQkR0gA6iGVVhlMgVIPmvLvv1y71uoR3lOIog6yulRpPKJzb2PVdVA7sFo"
    "oxEfAJ+qSOZho1hlIZ28NKdUbTraLbcESCZKRQ5GBUjQISS1ZYZv0hscqurGyvSQuxsNzQa4lsYLfUXHUnwl+Og0dhI0GzoftIa+TbeesPrSl5O0pchu5133"
    "yTZwbBx14XPoosN3YBzyeRX6ybXP9ZXVWywAlFxtCAXtOBr+tEhmEZL5Mo5b/SAxDggKllMUW7qCruTs0rE9BwMOfvrsQFmemSPOPJ8OD6OBy4pG7CDBhek8"
    "yZmDFu5OZUDFd3HDMPYx2wEAe6SrlP9kgC6UoSvZn3RbKt9yknHAurgHFZbNCdOfurdItN+PbZrH7Es2/k2lkncJhkszIBWpgXmIBeWql4bIarhO+aDKdchH"
    "ANkZqapuajBOfvfOAvk9yh1vLpLb314st+H7re8sxvkl+HsRPo9nWSi3vbsAdViId6Neb82X299CXVC3W14aJ21GLTHbQsFzuaC5Hv2+butuuaEBiE2lPycp"
    "ViOLV3aAJFepW2j6OQ5G5P+jiGU9H+htyXh5fP7NQugSsdWBgoJ1uiaQUMOBhTGSe1Ruh4OAm6lXba0RMwpCqpAHl6NboXICH5KyFNd1i00q5KnroVgVJi/j"
    "jGOxveB8Ne4aBpKqYc9s5ZQsE072IwDdkRLVGh5AM6c2Z+nWnXJVHUxk5trUpKYm/poY6FyvLkRuUFDdEfWlHaTx6NjmbI6/S7bvk2vgDZCKgc6FgBxkmi5c"
    "5GhIB8t4HomZNALhXBE4hTcZuSzqK9bTNqGiItJOx7Rw2p66LUZ5dBXVIs6RMpxPreQushbot5ekTi2oQnQxdnUBFuCHPx0Wacg8NYjW7Zn2IxSsuBujpW6z"
    "3RINaDxZr4q6nvGd9cTe6X9BHtEAVklPDly4AslQfonEzqkwcuiWDpoMhCqjEBZr0hX3K1LDmxqGwN1DXdTCEztzyIUNxtaBo7zyGSwECFqIQkPVRcslCPZl"
    "CPkNewTpPutUfXEhYFsLoK16iIQQXjIPuB6c+d2fDPc1kIwM+//wMR8B5Mo2M2XgSS/LnBMby6ITPpFF+JxzUhOZduonMuPkz2XWyY1l9kmfH9cy66TPZMYp"
    "jWT6qY1kymmNtC7zTvxcFpzQVIac0FBmf2BJDEJ4/ZF0dF7XboI/yi3INGIJO31rThWJ2RlhNXNfL12tMUEgbr2MSIwADuYOYbNt4ba98is6hVNPppZv28a1"
    "GBKwakJT6iOpL9EtFtyP7KGO8gI2sefk1cx8zjUFMFsCrfZP6nbAoHIRQ/0FCbCcEBwwBCBfjRXYyUlCJAFIN5/g4WoOjIcDJGudHAdp6oSES5Bp7RJOQvHh"
    "NnDeZjmV6c8oDodwR4K71tfqaqoH32rCuYWSFVvKqBWblJaapEF1g4ek3+yNiBwB1wEjTuDwuZOiiZTuGB3oQK6UXDQSG5yLvIGTVtnzjK6M5DDH/gCQwVjD"
    "S9bu3ifXP4v3MIEJE4YAINPU0GFZlkzEJJgnMVELmsj6ewBGgiPHBxZjiOqffRt0ZaSq+cBu3LNPbnrRaeBclG6fEfJ86rPcABhxYxi7sBa//bU/L9lJ44v9"
    "0AUbpGQliMPKILDdCeNYiDO3EE/vB0pZVQbJicgpOuEH3xCMdPf5MWreBjmpGsYyxGqTygDmuMe4U6+/qrZY+EwWV28pQCYywic2XMu+T7uJ/lwwCOT0kwVN"
    "ocL45NugFmPEk82znEcEkGu/nCbzUl6UPSmfyu6Uj/BpZVvKu7Ij5UN8fiRbUXakfHzcyja8NyvlfdkelXdlL87tS2kkC1JekcUNv/f2HK5cTbbfc15HW8Dv"
    "34HYTIDULQAsdC8ApFm2KQKGvVywUkK/9ErPafoorQmdhX0AvNBrOvzF2qJzuIoRWD0xBUW1SLnsekIOuEp95VdP9JBFm8j6m8OyqjM5f70D+0xbK8VhbKBo"
    "R6OOhpcpBxmAgY7YQS9EPSqeD3eiC2u0lTHLNxg2eAxv7gCZLPWCCG30z857Zh9tz3WdpgYiS+NmFlZGX1BvZBZn0jiIVaQHBjJA6Y9vDpJNey2B66ED5KCN"
    "73m5z3SoIaB7CxFOeJ4mvPVFyPaBsYS4uiE8J1jF/oj++Fa2BetABJCx9Hf6dAKQvWf44i1SCjkp1fOAE1eNH+SM4klLAsAfAZeYJ1hSIiCNnNMG93V+jTYy"
    "GduB2MGBYAaFMdCNnkZwiVl9DSAT+UQtm08wenDsfi0nI7xwyHxPSpIL15Rr7zvdG/YcB70q6KF79BCgAngZUCqwsU910SDIgf6P9pEbX4ehTfdSdYnPHc7f"
    "7YPwQo4L0jMGkIz+sXdwsTd9MhdAS6yc2J/dJDu+K1Z0q49QDEQZlVYMwRE04PBdpSs1g7HJ5gL7WlP35XJEALm6+XSZndJANgMg16V/LFmpjQBMH8vGtPdw"
    "7kPZmIrvqR/h+/Erm/DODSiZKZ+g8PMD1OkjFHCTKS/LgrcCQObatqM6yaH3tw/p0EwXhiBik/g2GGzicdJRxIIimHo/bLEZANKchc0tZMq67XIBMlHrbm86"
    "IQis2EpB9Zph5eOgoo6MAIFNq5CNpcXwpXr/ARojaFmlMQX/hT58FWCcUo4AS5EKfpQeKWGcg00CzYqj27oy9hb1R3tue7mvrHIl+VER57CbEjpG1yz4BAa4"
    "0H+QgIZj/U5wOvVRL+5dQ2u1isDWbk3CocYwcgbMMBR2/wO9IQ436GmRElEqNRc0H/kS0UiPEvhcx0UdLsHLQxYNFKhqoFWX3BjoX7aXlP0sjJkwIcxIFfhg"
    "28+c1mwzJryNDOLcrzvaPkF1ppy8gdsjN0OAPBbgaGBmqhZa95kmrZemxNuuPpokBMeX0bXRkDmSrglA2EbzdzRjTPCDZN9bWKl5OIBWFfvKdS/2Q3ILM1Il"
    "77gE31BU4F/vs1/AtUKvqW48KrUE/1sYSOh36hFhFr1E/WMneaKt65FpBFJdMp4HUt/zKaLAyngqOB0DGLNYiMxYyYWeul5y6wyqYIYjSmB8DwHTFyY1luZV"
    "CNawYmOeEiBpqGPY7h0NByJrUwBsMiIhqj/7II8AckXrGTI55TXZkPK5rE9rBG7xE3CNjWRt+vtCoMrEdysATy8GWlbCOX7Gz+f3W0H3bMH7tgGwt6Y0BjB/"
    "Jpvw9xY8n+cnptSX+W8fe4AkF3QP9D3qtEpLMLf15P4gYbWMAPJb3fFOk0kgDPBVcIp66Gbru9SPrjpSZ9H3S/WO3EmQhh9uxqR7UVMchDtEJBbSINFV7npn"
    "uGzR8Y9p6wPJLME2QVi/8rQGP2riCAGaiW45gPgsAwgOKIIxi+v0IOLfhVjXXYe7eh0Ge0d2IhF4Z+2nsYO+g6CA57HqNGG5nMj9dpRjJrdr+9RYglVzBbFo"
    "F2tDGnWzyIp9JqyNI8JeP5HaAm43u/fLHW+BjpWDYt7BQS2nAagIkB6VEcQzLD71u0/z5iUA0nbxCQkmCJAW787p80iTkaZ/1E23CDisP2kaA8RY5NKPBUqK"
    "72r4QcYjioTFYRz6Ysg8rbMOC+Wira5lm3yv+kQDSKeh0jWIpmac0DDAYOgAWFX68nvXJhhQJXvMhUP1zxC7r3rk2hjDarEOAGl9oDk8VYx1fTDOFa3YUbrC"
    "qKjDQ3XpRvuFm7bLhU9Dl1mR9SXA8dPmmj3XtiA2tYONbbr02B7kdPrnAhiMnq6iCaqa+KcHXmRoiCvugUW9SpuxJpWoqskSjOR2JDjIpjNlBgByM8BoA7hH"
    "cpBZAKJ16R8o4G1RLo4A+dFxKwTQrQqKn8tGACTrtUU5yI/AQb4oC98anmzfHtF1D30OQnJXNXKM7iqQAMigc6LTsburYAC80sMnnm+i9PWCTGQKR7IJePWn"
    "UlnNDay4NzKjDmp+p5PMlOiWKy8F+8uUqtoZe8yY6GO7Lbph1V1leH7L3v3yp3fwjEqmHFeQVY7GV1OPs7W9gV0XRLAEt0WltB0Jfu+ICJPHxQG4DcHBedG5"
    "2ideJkJP/vYuDUqwAoe6RZEWrF88NA2DXze7AvggvPLfuG+bjluzLofdJxesz5LL6qB/eG2IHFHdo8cDBz2eWnMtR6FutwAOtgl2zAs0sClB2EF0jgOkWcqN"
    "e5y/ebtc9QRdqdxYQvDRBBukdeBcXT1yjDhIcqfUo6bWAAAhdd21yAC0JNOiXTR01f1qaSy57iW0i9srqDcFpRlPa6bqCi6grFuIpEGdkaG9OMTZxrDK6yjw"
    "DDfJjoHuU1ci+oQJR8LiS1pQnDf6BE8MAzsudpgrMHZeULezLFxjFvMoVRy+90MEUCp8IzXJinKArtNXIw/VJUzWQt29qxwgZSlXWoft4oLKRdBE5nyLPnsw"
    "uEjUG3roIkgl1xRbWHiFdAwcLFAH2XiGzIVebzvAKBOi9MY0FH6mvg+QMjF3E/7elPoBQNQKv4cSzuU8H7/+SO/ZhHezEJg3pn4KbhbgjHfuwrvnpzwvS974"
    "ztqYbA8neV3lFliZEZ2RQp2W75Fsne76Fho9KMpUNWtxCix0r/Ryp1vUJhOT7L4PwYXS54+rYB2IlvTDw7OKwQdNRWoFCzNKqHL7kTbyZNvx7rZk7kHEBa7v"
    "TLcb9gxekrUNWX0gkiBVlW2U7mK/5t1jHW2F5Qb36iBN/Rv1cOX7yb+yAWSyhpj8iMYZ61mLQkdoBhtkm3Eup92YBVKsfAvldC2rDLOm2/YOQU+UiN0FPThZ"
    "yF3Dv7DFCPP7I2DRZzNYXcf/gG0qHv0C7WI6Ol5PjpQGDTN0aWJcGGospRpFTtIChjHsSths6BJ/ZhAu+VQmOPPgRbbHRdgBM1bLSY8gPlgXH3JGYTJSfUFu"
    "hJ98dlzHFybrUYrcyimxHcx800WeaztK62uOXsEMJjIEC+nJ5eEVoNymWfHVgBRJONTBBSON++5in50zq7SVsasNrMjJJTt3eF19RuwwM7ovauavyzlhaoZg"
    "JLTEGMb10YB559tDkD0/MY4Cs/ZC51EWeso205uD3LOm9zMmRMFXxXQ+H7/h7zQwHSllIJU93FjSHvgY/pNNQCcEduRXHsbvyLJf9IHPJOXuT+WsCjA8b3RH"
    "KVepKHDnMtQjDnJ5mxkyMfUVgM/nKCZOU9QmSFLczoKouxnnNsdEaH4PJS5Wx8/Hr+c1R3LP5tQPZXPqewrQfP+6tI9RJ3KVH0Nf+qIsch1k8kJCcgjZcNAC"
    "jWJRgKTzKf3SVGwgl2KDgm4zGdwQiiAHHWODngEgRbpPXy6nwPKYUgWTSN0RCJCcuF8DILEpFi3Ywcmcz0Rc7M+QNn5RSFhLLowrGrO2YLJyI6oADtPXbZLS"
    "FZu60cBXUl1FzffNAJJuQ0HP6Y7ZMFD8/aNhbugJHGR2TjL6y8zw2YgVvyPBgZKzI0j65QrqzAhjduAFcMX57XOYANgELB2Dm1sekIM2dxzSkpwIwYeTyYA+"
    "jcADTu/KlwbICt0li4ClgdKRm8vA2Rvk9PIY9L5hPJPeGkDyGR4vj+dpMmBN7UawQhwuHLW/dIAMGXosJomOUx5rrL6QFjveACGeGfeB1tSdcstXppSjPplp"
    "6sDdMdWX7reDLRE08bF+x7jhNfA51e1v3RIbT+AQt9jbAsn6BUOPgxq41tKo87jl3LyWvrBMcWacLo93Bs2VdDiWW/Z4cs7U2/kWu6ojd5BS3aPp8FLgJH/r"
    "K/1lLRNzuOVemxvrQuvMmLnN0SwLp//UkFFLtrUC+9NCW1nfYGAznbDtykna4zxckF7qNcdqrVmIrP5b8fHH15iPlPp5jgXWl2on7t9uHgMmtQUpAIsq6Hrx"
    "0z3k3dEL5PMRs6TpiBnSfNQsaTYye2mKc6Hwty/xdxN8tsD1X46YKZ3HzZKdwa1HM/JDxaKudIfjQ8LNp+UMmQq93laIstTzEZw2pBIYP3HDDMHRAPJ4FnKt"
    "1DkSsDehTuRstwAs58CgtCDSQSa7DiYHkJ8MmCsnQsRWfR7ZfvXodzcfBUjrPF3teB6RE3WbDdeHb8a+Ibdj/xAORnPYZVSLR3TwXjiBU89krgfgQMGFFofL"
    "QZtRi6xyuoGRRZ6YaIlkF/g7RJJMAUCejKgO7tZn4ge5VA5Q1InJHdTFhZONdbSwLwUiTObfvTpA6MbE56orO8hmKeoT/ou62KhLhzl46/+8zs9EcTG+m576"
    "EGqaNOoczXrP3qD+7tmO4yTtYRg4PKyyKLgdtSIqMDILzFBdfHRC1IGqgFwCxWZYSd92vRufZ8EAfKqxIf3nb5STK8D3z116lANRbpGLAWiLvX2ozlADmmbU"
    "Zp+hL6BmeL67+asGnaM913R6+nT1DT2oWzzUbTJYrnuyk9wAsL7xxf5y40ssA+R6lBvwneV6/RwsNz4/BL65vfE79nh5cZD85qXRctmLo1BH9FEdazeNaQaG"
    "5vtq1urgzWCeBzouKBrD97Fqy9Gud+SCyfhzZMxGFVnu+gQSCvaCoSO5jlO3qkd7z6jYaWGm+i4uEgite6zDVNM/Ip6dxYxfBomJMcd37cb4sIWJF4xcu0dK"
    "P46FCNEnIZTPxp2rcoKDt7utqb4dzEUxuGn1neruNEhoaztvikxHDtILn6BqiWoncvxUE9DP1BKMmEdDzLWHYwQLwiNfjNb7f/zBgR2Z5vLkpCOAXNVihkxP"
    "eQng85nqHDelfQDu8RMFI1qxDRSPL0DSOMRCkDbu80MV87cGgHwnp0Xyx5ONT+g4ZpWcBh0Yk88aV2JZqHXwKRdpnKNa71Tp21lqNP5aX/7l8PlQrDNtk4l3"
    "tppadEDYC4YcqG7SRW4JhoO73hsCXRtHKdh+TARLoqWjWDsxIVgJUtXDn5CuHTAQKUACFMxFhgPMdEKmf6R+FO4NumUDfsOAvaR2B5myarM+mVyZFuceDAnN"
    "GGTQSGMAJxGNLvzNtHS+h5/NKmUYCShQAqg+y63upOG4FdgtD+5N3KhdRTKGgHnyCNWNQX8KHW4R6IRUjNKEAwSxDnJz/V6yKshkntTApASjyoB5WCQAkNoH"
    "ui0GwYAiGSNo0D/YCVANVgo+tH66NRVx4A98GgwUFl4ZccrqS8VNyUR2omThz3XbdskqJLRYsXUvElvs1U8re2RlVHbr9x/g77p42275YfseWYbvP6Ara3Wd"
    "Be4ZXCWNaQyJC2InVQjqZO6O8u4ArWobtgVbGJyD2OQRnulbQQop0DQPJGq8AsEM1zxPDta24zBViutus+lCOQbNkJXGxfqhttJoqKc38/637Q4MBAP4KgRj"
    "awXdSMun1OdD5wFg22LhgWM+54X6QXIecKwlwFg37qIUwwWLksCLvUAPGi75mr3uLyDSaewiZHKH/pG5JJn6T91yYLBjxI9a3S1izRZ6Lqho4yOt5bPvPJek"
    "rvNhvAWON79PXwkicSfOVIXfDsePQoDMBVO7ToIzLFOSaYSG6bUs3pMDzsRs84WjSEtlekep3fw7mbvtgPz6BYAWRDnVseFecg3me0XXG1q+MVixGmbQlQOD"
    "7NQqreWbRe6Thi1VbQtM149p0taw27VVdNqaDXJKRQAwo0Q02zgt7ZxwHKzB1YPiJq18AGL4SzKhqU5SpHFrFpLuEvBU/+JRJGpUMf7QANpFLQ2XZOFgpGsI"
    "p0xCX0e3E9uPms8xABvzQyYc2UEHbO+pG0fpBLJkCebOY24oRcHlcXMq222RID4AMcItpevY4MDrIIyqmUOGweQw0KtkBW5XQd2VgaMlaKBF32gdRT5hknH/"
    "G02eULkrokC6yVINseOzEStP/a7+RYDcBSzag5yPAIaIY81PgRO42gC2icE0Gc7oP6mNWOzyWBiqMklCcDmy1Hm2WZnRRblegLmGQpKDBhA0wNYafGrYSVLX"
    "SpcBh81YIqUqYoFQ1UHQNecCkCoBcYxSUukvp1XtIoPmeoJkzTLlsfd8E1UYaI6lecPLYC2ncGE9fkCqMOemp1SjeinBPXJx5t+u/1VGwPsckkCFL4b5dhQc"
    "HwbwbNfT7ejhgYUekkQiZNBSpZnKwKQvlTY8f2Xp6u1kZPBd1Mgn9p/Nj/xKgM0gBYVrExBpjEgYX3FIKATIXACy39zNiBZor5uhJwCSHKRb2jTJgvkhptSC"
    "qIMsyVXaTJEX+yJmF/uucFVXB24ApIKhrriYtBD9MjRM0XwgU8q2knpdxrloB24OeSMPYOAaX8MJzDXdBkBQIS/ctAVJR6ELozGDAxOWPXOgJbcaDECoJ/0q"
    "wV1xp8R0tZxDpIUO7d/v9tH09mGgqlVUQxijvex00AWuVXkLXWCdQ/QNnWw4kYWkk23CiDR2ZZb88jlwTUjckAbukDrcbE68qhNDnUA/E7kJjqQjClxayjQe"
    "rjpXcxciaKnqylUMNqSnrlgnZ5aHcl7FUQcZ0lzdppzD19h5bq9AgCQHT3VIVylSrpm0RZILPVx8N7drA0w5wH2/FSYA/J6YWBeJsJhYfTza3own1A+Shr7P"
    "8wqkbrq9IcYHklkURxRJMWTJTmN8PZNGsP+pFqE6QF2HkKAB9aN+VhcP6Guvfb47tuXgkoD9MCHqMjmFGVmt/U2+niupD2KBCGI0F6DcttcIuj26AUHF8tN6"
    "vWS6pnHimmdeEgb/JkqznyO+WgeAjb3pG3bIZdWxKDMhL8ecO1sHScr0kEH/69wj6F/i0abSddwye59LJvzOPKS3QhVhW45g4zBy12rooirKv6srDxc2t5JD"
    "t3trw/6IHgrsg+U6COk12H95lZB4I4z50MbE0se+jdgSGxt+FAJkNnLYH0MREnV6JbozoON10y6G6iX0euQA1LVGxQJMbIDhL14eKhc8zsFDLom6FIp1XF1p"
    "LAAQqEGHIgM7HoCFDr8G+zczbFCHKUTrg1i1CQ48Y8605No8ysNFgw0YIH/AJlsc8KrjROJdOlerXsg3pNLVl64QtZG9hedpbSeni4FYqmITGTxjqbZTgZi6"
    "LYAkB4u+V4GCbjqMXkms+AltTUJfpZZqXXlt4o5dgY2pMLlTkMGG24tqPTSZqinggzim1krNNEMaETjg01iph1z1ZAeZtdbdQYKhypjICBz4fXnWTrmBxh/u"
    "/Oex8UxJR65cuRvNC8j3E3wowlmGmRRu74qEIX96fziSgFhfk4tWPlCB2ACfnDDPxaeM1SHBqbDtxuFxP/UIG2UTvJ//8fFwTSabWoftpLsKJnlt1FcTiFi+"
    "RPYNN0+jD2M6VA3FuNth5Z4IyWsrXaat8rpB/8x9VtRAo7XVf5/qNMcyqMfE6ty3KHALMxbyVBg4rqzXTeZu0oR26hWg/ojaZi5ETmj8qqcJ+JBoeLzabw6S"
    "fIBhIKerEgDHddA/OgfpAKkuQFwUAfS/Qh+thjqAxwGoa/Z79NaY5TuxKRmMWdzZkXlPNSOPuexEGcjVUd6c3NO5lzm2Nane0iz6QarS5MVhMXUViYWd5Vas"
    "N3l5tsU/oqupGXIehQB5GElERoELOrMKXCiUuyEn4uIQuUBdLd31QPWTpiRP5V4rDP8jCPI63y5BdW8aGkVux0GVe8wgL2Tz78wfLaxehJqQo1o7ywFSXRB8"
    "72dGM5QFl6WO7ARIcF+66uogo/uMD1DV6TGRA6OByEESJBn21U7ueKWnLLKNpjFyIaY4QOocCbpE3TPYJlMQTXQI0TVExTNwujE3kWHz1iBTELgiZGWxyIrg"
    "EkPR2jiM4JenOjFyCL4LHrcqLVGmubQcYfTQlT5uVVRVA4wGPI+yHTL9Ax/QOIZJy8gL3TaW3Aet5BQrqXpgHQiQ5jSvahI1kuEcDCCfDgq0N8dl7nIQoF45"
    "ZgePsHBEy4CKdgBQ1eHx3sTEWr79gNz/Ifr+Iefu1BhFEZEGGXJftl9MlMhYRWAa1yCCk8tFyq+6nSb7BmCW+9NovRvtNjMdmdS/fzhSjYAFOaWrRV/DDkFv"
    "7D/0cwDkbGa7DYda32gAsh5W/bNiC88THPfLPDiHX/oYMwAFsCWoJRY7+84SrPFoP9U9ZVrJM13H+fjmUoukzB7P3QT6zHSOX1UL2XYmNkasWIYeo1265hnA"
    "ogKAbz5ujYOY7epo/KOfisAyt78pjXATNupCTW0V3Hrit+UCBYltXwuNNAnyzNiYKedUhUihewwzGa1FvBjQ+GRXtxJwBAoG5JDMtUKV74y80aS42FbB89FZ"
    "NnKKVuh8uIz8Ey43ysWo6Go6vaAkN9bf2X53uFZOxx2YX+k+1vy+NEsJ00t51ha1JBKg+W6CsYX0KWC7866CKnwMH/hsqKzBDob2Kg4aiHEECvypLg/KVZjB"
    "xmCS1+EcB9k+aKZ8O9PNONVo0DQ5vzY4btXbcsAP0mxFZrH2yaMAaS44Vj/QErRNx9amRcu0kCcR2UDwVzEJE5T4FO3pohthcS+/hLvTs20QDkoaqBuWGc0s"
    "cWuIdjEDmnH+5NqtEIy40dRFWAB7jV/inc4XE4xMjDbjFMMNTR9LPZzCioratuH8IUYLudM2fxq+ZIPcRM6+bDsYxrAgcgxo1JRl2CGXaEY+Ag2claF+KApf"
    "VjWykJMs31r+/s5g7JdufXEQ4r5uP6vITF2bcWJZSJ5xOVxdmNC5QIBUv0IDH2bfOROGn8ELN+lzlGNmVI6GhNI1C1um0qWM+K8v3auuOA9/Bl0huLdEYooA"
    "iP6ZLaLIxmNa1d5ICtJCRq/0+HEuKJajXRe5qs1Ga/ghc62accm4x0R72JcmHenWEXhe6eptZcIGhvDyCZQyCHbRCPF+zOODus+wC6iOdS/OMQbJ4HD+UQoB"
    "MjeSzs/cKpfVghIccauarRurvwEkBzwnvVnZbMLzGrcaM50WVmwzPHAFDPHGLpIQnJCP74xqzeVbH6imHHZJQT+DEODij1fQNukyLqLf9DUw7rR1J2G+B5Mg"
    "OJ9rolFyleQczIJsulLXEakuFNfAjelexMHOgQU2HBRZVPPl0pZNkyB6c1YlhE6eHz5vtdz3bj84gsNxl/5/ym1T7wdQ0PRt9GczMcq4DHJ4rAfFbtIS38s0"
    "lvvf6Q3dktFhnzuFsxKRNV+By9QAJvQfku4Tf9AM1TTsaGJUdfMx40DCeZ4LmomAGstLoGT9uHjBNeusqh2kyfeLIxcqowMbHyaRxb/bghXAkXDpCwu+cQvf"
    "d3uPlzNqgAasT11fVNVNhQtodzXU2V47HDOsIwGTxhtulQG6VGorV9RrI+NX0+eRVeCCCW8GXajYIVGcDyzm++TCegTI4DKE+/OM4iHXTKOJcZKMDqvXbaqp"
    "BPQ1aKd6MpAThidCZLQR2YT31mwPx3CEWaZpX+YHyAlDmYrfj7SSWnB780hnX3CMZj8gF+l1z4NRQCYqzYYUxPWQEUkBk4ua1b04vTDKdICnx1DZGo3UI/8S"
    "pA/mKTHjZKRtjVQnhQBZEF2dQqt27pVbXqGbgin86U5jAMn4U4IfxRt2onMuAAYGzhMYGaxvOQKpWyN4etJd5SBg0HmwrdRrM1wnHVdwN0zyL5+Y5FwSom0Q"
    "7fR6Twu/BEaAK2CNTQFXoO5HAbAVDMiV0TDhUQiaBw8rMa6hLie9KnWmJpamQa9Eq27rkfNlw46QgjV/Iq1F4oleszbIw/BHK4UUVgwJNNcOxpe7vjFYMb0+"
    "aiDSBYO0AUgCQFUnCmv/Vc91gF7M9I7KuarItw10YTZ2E4CCzshEX+VxZR7E2YsIFPR3JPhxUfLEwDRWWdYltp3naRQahFT+zOpCUddVIcglWRwO/eWajZXv"
    "F23UHRLzPgwmeRBgZm7YAmftefB5hOM/oqA0ozcBn+K0Wo7JsTIMNMSHc7wws7eJlFpX7JFDV5gLsV/44DlhW1vf81qd18nZobXk6vzdczfukwuexDMRFVMQ"
    "Bxky3eiOjqQHuNYLkNGp55TF+XbyvPVbpFzTcRaDDpplMPDAU6gdtkumu1jZXEAbEelyetU28GTI0ndwzKp3g2da+Xoe0tQ9CvWVZvd23aNKQjRY8dO9RZTb"
    "xnwirR7pLo+2my6T12xEFNBGGY0tYsfhc6KX8fjMXjZk+3vi6g0yeeVamYsQVW59olCtekrq+OPa3cPJUqiDjNPE50cW4p3vfAMdxP1TVEcSklbE9GnBRUNd"
    "f2xvmQxN3xXEBdOLcSJwIGRQrCzXF5tF9ZUlITO0irHaW1YizsXcbRQU/VdykAd85SMf+WSniRg4TKPGZLnubOzp1MhNFYFy2wwTrB/rQk6C3CWB2jhaFXex"
    "p80J2Ibhtvrd5MXOY6ULts0c9sNemQFd1fys/TJ53X4ZsnCntMImX093miC3vYLcftBj0oXH0lpRDCJ3RDGeSRMIfnynWZcTE8qTDKiPJjm4LtjaooN8HyaS"
    "Tib/R41ELtg7VxN4NoqBFH8JUpXbgAbIpq3Wa/UwQBgjDRK6/4iJsiGHIH1Q6WOo6cqYyRxARg5F9czQh5WCyH3vh4PknYHTpd+8bTIR7Z+RuV9mehm7er/0"
    "nL5J3hgwQx75cpRcyCxNpAOMS5q9nKnn6NOpnCLap3q0UJy71Z0qLTmHghfcjs6q0lI6eiIHMz6Ye41uHabO+GagCdi9FD6W5z9OA5V7UWhKNw8pVXcb32pC"
    "E1awD3iO7+MCBVcveDJcXL0Z9IPjpe+cTTIR7ZsMXc+Itfulx/T1GAMjkYgYgA+uTQ0uCvR0yyIAUp1jTufahlgeTNJf5wC2gn283XDTpWPwqraQfrI+txoi"
    "sTS3luDCRu45Mi4Ff8q4P2eUJQnbj0AVUxK7e56GHSxPwcZ3p1VuLiXx/XSkLTu9clN8bymnwWWuJBYbfi8FV7hS+I3ldN533zvy1MfdE3tQmiUKs8x0rz4R"
    "D0PIQoDMZS3dC+I9+AkGHpPcKuhwrwtMAHV0dmV1nmKNizw6iDjA6KtHK3Y/xPW2i+KLVf2oCv/8uJZQOdM/mtXZVrzvkG38DGxzSV2UKrjV4EHu0LaDUIOF"
    "ithuUFIdnEVwRDn8QvJRAgrD5R5sLSfBMnw2VAqXPdYFvoztsTl8VzkTYmIJitAPtjfjgMZ8kzszvatFbJgfXJTjUlPtm7jLupALSVUABecIQ87PH28nI7D5"
    "Ew9yjiGELj/WJuhpg1pi0Ny1cnrFLwFK4KaxS2QqxHZarFVvpdvIBrHQORNVARC0KJ66+B8mObNj0zIM/WxpcJmX1O0qP0WShZ/W6SiXP9ZZLoQFvGRVjAda"
    "cwke3FhKgwFsIbSMRLYohL1hqI+2JK4G1lGSC+pBwTmegc3hWo+cY01W67mHO2JYJBy8LJJKLez4fRO2rvjlSwBHpjnTxYjGNwNI6wOqOKjaCW5fHI9cKNyL"
    "ggsXt11A0omS1TvKxQBbuo2dCyv7KTUwBrDoGjfsNAv5Atxjg9KRhnCquoDjgAsNOXQ8FztWXlevI/xMTVVAnfL+YNQiYKL88yOMG8Zfu4tSgVxwmGdcyNhH"
    "OQvHrhbS3sY/XeDURY9x6rolBn7Dgtx00EytlzIfjomeLtkB8vC5WAiQucxIQlCZzyEOMwZX0ykZSEZ59nSlLrjYjoWmk2Iqs/s/HCrbIGqoewhXLfc/zA8U"
    "7Dc62ZpPnoqb6EeKCpWaYHKUwYBm6qnAKXhdE/tzGKfIqA3GgBerBn1qPLmCr9yWKABiOUSktMrkUAAEmiiW3zmgOWlycEVKG+OIqFog16rRIRTHfMMly1ZE"
    "KzKV7gBIKPyveay9jF2yWduyHe3Y5zrGgulAUph7CvFiOx7wcCM8kxtbMZckvA24gVfYmtcc6MMWBF53jVoxdxVy/CyRDymBXjPQIAMT91HG9rBpyPWZURl6"
    "ONAktQr1iVSZ2HN5nz5DVSoOypppBtwqJmUx1KUo83WqegbRPRr9hMlbGWALLqfFiHlusTdHe3PQNpcbXQzdf8AMZ6ZFI+g82ny80tEWRvQL3ZpAY9bFXKdY"
    "B1rvg89tHCRNv5eB9qdTRQODYQZKeiVs34F6mRM4AZXFONAodykkFbrd0PHeQiVNMqGkkoZ7T4alvN0kc1HS2qu+jzKt8f+LNu+QK59FfRmjnsT8SeoabSsX"
    "/bAI2MIc3es+oqfW6CwTVyQyCpnQpjtHReqL3MZfIUDmQhXSrkYb6GA0cwm5BHM50EgHzS1YMDiG1E90Z0mBMr40chuOXmhRDJwIFJ0OUN+WxGECODkIcBNu"
    "WeZtk5ZvkouhU0qpbGGMljg3DHAXexlxo8lGmSSjBxyS3eCk1l13CXL/TONC3R2pDj7rsu10jXF1Qdi4LBZNxJ3llFvx+OKwT4vqrDz7uqV4A5BBX3nD811k"
    "9OKN2mr1JSQtQAf6YxZ8kCWhbyDWfddpfYeww9MBYEzakIqkIOmYxCpKg5Ok8Ya6SQPBwOG6ZZ36LUx0jbLhQqaLGScZ3XLIGdniplbWaGFwmmlCBgNJJnAt"
    "Wg2JWGsCsDTJAqWGkJnGFgrT+ZpLC/WuF1VtJu3HzFP44wiA3zSMU8R+GodoBPJJ6wsHr2GaCstTKfLptwhIeLi1AlPR6nwvQQJRWjo+wVnSxYuivHoUhGxP"
    "DiTuQG/hh677U9HZOF9L6uz6xJhkQDVGEXKY6oZjC2JGEL1Bw4wHm0l9bIOrmmyOa6a88z2Ugk2r5zTsQsm0d1EWpILnUYEgGcBROXgT2SOxXduK9kANch2S"
    "Ra9xQ6DuFaUcJOnsUWF5DL5CgMyDME92gQUPOwvaBAn+Ws4pJA2QHAAQKbDaP9V5or7JXEnol8WwtuAEnj80ECA13Rl0bxTDLH7WxIFX+05HajAkbqABhhY/"
    "FavpMmGD3bIx83uCK4gGksaJcxLZAmD7fxuHZU7t5A7NCm4TyH9T4xMdoDFpaPTQiUZDDI0QPnmYxFdDxcCJM+rn4XbyD+zfvAhJUnUOOYeh3AUm00H4qRV8"
    "kBuBPx3oto9cpINk/U5YzBCVFIBNNzLT3fUYz5uIezYLOgHSdaQ5tguwNlN/aO1V/S7bTW5YLfHU35r7VrCOq2GG8e8Q2YuizYwOsggn+jfSaMV7SHvQDGLt"
    "z6F3HTx9mY0F9KVmalL7XOhfC5/T3vWMTuQg0WoHz0MIN90qZzFPJTjdExG3nUG9J53t1UHe04Upd89xS2CM+xjyuy8Evle4uj6xaBtNb5xdBLacparrVpUE"
    "Fwhy0ngW/CM5/mq0+l6ywriEC9RBuIHRV1T9RN3CVL/PXM9GTrocA3BUPa+rDrgoap3CghDUH6ANYsgf6zTJne1ZnwCQTCCXv3KnECBzzko3Kz/XbTL0MUhY"
    "oaBg+pbIAJNrYoDDO1x1UrCEX/NsT1nqEQVujkEn0XE2kcIrP3DgPRrbob55JnwF8/d6/Hn3x5jQSJBhOkeKXa4fUsfx4HiLunDC0nKqHGCIMSeA0CWFYkkA"
    "Dm4ixsSi8GeE4SHsMayic3Do9bRvylGTPtQBqmuLx5+r3g2DF+4cxcq3lSfajJTN8OEzYFCtmmG8istU5JtutaClgp6ZJhhFJJBMRBf9rSHe9whAg2J8sIY6"
    "4NOKy4lj0UZcPIJOkoDC/g2TiRwi22BFRWm6uKg+z35L6N7YZt5PehNYmZmIxrFuKAmjkf6GpMkp5VrIvz4aKBPXbjfBWS32nKykQ/BcMH4mmrIOkLyesGm+"
    "hJYVskIbitnYc5v7P9N7gPpA1bG6qkT7JWZIUc6PIBfjKJVT9kVBd0W0cUAx25L2xrhqXxQsJybVJQhWqAq1EzaCq438qVnBzK4x7RbuYIYmW8ipGrj1VdSJ"
    "4YX6zmMFkLZwBTVS5Pep4r+N6XQY8pp/P8uGFseZ0pxUNv2u8xq5Dr1CgMwDIF/rDQspUnVFhgcHSMtOHdNx5NnRFpNaDPnwWrmVMtjLrE/ys51lr5QBJA8O"
    "PnOOVWxxA8/8LTvl5gYAB0wYAwdyg4xosHT7GheuFm1yNpzYnDwEsACSBhxWXLnvk8QmSoizNQ46EmO4UGiUCEV7S16rui8CMfVyDzbB4tBZ2k5Y5m7ObDhU"
    "C1EmawOD4MpUEDxyDkb8lc07owmO2QCe658llwZ9qVo/qRpgewlgJi7TUEaRO/IZVUs+JQQv+p1cibdTDR+W+t+MXebDGLwAzBOAY4E09+gpzRUKfSwjY2gs"
    "QGjjWdVbS/2uEyVL44ihP2Y7fHGwLRQY1hn8PGPz1Scy221hoHYtj3Grtst5NaFaYPQWgVGTqpCzY+gp04e5btJdjiyRCRcJxr+T4w07/4X+NG8LDY3UqKyw"
    "sJJDo4eC6U/N4IRzVeBrC6PW423HyHbdBI2O9sFPlm2xnAJh2Zu1fodc8ATEfzh9Z3cK/5FAqVyvRUqpC1kwogbjGzjsi8Ftz8D+UHqo0YjDxpIIKuMRjaXD"
    "R2AhQB5GE6PWR1/NhoNqZ9sfxLOiaPA8AUN1a0l0LJIV3PN2P2FwBA92hvlgmRLeo4AP75UcZ3QuWe/6BLHJZLHDNmEYZXBdA/qMtdaBktglkFvDdtei6cCc"
    "80tkm84Oktyfm8X8BwkuND64AYIhfBFAcgEgLaC0Z2y6+heSYwE4wtJ9UgU4C7f4BqFqPjA5vT1jSiJ8O7GKJ1yv8yYH22wpHHygh2Az51Imr9wqv2YseJnW"
    "4HwttE+3YahrBgfdGY8AqX1n/qk0plhKOj/v2YaMTtQpstBtKLgOeaig+1paYmKCKnS7iKwyCzqBEQYdiJ5/fLmffDvPM+ios7ml99cU/6oy8cQgvmgmICW0"
    "0XN3au8HycHG6Cu9poEz/QJ1w7sB4BlVzQfWtqJlvLdLP+ppYGAexFD1cjisBOD0iKdglWfuU/pCYmvWImQQ4KZTCoEKbw+erTEtelBFgsXbJCT+R5elkC1I"
    "pPWouVJMd0MMrnBJzJ9k5phytgRuGstooCOAkwlw/St0vn96a6DAbdZmUDbn+0KALBB8sl9gHA6PTpPXS3FY29LokBsDyJAUIn+ABFBg5TqraisZvmiDdYxm"
    "T6EmiVyAhdIlomaOpJp2F/ubmX90gyyK3jimb94lD3w4WEogtTx3u0vTTY08Dlkngw+gyDeOEzt7Ue5QDQpBZxnXO8Z1sC6+6N4j+E73C/gTFi/XUv769iDp"
    "NX1VlEswJOIgo6G7x+nqnUi1ZTq3SEbLnxgqW9viEoRR1UU6Nz0R26Pe0QC6UTjkpyFhgqZbU+OTZfaxXJlUMXgIXkyxn/DJi9MkzmEa95jdWdos4mlYTDTH"
    "J7OKw3H8knpd5N3BM2XzLjfEoYr7PMWY6l3VMGW7KSoHrchilurA0QT+OtAmUs84z70BoaIPfYQ6cUsBBWrb3yfB6Xo7ApcbGWICYIbkEIGbdOlIATWoG2wh"
    "1D5mPDbizK9/vr0MmBH2Vyczxlh2+uma+G9LmInYoc51209QiczA7Fhasa1/VH2gYb+ulwTnrEEbyK71bO/gSuUbhkUGmn2277ytuF6yD78EB4ltX6elvKAJ"
    "crlZ12bdj+Yz3eog51YJxyujOJPlbsRGXWHbV0vcy726G2H/nPqy8JhnFA8cmkjXqci7iGSv6ZG/Gx2iuUp5bkhflc0h2MUs9wlUVwlYGet1HGNhbOwDBUiu"
    "t+QaOIbCWpssOBrXEIGjzicXuWHwCXHBmQimfb3nVDmbUS7cnlMNJcGwYINedY5hU6+gr4tEMV5j4YpWuBLznHFdlmafXJIr++l4XrYH9kruhEzqA6Td6MWS"
    "5Wm/WF/Ne6CxsBTBrNkhR5Clp+JBgKf9Mx9ZJzTepTiLp7GDSXtp+NKUYziWbtsnVZpPlpMR402rsXkiWHvUchvi1F0XGy0S6gRt2eNTmQFIC/qdn7prJH9z"
    "jjEYesgtEmRh0Et5qA2ie3rJk12nyqwNHjZIYZN79KDhrK8uEASSECSgTQ696ouEk4G8pWmpXV+rtFOq4jmk1wFZh20tHvwAcenYIjeFiWY1SQf7KRia8N29"
    "CcJ2xSYyxw1NMSDVCDA32imgon0Ma8TWrKWQT7J2+3GyaKvzjWoNtnBMqgyoV40DZFgCN2F71VteQT2w3Sr3mKcKICkJLCkOkou5+dva2LS+1ixBkP5ORA7U"
    "kB1Jwyp13tg/ZFaSBsjVXzCj+PO6addOZBXPwuemlGYo3Hb18wg4CZjHq3Cr1w0pTQCQjf2djVC3T7Fnzme6wdjChsN9ihyrDyKZiQmDlmyXc2vCH04zYtMh"
    "NzhFEwypbyJ3Zp1j1lMqyhHJwUkGd5ZrYZhZ5nHO5rfHYroaPTyM7mhqnn2t46zx5zKRhE+pr+FK84+PhsLBG9Zd7utMx1m1rGLw03jDekbZZtgeT2gRDFBR"
    "DDfbxkJLLqMfIH7RYZqJKWAdPLVKJ31Pa2Ti2ayJHP3I9tX/yIZ/8T8CQBRAjVjDD1/vs5+hs0zPqSvUMJDG/YHo9K8xxRSrIY4yTpoqAoiNjOhQdywFPxoq"
    "GAZngKgZktxgYft4O2dNtyDuQ4M9VUrAh/DO+t3l/T7jZX5mzpDNUC8HukCew9aC3Frk5w77iX3ORdEs/yt3HZInOkyAuxPEbURFpcDtKFjQVXWgBgu3wuu4"
    "dSd/jYNOgKPqMHkddadsJ/2AwfWdCr1hmU9HyriFa6JFScdvlGTWkZv1BPBbJHlYCJkdayuc0tEHoF8G97yOrM/u+uReIok9aI7gvM47s77bJnYEYLSrDn2D"
    "ByCUtqssd8+JbAuwz78CluREsooVn8yV0Sl1ZAU2w1oFoPwBZSm2YFiS8jTOvSTLUZZhH5gfsDXs8S2vyGq8ezXqsxzvX53yhqxMeUuGo66zG357NPiSzz1k"
    "cWzQfbd8h5xfEQOOiUnp+Q9Hb3VI5j4gLIgG0a1MyyEsC1balPLgIKCQ5yZFRR76QhpNcIdZz4GX1767x7IBahVXpb+B8E4A84Cpy+XhD/vKucjGrBmcuekS"
    "HJ41mkIHKrkFd18hx6jZcbj7IpJQUKGuBRwECxyLuc/3aUjy8Ov6/eS5njNkxOLNvjUr38g0YCY2BnV9QUzhsWx/9KyQbQMnMnftl7ZIK/eXt7+CSwytyegn"
    "toPf2aYAIDrBwkJIKyuBkBwUuTFGZIBW0PGlVMTkg3/sCY+2lhsw+ap/Nlh6TV0jy3ebUUIP39lRvypQhG9JqhGSIIoaQDhcfVMZPnnk/NVSGfU5lxFWyJaT"
    "UgEcpfY1o5+oVjDOShd4RpyoIzkXATr3s3/ptoN2lyMwtpSfP9FBHvvyazx3g6B50aGqb5eA4i00RpiitW3fEMDnw4FTJeX+Jngu6oN9gdSQht0OE4V/hxI/"
    "z+95/ebP4Lxj6CIioFRaQL+oyxfn4v3N5G/vD7IEzEd5RCL2toVtZUWzv0hmy9/J5la/lc2tb5LNLW/E9+sksznOtbhdNrW8FefwWwsv/B5KOMfP+Pn49Tl/"
    "K+CezJa34N03SlbzGySrxc2yqfWtsqnVrfj+W1n55Z8la2pbdMb2ZLVXSZAIw0yThB5AuNRe+QBZWl5CzGqD7pNRJsrLKA16TJL6PSbj+2R5rdt4aYicd691"
    "m4C/p6JMlwYdxkjjAZNlQ1AluQHBOMgf0VNJ1F6z8TB2WZPNkpOxKUtuasyidfJR/1ly/wdfy2VPdJEzwB0Xp9L8AQymhyiOY5BxIeDgJZDCbzEVYHIygOS8"
    "2l3lqnrtEV30jbzdf7oMmrNaVjBLVgQI/ANiF1KSUXBJZOFBhXyj+ySqfwwvoeqBHZCY1Vno2sHz1ssL2J73jjeQtOHxblKC4qhOLixw/CSnzU/uYMjvj4A+"
    "cPUqUqGrnIFrL0PI4Z2v9ZJXoMLoN3WpLM3iZgR2aJvZvXBkF2SGD+Jcdr7x2DVRDT26+DLoINFO1mMspId3+s2UP77eT/ekLs2+Zu5MhJKmPIy2sq+pfmEy"
    "CuoFEV54EqKlzkO44a/Qz1WbjZBmkAhmrd+eWN84NVAS20cru+jF8zKyeRzn9hFxmIPHzpQX2o2W53rMw9axM+VVuNC9gsLP8J1/x88V9Fu4nvPyVcy/17tO"
    "1m1pG/QYJw27jZWXMD+f6TJRBk1ebAvWUc69CCD3yTPQpV8OChTDw1CkKB6chu+poMxJ+DwR54rY38erCN4lKfa+g6xTCRTUied2nY1R+Ro6Y8MxBEgujZZM"
    "M+RmzMs7Lzeoyy5BMhOLxdYGYPyfBkjqt/apnoVtoN7K02XF5iUn0Lzt++XbJTthWVwsDXuMkRe6jJU6bcdK7XbjpRZK3c4TwB2Olw+GzJZuU9fJuJW7ZO0O"
    "WPyyNdCCtBjZw4gg25qV73XgVDaDWxjAiu2uKccOHgp4kio7zUCgjtjB+OG3bcEsn77xgPSat0Xe/2q6PN9ltNRtM0bqth0nNCbUbTcB9BgnT3QcCzCcKK3H"
    "LJehi7bKwsw9yOGZfUSoblGlTb6T4IiFQhcL03cFXekxXxrVJ5Y5Mi1vZEhTF58MWaDBtM37ZeDcTGny7SwAySh5BlvZsp2126KdHSfI0z0nyJsDpkinKWtk"
    "9IodsnrvPnfJMmKpMcxxMNoiVsd1gH4zlwW4jAxOeg8t22Q4jh3nnF/PB8klXBPRfJ+pno7mSADkrjdl/4YL5dC2FDBRKXJwG0BpK4AIf8uWDDm4PR1Jhv1v"
    "nkM5tD1R9LpczvOavH4r6B7ee3AH6sFnbE2VQ1uL4J1p+t49m0vJ3p1voM1bj60Up6BiyQES4lFCCcRzwb/L18qI7iphhGHFDYVC8gG/4n8aICnUWFJdOuia"
    "iKM4xe+Jxhz1YNFm6KgjEDLj9X7f1MlPJxgKtwnSqplkUtOjGb153KOLkrY3OKEzrRXpQk4iQFYyjkV5V8oy7hhtlZHTpL5MtAvuUR2/E2najm5q5kcQjk86"
    "OdNNhbtl8183eWkarzBCFa6OmLK8y8IbY05H+hibB+bY78UgNDIzRVm+lf40M9me44lJcnR1yqcnrF5cCPGNfrLaM/a/u9UdfQ9EALl/1+tycMNZAKQU2bcz"
    "TQ5uTVdQEgDlIXwSHANw8m8Wgmgo4VzO8/w9r9/yvUfvA8eaVRwATZBGyWKd0pAuMEV2ZJ0ju3e+CRIgd+ARD4F8bggDLNu+FpxgxpHY0OOA8EGoopxxa1aR"
    "xFVHMziPRVPiq3lY5xNDOBtS5vK6sBgEtLNW2PQLrjXm3s2/LOsMD9LAuVanE9dt2/7q+B7smcixmqCliWFDO/hpU8m2N/W+y6uKgRzxT+/lQAVz2w9UdxBx"
    "ukRt/xFGuZxV4xtCImMjPUHRZB6CpplJclTYa5j9WbF6q0WadAoAG1yvnHA6tsOYsM8wIrKTyLXPvoAQHgPQGoeXCAzIjbRHek5npLq7cQqyRr4YauhmosZH"
    "OwKzAeSBjWcqB3lgG8TZTABTZoZxbVtQCIrkGsklkrNEORQr4VzO87wmr9/yvUffA25R3w+gJjhmQuzPgogNsNy/uaTs2/EKKJB1TGdggjv0EaUTyPVZATRj"
    "upcIGGPcUzQ2j7ZXjvY+k+cjMdd5Jh3KBhkQdzTXYmhb/EUB/GwIBzoYnPIGjsCE/2LEFyp7yqt5JhRyDom9bI7pApYEbQy4jXcxfppkYT/qLAqtcQgLfRtf"
    "4BJLTPDXTNDBaJwoRsvs0JEgsVY30DuJuid7idaH/4TVwJqm7VUlQB7Abs+3VoU9zhXSouu9r6MRQOAk0gT65AeOpENC62qcpVdQt7HgdhVQPyiYx2gYr2w2"
    "2uZoRPy3cI/mNUCdlDF13pd1dZwOfZMsXXNelxCxd9eX/ZvOcOAD55YJUEI5sC0dIAVwynIuLgfgxUHumH4HCB4CF7lvRzoKRX4AI8HR67E78yxwkG+hPcdW"
    "xLZODRPEuaKoMwJYhtAwG2ZWTCeX4LOOtkt+xH3KBVDU81U02srSasVaW81tPCU4OwcPchC6TwmzQPNKjjKHt4ByzkloTLTGtXJmmjBjkcQc/IE7C78fZ4j0"
    "7gv60MD1W/+E3rVlQBPUar3Db6FHjVoRlXSuOqcWYqc17NMgKcBKHB5+PP9SwFgIMqS+lChtdA6wEv856oGwUGhfm944jAe7kcAVzsT6PgfI58TfRE2zLSVe"
    "G0NyC5VkCRr+Y/EZ5Dn2oO/b7thrMzI4rx/dvIoAcs+e+rJr87kKhoe2ZigQHQJI7d0JkITeTzLJwbnYTdH7eBSCJAD5AHSO+6GLVJCkqI1zuzafJ7t2vo1W"
    "Z4YpfHQUOOyu+PCy6RSUvzpcgijinEgYhDq9gojpHXSMKpT0Y4LIp3kFo2Irf9BFRrrJ2EQKi4Jdk0jOakH99voA/pEorcAYtFWep1KfGRaXAEsxkE26JT/y"
    "wlAF78rDmRJb6OLtjlAluth5wsgYYUKlTfOQh9q47uz8Y6z9kb7Tudgf2azE7exQ623WJcC0LngRpoXlICcz6cJ31NfWu9avPlYUzpwpCz/HADK00KAweqWN"
    "OV91A1WMfQ+0S9AhrgKwfghhA3F6hvNhWUvMzYSFwL6ZcxElJHKptlZEHplHacFm02MA+Yrs3Hw+AIliLYtxjAcASge2AxC3kIsEcB4PYAzvCLpHiNiHUA8F"
    "SBfx92w+Q/buoJFm0zEbdrk96PBV2NEmWpJz3BWfnP+jNTv84Qmwjg/a2CiOj+gczEF8WKq0kqMd2aZb9Bu/mNI9mijZJlLgy44dIYK7VM7P8Iawj43WSRcJ"
    "zlub+AnNSFhAYrQxzLF2+z22N7lRJmxoFVQPVHFFdNJ77H2JYAB6MSR011qF2HMt+idvumjkkT7TLjq83WE7BgNIS7hrbbWUekHGTLxDSYDrGNkUeETNKkRw"
    "ibJE0Z0nLijnREF7Xujm8MyI7k7rqO4BIPWdMXBmPX2c5bpoE1S924J6P3xGWK7dSBo58FLqiUUoRYv1sQDIvbtfld3gIAmK1DMeVOs1RdqgewxcY3ad4jEV"
    "q/MS313vaFZvs67vzSwJ7rYhSHysOchjN5kLn3TsKRAmHgEkXgKQxIHlAMM7/bqcHgSq0NffEigVv5bfeT+vOwAwC9/D8yMdpIOXpm2j94O7F9mi4QAXB108"
    "6yA2n2ZRIMvlsHda3cNnHCBDHXLWl65GWrwuuQFwot0GwBoCGwPj8L5kei4O4Dn7gs+xRcMK6azXOE2Vq9S/rR65HSaVJWih1wZVArlE/9t2/LQFyOifz8qT"
    "TMNi10QcZCFAHiHlCi//j1DgSFyl8gLHI614AMc4SIVn5Fafw94b0w/qRA5cVgyc43UKIJYbqMdB6UjbkfN6czQ3UDnaI17XI61bXu2M1yXZusUXlKNtS273"
    "FQLksaRm4bOOCwVGjRolXbp0kU6dOknLli2lVatW0rp1a2nevLkMHDhQlixZEtUjN1ALP06cOFG6du2qz+EzQuFzevXqJbNnz5YdO+jobCJubkDJ8zt37pR+"
    "fftJl06dZcqUKdH1CTHTOJqR34+QcWPG2uvJWeUDkKGOmzdvliFDhsju3RbjHUBzw4YN2l7WOdCgVYuW0rpVa+netZv07tlLuoJG0W/evhYtWkizZs1kxYoV"
    "2cR2tm3kyJFKj/bt22ejR5w24XvTpk1l9OjR2VQAAcwWLlwo3bp10z5iv8TpSlpPmjRJNm1KqMZyA0GlXWw0sS9Yr4avN5RXGrwiTz7+hLz/zrsyetRopT+P"
    "I+F+kx2ohQCZLKUKr/tfQYFdu3bJ3/72NznxxBPlN7/5jdSuXVuqVKki1atXl3LlysnPfvYzufDCC6VSpUoyZ46lucqNy9uPiJh///vfcuqpp8pNN90kNWvW"
    "lGrVqulzKleuLLfccoucc845cuedd8qgQYOyPScu5vMHTt6LLrhQTj35FGkJkAqHXhe4M8z2B+67X35+xc9k0cJFDpK5kzTOiREcH3rooQgEwh1z587VNlat"
    "WlXrzVKjWnWpWrmKXH3lVVLqtNPl2l9dq/QJv/OT1/O+RYusDioK4yBgsc0lS5aUO+64I6JH/N74d9K6b9++MRE6ISq/8MILcsopp8iVV16ptCRNw71//vOf"
    "5dxzz9W+I8CzH/LqI56fM3eOPuOaa66RJ598UgF38ODB8vZbb8sdv7tdSpcqreNhxowZ+T7naAdvIUAeLeUK7/uPUGDp0qVy7bXXysknnywfffRRtjpwspM7"
    "4YQpWrSoTsIAkjkru3btWrnqqqvkpJNOkvfff/+wtqxatUqeeeYZfQ8nNCdlmMg5AXfQwEFy9plnyZmlz5CRI0bkSpctmVny1z//RYoVKSplHn5E9u21HJ5x"
    "Djfnd/7+zjvvSL169bI9Mz81w6oVK+X6634jJ51wonz4/gdJ9xEB5vzzz1dg69y5c9L35SZW/+tf/5IiRYpI+fLlD3tOZmamfPHFF7r4cJH7+OOPsy1g8eeN"
    "HzceYH+lnItrv/nmm8OetXXrVrnnnnskNTVV/vnPfwr/PtZHIUAea4oWPu9/lAJDhw6VM844Qy644AL59ttvc30XQZEcB0GSIBcAJT75CKQ/+clPFAD79OmT"
    "63O2b98uf/3rX+WEE05QbnOvg1q4ODz3k08+kZKnniY3Xn+DEFgD8MU/p06ZKuecdbYCaamSp8vnn32eL0CGdxBk+HweyegMx0DkJFBfcN758vVXXyfdF+SS"
    "zzvvPKXt8OHDk74v54VUCfzud79TgKxfv36ezyE3SIBkP1Hcjy8W/L5lyxb5y51/lqLpReSD9xILWE7uvUePHgq2pUuX/lH1zquihQB51EOh8Mb/BAWozyLX"
    "d/nll0e6xpwuMazXH/7wBwU/cjNBRxUHSOrCSpUqpVzkrFm2oVNu+koCLIGW4ufGjRujJsf1ZhUAYuQMKULv2WPp8nJyVj2791DO7tcQe88640y57JJLZQI4"
    "pPi1OdvBej/88MPy3XffZQPd/AwXnTp0VLC++MKLZP7ceUl30YcffqjcI9u5evXqpO/LeeG0adO0b84880wheOV1dOzYUU477TS56KKLZOxY08vGacY2n3fO"
    "udqWEdDd6u/q0pMoPKfXAdjT0tLyXOiOujG4sRAgfwz1Cu897hQgV0LgIwCSw+NBPVZ8cvH7H//4RwXS+++/X6i3zDkBn332WQW+3//+90KuJ/57XISl7qx4"
    "8eJy++23ZzPYhIYHLrN40WJSu2atiB5mZEiYGWhQ+LTRJ9Kja3fl8MhF/vPv/1BOKRw5QZXqhH/84x+yfPnypOncoH4D1YXeftvvJCsrK+n7KlSooFzffffd"
    "F+kFk745diHFc+p1CZKTJ0+O6JrzWY0aNdLrLr30UlmwYMFh/cOFkBz36aeVlL69+2SjUdRXoO+yZcukQ4cOWnJy70dT/5z3FALksaBi4TOOCwXIOVHkpWhW"
    "p06dwyyoAdhWrlypesoSJUrIyy+/rHWjr1zgvPbt2ydly5SVYgDIB+5/IKEDU2fuhHM29ZQ33nijvo+AahPTrgnPmjd/vvziml/oRP6i8RcRHSJ/ST9z7933"
    "SDdYl3lUr1pNSp9eSu95uYHVL5r0sfdT70agZ32TOchh/esf/5QTipfQdyTrIkNO9e9//7sUK1ZMatVKgHwy78x5zVtvvaULE+m/fv367D/7ekH6PQzDExce"
    "LgBhAQs04GeXLp11ISFI/ukPf5SZboSJ0ypsKRt/SX762aNpTyFAHg3VCu85LhQIQBQGPXVVv/7lrxQA3o/ppTRILebLR8MGJyl1jMG6SYAMz1m3Zq3cdvMt"
    "cho4rbfgNmLohJLD7YbcKvWPV199tdBqzMP2lEmA6MD+A+S0U06VC2HFHjV61GFgxxPk5P7yl7/IuPEmUlOEvfGGGxUgL77gIhniusKcoEp3HC4EBR4OPOvX"
    "rpOb8Fxyxm++yUxXyR1zZs+RK3/2czkDFmG6CCVzxNUBcVCqXqWqnHziSfIIDFHBQp7zeR3bd9AFggDYv1//iP5xui5fsVzbQk77LFz3ayxCn4ED37AuAbra"
    "777wxZ3nk6l/stcUAmSylCq87j9CgfjkmzBhgvz88p9K6ZKlZOCAgbnWh35/1H+dfvrp6gMYQC0OkJMnTpLzzz5HzoXBJIBT/GEE4qB7vOKKK2TYsGGOoSFO"
    "PSE6t2jaTE6HnuyKn16h4l444vWePn26iq5r1q2Nfu/Xr5+cfy6MIgCl3wKs165ZY++IRZU89thj8tlnnxVMd68ODUE/Qz3OhKElP/1feE9412AYaFiPC8+/"
    "QN5+8y212H+Fws94oSGHfwfOMKc+kOqCm2CoOgUW9PgCFhpAC3YTtOdMvIvA16JZ80SbPQomToMx8BmlqqAUFpLSKOcAKG8BaL779jvZ1A7xehRykIWhhgVP"
    "mP9jV4RBT47qdHBrV//8ShkEh/D5EG/nzZsnU6ZOVUCgvpGGhhtuuEF69+4dUWEfMhTFuZPePXrJ2Zhs58MIcOcf/yQPPfCg3HfPvVIV/pSVK1aSP0AvSR1a"
    "jRo1chhwLOlCXLR76sl6cgo4prsgogan8pxA165dO7kb7iihHYHbffbpZ1RfSLCoW6u27c7oB/Wq9MEcMGBAwb3pANkBnBm5tyt//vOo3rndnFPX2QL+iKzD"
    "5Zf9RCqWryDVwAVWqVRZ/Q9Doe8k9ZT0o8zLqEV/0KvBiRIAb7rhJnnwwQe1T3gfdbl/++vf5Aq8437Q+jtfdJRWHuqY2+KyDlzxW2+8qX1Ojp+LEev6m1//"
    "BiqNJpEDfV4cbcHEy/+KQg7yx1Kw8P7jRoEXX3xRwehn4CKpO+Tku+fee+Rfd/9brdWcxIwUiVubKeZxW4g4KLwDJ+MzMMloVW7wUgN59ZVX5VXoAukaUzSj"
    "iBpkaCCJT1iL0uB2qwnTy57de+T2W38nJWCgefqpp3LlHnmSerlnnn464paC6Llxw0bVr1HUpvtP1y5do2dQj8o2MSol2eM1tKNEseJyA7g4cmt5HTm5LEal"
    "EFjv+ffdSCFoBq2CjtxEWnLFF4IrpoX+sbqPyeuvvy6vvvqqXHfddZKeni6XQ+UxesTIxKOjOO2E+iPUTVNoxLhpOte/2fANufnGm1Q8P4M6XIAlHdZDVE5O"
    "FUVBbUjm90KATIZKhdf8RygQN4Zw8FesWFFOgv7xURhYaD3etm2bbEXJ2rpFtrlFO17R+GQLQjGfc/+990lxAGGlChWztes9WJpPOelkufSiS2Ts6DEGaO5a"
    "ovHVANr9vmcRf/th2Q9yHdx2CJDvv/deBIB6X2xy161bVzrBrYVHiNMOL6bf4qUXX6KTni5A8+fN159GwOGcHFjcgJFrJ/h7mASCXB8dxO+Hu1GyoiY5VYI0"
    "rfBP10uAfH4dnlM3HN5FQDwRAH3LTTfLtq1hX3DRUEmqPdjGVs1bRHTNzT0rW/85veLn6Aj/0vMvyE8BwhdAAmB7n3/++UjfmWy7kx3QhQCZLKUKr/uPUCAM"
    "eLpwXA2nYuqjPv7gwyOqS0jJxZvItd1x2+1SAn6Lzz1jlmk1TRPwli5TkOKkq1m9RpSOLICrcqGctH7993Covgzgdi70mYwBD8AY51ZpIaaozLjlvI6P4INI"
    "HSDbRl9KHgzDY5RIeGZu9yptvC6b0C7WnSL7O9DR5XVEi4bfx7h13kfDF92QeBSUii08m88K3DA/KUYTIO+CKB030NA3lDpYcqm/R3jgFnc/itMpqB3ocrVg"
    "/gJZtmSpis859Zzh3X169FRxnaBLS39wlyoEyEId5BGBw//vF4cBP2HiBA2Fow6yF5yuA3BYeqv8DwU1t3LPnjVbrkGscmmI2G0Q12sPSnB8L8OPkBOZwDdl"
    "0uF+fCEFF2/7oskXcjLA9Fe/+KUsXrw4qkR8klJHythmisx5HeQS74VeriTaRlG7GcCRomluIZBxcDKAtDOzZsxU0ZYccDcknCgIWAM9GI1E1QIdy78b6sao"
    "AnJVxtsRnkNuniGep4J2z9QzdUK8Dt27d1cwI93btWkT/ZZTH/rxhx+pmuA3v75OyC36hVFKtri3QmVIAOwrAmRQKSTr2lTAkIl+LuQgk6VU4XX/UQr07NlT"
    "zj77bLkYltaJ4ydmn2AF1IwYEjia7t26K4hQjzk1nnnHXXzmz5krV8EgwGvqQTeX84iD8YsvvCgnYDLfCit0iNbJGQ7IhA7Uje7nVhaKGIknKgcWA+5fXHW1"
    "urQw4cb1118fxR/nxRXFwYXuRuRkCXTjx43LEyBzAlKTJk0UkH4JN5rFnkRDt7AtaNXJAYAM77zk4ovV+NW5Y6dsAMl3EkD/glh06pAp0h9m0PL30fpNPTAX"
    "nfVwx9Ij2g/H8z0690vjWhFcW7Zs2SgMtJCDLOQg/6NAdTxfHh/s9V96SR22/wyrc+amRORL3GiSV93ic/3jjz7WSUqraHCtUVDjRX4hrcsnlzgBLkVXyFzP"
    "CBS4ofAsio33Q2zksypXqHQYIIS6N2zYULPQJC7I/jVuiOjQtp1ykGeffZaG4IWMO/nRPLynwUvw2YSYfNtvbxVafkN9c7s3TlcavnjfHXCn2b/PQTzJTo7r"
    "iKlnPAcATePJ98O+87XAEwg7oLVD+2hYOeuss6RP3z5RHRW0fYEaAL9IWqkvBFc79Osh2ekaQ21G3/wMLlgloZYIHgsF6TSTbFa2y2Kbdr0pu7PO1W1fD3I/"
    "6nh279ie19xpUDONs/B7KOFczvPx64/0HtzLzOYsuqMiNw/TZ2Bf7KzTZc/O19EYmyyFx/89CsT1ZYxJzsjI0NjkkEIsmhAFND0OkLUQDngCImz+/re/Z0s+"
    "EeesJsLf8tJLLlHx+QUYAMIRv4YiM91pToHOj2FzeR10k2ncuLH9HAPhAGBm/LEaHoDBhK5FjDC599578zXQxMGJ9z4Kay65qX8ikqagyJtAVybf4HtOhDP8"
    "4/C5DHXKWc9wPugcw7vj9GgMTpRgdRWy7yxZbPk449fx743IYXnrb3+rEU7/+Oc/soVuhmeR0yzzSBlJQ4Ye+lTOhOog50Hn/9tuu01ORjDAm3AByqlXPZYz"
    "IQLIg7tewJYLp8n+XbZZl2RyH2ov4Xu0P7Wfj/aMiV0bdj9M5rfw/HzuOaj74lg5xOtYFwDk7g0ny57tDUALAGSS4sCxJFzhs/7nKRAGPkP+br75ZiYOyJb6"
    "K6e4WFCNGNHC5BRMj/USONKcRxx0qteorqF3zC0ZomjivzNpLzPfEMwo/ud2cLITgIKjeW76sZwgwrhr+mCGVGF5iYxxazjbdQd8N0mfeORNQeImXZkuhlgc"
    "cmLSZYbiavjk93hhW+jTGQe/0G7+xvfffTdchTy5b6hj3P3m888/11h6RjrRLSgccdrSEf0FcLaMYLr8J5fL448/rovMBx98oIYg+rkyu1I8C1NBbS1obOT1"
    "ewIg97SRXVkPAHR+J3sy/yz7M/8uB7b8SfZn/Q3lLpS/yr4tf5T9W38flX34Hkpe5/n70dyj9+F9B7L+gvvvkL3bbscn4lIzcT7r97J38wMA847ARiQsKATI"
    "o+3//9X3BQCkhZIiHCdUyBZ+pD5vfBb1XoxvZrbwvMTXoKuk1ZyiW//+/TUJQpjA4b08R8s168Ts3rkdfB/ddQiU+U3gnCA5FY7vLHHOLTcwD3Xle5gGjvrO"
    "YCzKz1gR6EqLMaNjWNgOGlLocM9Pgn688BwzIM2caRxdeEb45IJBujKNXDgCQKovqu9Fw5yNX3/9tdI2Ct90PWxOrpghmex3Zib/8ssvhRmA+A5GJsWP+H3H"
    "GigTuxoisfA+2YbtHtchWiALmMMsI/zcjrID50LZie/HpyCZPepAx9VteCfrge8HkXr+wC6t0z7ZzY0ebaP0wuP/HAVyTpg4txEHrGQaniyg5vRTzO2dyUzC"
    "nNxtXveE6+JgkvOd+XGRAaziNIhv9JWTNsdCTxff2iC3Poq/n+/Lb7+YeL8caZ8m0w/JjI38rokAsvBLIQUKKVBIgUIKHE6B/wdq5uY+aI+nEAAAAABJRU5E"
    "rkJggg=="
)


# NEW ADDITION: categories for the "💸 Daily Expenses / Staff Expense" tab.
EXPENSE_CATEGORIES = [
    "Bike Fuel (پٹرول)",
    "Bike Maintenance (بائیک مرمت)",
    "Bilty/Vehicle Rent (بلٹی گاڑی کرایہ)",
    "Loading Wages (لوڈنگ مزدوری)",
    "Advance Salary (ایڈوانس سیلری)",
    "Miscellaneous (متفرق)",
]

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def q(sql, params=None):
    named_sql, pdict = _qmark_to_named(sql, params or [])

    def _run():
        t0 = _time.monotonic()
        try:
            engine = _get_engine()
            with engine.connect() as conn:
                df = pd.read_sql(text(named_sql), conn, params=pdict)
            _log_slow_query(sql, _time.monotonic() - t0)
            return df
        except Exception as e:
            _log_slow_query(sql, _time.monotonic() - t0, error=str(e))
            raise

    try:
        return _run()
    except (OperationalError, SAInterfaceError, DBAPIError) as e:
        if not _is_transient_db_error(e):
            raise
        try:
            _get_engine().dispose()
        except Exception:
            pass
        return _run()

def scalar(sql, params=None):
    def _run():
        conn = get_conn()
        try:
            return conn.execute(sql, params or []).fetchone()
        finally:
            conn.close()

    try:
        r = _run()
    except (OperationalError, SAInterfaceError, DBAPIError) as e:
        if not _is_transient_db_error(e):
            raise
        try:
            _get_engine().dispose()
        except Exception:
            pass
        r = _run()
    return (r[0] or 0) if r else 0

@st.cache_data(ttl=20, show_spinner=False)
def get_calloff_list():
    return q("SELECT DISTINCT call_off_no FROM sheet_orders WHERE TRIM(call_off_no)!='' ORDER BY call_off_no")["call_off_no"].tolist()

@st.cache_data(ttl=20, show_spinner=False)
def get_contracts_for_calloff(call_off_no):
    """PERFORMANCE: this exact lookup (distinct Sales Contracts for a
    Call-Off) was being re-run as a fresh round-trip in both DC Entry and
    Bilty Management on every rerun. Cached briefly since Call-Off/Contract
    lists change rarely."""
    if not call_off_no:
        return []
    return q(
        "SELECT DISTINCT sale_contract FROM sheet_orders WHERE call_off_no=? AND TRIM(sale_contract)!='' ORDER BY sale_contract",
        [call_off_no]
    )["sale_contract"].tolist()

# PERFORMANCE FIX (DC Entry 5-8 min lag): these three lookups fire on every
# single Streamlit rerun — which means every keystroke in Step 1/2 — and
# each one used to be a brand-new network round-trip to the DB. With ~7
# item types × 2 lookups plus the article/bilty checks, that's 20-30+ fresh
# round-trips per keystroke. Short-TTL caching collapses repeat calls with
# identical arguments (the overwhelming majority within one typing burst)
# down to a single DB hit. Cache is cleared immediately after any Save so
# the just-entered figures are reflected instantly, not after the 8s TTL.
@st.cache_data(ttl=8, show_spinner=False)
def get_ordered_qty(article, category, coff=None, po=None):
    params, extra = [article, category], ""
    if coff: extra += " AND call_off_no=?"; params.append(coff)
    if po:   extra += " AND po_no=?";       params.append(po)
    return scalar(f"SELECT SUM(order_qty) FROM sheet_orders WHERE article=? AND category=?{extra}", params)

@st.cache_data(ttl=8, show_spinner=False)
def get_received_qty(article, category, coff=None, po=None, exclude_id=None):
    params, extra = [article, category], ""
    if coff:       extra += " AND call_off_no=?"; params.append(coff)
    if po:         extra += " AND po_no=?";       params.append(po)
    if exclude_id: extra += " AND id!=?";         params.append(exclude_id)
    return scalar(f"SELECT SUM(qty) FROM inventory WHERE article=? AND category=?{extra}", params)

@st.cache_data(ttl=8, show_spinner=False)
def get_total_ordered_for_article(call_off_no, article):
    """Sum of order_qty across ALL categories for one Call-Off + Article —
    used by the live Bilty dispatch % indicator in Step 2."""
    return scalar("SELECT SUM(order_qty) FROM sheet_orders WHERE call_off_no=? AND article=?", [call_off_no, article])

@st.cache_data(ttl=8, show_spinner=False)
def get_bilty_qty(call_off_no, article, category=None):
    """Total quantity already dispatched (Bilty) from the Lahore factory for
    this Call-Off + Article (optionally scoped to one category)."""
    params, extra = [call_off_no, article], ""
    if category:
        extra = " AND category=?"; params.append(category)
    return scalar(f"SELECT SUM(qty) FROM bilty WHERE call_off_no=? AND article=?{extra}", params)

def _clear_dc_entry_caches():
    """Called right after any DC Entry save so the freshly-saved figures
    show up immediately instead of waiting out the short cache TTL."""
    get_ordered_qty.clear()
    get_received_qty.clear()
    get_bilty_qty.clear()
    _get_pending_articles.clear()
    get_calloff_brand_po.clear()
    get_bilty_breakdown.clear()
    get_dual_pack_articles.clear()
    get_category_totals_for_contract.clear()
    get_ordered_for_article_contract.clear()
    get_received_for_article_contract.clear()
    get_total_ordered_for_article.clear()

# PERFORMANCE FIX (SHOW-STOPPER — 5-8 min DC Entry lag, follow-up round):
# these six queries were STILL firing uncached on every rerun (i.e. every
# keystroke/selection anywhere in the DC Entry form) even after the first
# round of caching — together roughly 8-10 more fresh network round-trips
# per rerun on top of the ones already fixed. Same short-TTL pattern as
# above; all cleared together in _clear_dc_entry_caches() on save.
@st.cache_data(ttl=8, show_spinner=False)
def get_calloff_brand_po(call_off_no, sale_contract=None):
    """Returns (brand, [po_no, ...]) for the Step 1 auto-load info box."""
    if sale_contract:
        df_brand = q("SELECT DISTINCT brand FROM sheet_orders WHERE call_off_no=? AND sale_contract=? AND TRIM(brand)!='' LIMIT 1", [call_off_no, sale_contract])
        df_po = q("SELECT DISTINCT po_no FROM sheet_orders WHERE call_off_no=? AND sale_contract=? AND TRIM(po_no)!='' ORDER BY po_no", [call_off_no, sale_contract])
    else:
        df_brand = q("SELECT DISTINCT brand FROM sheet_orders WHERE call_off_no=? AND TRIM(brand)!='' LIMIT 1", [call_off_no])
        df_po = q("SELECT DISTINCT po_no FROM sheet_orders WHERE call_off_no=? AND TRIM(po_no)!='' ORDER BY po_no", [call_off_no])
    brand = df_brand.iloc[0]["brand"] if not df_brand.empty else ""
    return brand, df_po["po_no"].tolist()

@st.cache_data(ttl=8, show_spinner=False)
def get_bilty_breakdown(call_off_no, article):
    """Category-wise Bilty breakdown shown under the live dispatch indicator."""
    return q("SELECT category, SUM(qty) AS tot FROM bilty WHERE call_off_no=? AND article=? GROUP BY category ORDER BY category", [call_off_no, article])

@st.cache_data(ttl=8, show_spinner=False)
def get_dual_pack_articles(call_off_no, sale_contract, category):
    """Returns (jersey_articles, molton_articles) for the dual-pack UI toggle."""
    dpj = q("SELECT DISTINCT article FROM sheet_orders WHERE call_off_no=? AND sale_contract=? AND category=? AND variant='Jersey'", [call_off_no, sale_contract, category])
    dpm = q("SELECT DISTINCT article FROM sheet_orders WHERE call_off_no=? AND sale_contract=? AND category=? AND variant='Molton'", [call_off_no, sale_contract, category])
    return dpj["article"].tolist(), dpm["article"].tolist()

@st.cache_data(ttl=8, show_spinner=False)
def get_category_totals_for_contract(call_off_no, sale_contract):
    """Ordered-vs-received totals per category for the Live Contract Status
    Counter — one grouped query each instead of 2-per-category."""
    df_ord = q("SELECT category, SUM(order_qty) AS tot FROM sheet_orders WHERE call_off_no=? AND sale_contract=? GROUP BY category", [call_off_no, sale_contract])
    df_rec = q("SELECT category, SUM(qty) AS tot FROM inventory WHERE call_off_no=? AND contract_no=? GROUP BY category", [call_off_no, sale_contract])
    return dict(zip(df_ord["category"], df_ord["tot"])), dict(zip(df_rec["category"], df_rec["tot"]))

@st.cache_data(ttl=8, show_spinner=False)
def get_ordered_for_article_contract(call_off_no, sale_contract, category, article):
    return scalar("SELECT SUM(order_qty) FROM sheet_orders WHERE call_off_no=? AND sale_contract=? AND category=? AND article=?", [call_off_no, sale_contract, category, article])

@st.cache_data(ttl=8, show_spinner=False)
def get_received_for_article_contract(call_off_no, sale_contract, category, article):
    return scalar("SELECT SUM(qty) FROM inventory WHERE call_off_no=? AND contract_no=? AND category=? AND article=?", [call_off_no, sale_contract, category, article])

@st.cache_data(ttl=8, show_spinner=False)
def _get_pending_articles(call_off_no, sale_contract):
    """PERFORMANCE FIX: this correlated nested-EXISTS query (which article
    numbers under this Call-Off + Contract still have at least one pending
    category) used to re-run uncached on every keystroke of the Call-Off
    field — the single heaviest query in Step 2. Short-TTL cache only."""
    return q("""
        SELECT DISTINCT so.article
        FROM sheet_orders AS so
        WHERE so.call_off_no=? AND so.sale_contract=?
          AND TRIM(so.article)!=''
          AND EXISTS (
              SELECT 1
              FROM sheet_orders AS pending
              WHERE pending.call_off_no=so.call_off_no
                AND pending.sale_contract=so.sale_contract
                AND pending.article=so.article
                AND NOT EXISTS (
                    SELECT 1
                    FROM inventory AS inv
                    WHERE inv.call_off_no=pending.call_off_no
                      AND inv.contract_no=pending.sale_contract
                      AND inv.article=pending.article
                      AND inv.category=pending.category
                )
          )
        ORDER BY so.article
    """, [call_off_no, sale_contract])["article"].tolist()

@st.cache_data(ttl=3600, show_spinner=False)
def get_cached_item_description(article, category):
    """Looks up the last exact PO wording an operator typed for this
    Article+Category (e.g. 'INLAY CARD FITTED 36.5X40.5 CM- DIXX JERSEY'),
    so the DC Entry form can auto-suggest it instead of retyping."""
    if not article or not category:
        return ""
    df = q("SELECT description FROM item_desc_cache WHERE article=? AND category=? ORDER BY id DESC LIMIT 1", [article, category])
    return df.iloc[0]["description"] if not df.empty else ""

@st.cache_data(ttl=3600, show_spinner=False)
def get_cached_item_descriptions_list(article, category):
    """NEW ADDITION: returns ALL distinct descriptions ever saved for this
    Article+Category (most recent first, capped at 5) — not just the last
    one. This is what powers the Combo Article description dropdown: a
    combo article carries two DIFFERENT quality descriptions (e.g. one
    'MOLTEN+LYCRA' Inlay Card line and one plain 'MOLTEN' Inlay Card line)
    under the same article+category, and the old single-slot cache used to
    overwrite one with the other every time you saved."""
    if not article or not category:
        return []
    df = q("SELECT description FROM item_desc_cache WHERE article=? AND category=? ORDER BY id DESC LIMIT 5", [article, category])
    return df["description"].tolist()

def save_cached_item_description(article, category, description):
    if not article or not category or not str(description).strip():
        return
    desc = description.strip()
    conn = get_conn()
    # BUGFIX: this used to DELETE the existing row before inserting the new
    # one — a single-slot "last description wins" cache. For a Combo
    # Article that means saving quality #2's description would silently
    # erase quality #1's remembered wording. Now it keeps a short history
    # instead (skip the insert if this exact wording is already saved;
    # otherwise add it and trim to the 5 most recent per article+category).
    existing = conn.execute(
        "SELECT id FROM item_desc_cache WHERE article=? AND category=? AND description=?",
        (article, category, desc)
    ).fetchone()
    if not existing:
        conn.execute("INSERT INTO item_desc_cache (article, category, description) VALUES (?,?,?)", (article, category, desc))
        conn.commit()
        conn.execute("""
            DELETE FROM item_desc_cache WHERE article=? AND category=? AND id NOT IN (
                SELECT id FROM (
                    SELECT id FROM item_desc_cache WHERE article=? AND category=? ORDER BY id DESC LIMIT 5
                ) AS keep_ids
            )
        """, (article, category, article, category))
        conn.commit()
    conn.close()
    get_cached_item_description.clear()
    get_cached_item_descriptions_list.clear()

def autofill_item_description(widget_key, source_key, article, category):
    """Refresh a description field only when its selected item changes."""
    selection = (str(article or ""), str(category or ""))
    if st.session_state.get(source_key) != selection:
        st.session_state[widget_key] = get_cached_item_description(article, category)
        st.session_state[source_key] = selection

def round_and_format(val):
    try:
        if pd.isna(val) or val == "" or val == "—":
            return "0"
        if isinstance(val, str):
            val = float(val.replace(",", "").strip())
        rounded_val = int(math.floor(val + 0.5))
        return f"{rounded_val:,}"
    except:
        return "0"

def round_bal(val):
    """Strict zero-balancing helper: rounds away Python float micro-garbage
    (e.g. 0.00001 / -0.00001) so true-zero rows are correctly detected."""
    try:
        if pd.isna(val) or val == "" or val == "—":
            return 0
        if isinstance(val, str):
            val = float(val.replace(",", "").strip())
        return int(math.floor(float(val) + 0.5))
    except:
        return 0

def build_article_blocks(df_ledger):
    """BUGFIX (data-merging, priority): this used to group by [Article, Item
    Type] ONLY, ignoring Call-Off entirely — so the same Article Number
    running under two different Call-Offs (e.g. 291 and 292) got summed
    together into a single merged line. That's wrong: quantities must stay
    strictly separated by Call-Off even when the Article Number repeats.
    Now groups by [Call-Off No, Article, Item Type] instead, and returns an
    ORDERED LIST of (call_off_no, article, {item_type: balance}) tuples —
    not a dict keyed by article alone, since a dict can't hold two separate
    entries for the same article number. Each Call-Off/Article combination
    always renders as its own distinct block with its own individual
    quantity, never merged with any other Call-Off's total for that article.
    Also drops any category whose remaining balance rounds to zero within
    that specific Call-Off+Article block, same as before.
    """
    blocks = []
    if df_ledger is None or df_ledger.empty:
        return blocks
    has_calloff = "Call-Off No" in df_ledger.columns
    group_cols = ["Call-Off No", "Article", "Item Type"] if has_calloff else ["Article", "Item Type"]
    grouped = df_ledger.groupby(group_cols)["Remaining Balance"].sum().reset_index()
    key_cols = ["Call-Off No", "Article"] if has_calloff else ["Article"]
    keys = grouped[key_cols].drop_duplicates().sort_values(key_cols).values.tolist()
    for key_vals in keys:
        if has_calloff:
            coff, article = key_vals
            sub = grouped[(grouped["Call-Off No"] == coff) & (grouped["Article"] == article)]
        else:
            article = key_vals[0]
            coff = None
            sub = grouped[grouped["Article"] == article]
        cats = {}
        for _, r in sub.iterrows():
            bal = round_bal(r["Remaining Balance"])
            if bal != 0:
                cats[r["Item Type"]] = bal
        if cats:
            blocks.append((coff, article, cats))
    return blocks

# ─────────────────────────────────────────────
# PDF GENERATION
# ─────────────────────────────────────────────
def generate_ledger_pdf(df_summary, df_articles, sel_coff, sel_cont, sel_art, report_type="MASTER"):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    if report_type == "MASTER":
        title_color = '#1e40af'
        title_text = "MASTER LEDGER STATUS REPORT"
    elif report_type == "CONTRACT_SHORTLIST":
        title_color = '#7c3aed'
        title_text = "CONTRACT-WISE SHORTLIST REPORT"
    else:
        title_color = '#b91c1c'
        title_text = "PENDING SHORTAGE REPORT"
    
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=15, leading=19, textColor=colors.HexColor(title_color), alignment=1)
    cust_style = ParagraphStyle('Cust', parent=styles['Normal'], fontSize=11, leading=15, textColor=colors.HexColor('#0f172a'), alignment=1)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, leading=13, alignment=1)
    h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=11, leading=15, textColor=colors.HexColor('#0f172a'), spaceBefore=8, spaceAfter=4)
    
    story.append(Paragraph(f"<b>VERTEX PACKAGING — {title_text}</b>", title_style))
    story.append(Spacer(1, 2))
    story.append(Paragraph("<b>Customer Name: Vertex Shahzad Bhai Lahore</b>", cust_style))
    story.append(Spacer(1, 4))
    
    filter_info = f"Call-Off: <b>{sel_coff}</b> | Contract: <b>{sel_cont}</b> | Article: <b>{sel_art}</b>"
    story.append(Paragraph(f"Generated: {datetime.datetime.now().strftime('%d-%m-%Y %I:%M %p')} | {filter_info}", sub_style))
    story.append(Spacer(1, 12))
    
    story.append(Paragraph("<b>📊 SECTION 1: ITEM-WISE REMAINING SUMMARY</b>", h2_style))
    sum_data = [["Item Type / Description", "Remaining Balance (Pcs)"]]
    for k, v in df_summary.items():
        sum_data.append([k, round_and_format(v)])
        
    t_sum = Table(sum_data, colWidths=[280, 220])
    t_sum.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (1,0), colors.HexColor(title_color)),
        ('TEXTCOLOR', (0,0), (1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (1,1), (1,-1), 'RIGHT'),
    ]))
    story.append(t_sum)
    story.append(Spacer(1, 12))

    # NEW: SECTION 1B — Article-Wise Breakdown blocks. Fixes the bug where
    # only flat global category totals were visible — each article now gets
    # its own boxed block with its nested, non-zero category balances.
    article_blocks = build_article_blocks(df_articles)
    if article_blocks:
        story.append(Paragraph("<b>🎯 SECTION 1B: ARTICLE-WISE BREAKDOWN</b>", h2_style))
        block_cell_style = ParagraphStyle('BlockCell', parent=styles['Normal'], fontSize=8, leading=12)
        block_head_style = ParagraphStyle('BlockHead', parent=styles['Normal'], fontSize=9, leading=13,
                                           textColor=colors.white, fontName='Helvetica-Bold')
        NCOLS = 3
        block_cells = []
        for coff, article, cats in article_blocks:
            lines = "<br/>".join(f"📦 {cat}: <b>{bal:,}</b> Pcs" for cat, bal in cats.items())
            header_txt = f"🎯 Article: {article}" + (f" &nbsp;|&nbsp; Call-Off: {coff}" if coff else "")
            mini = Table(
                [[Paragraph(header_txt, block_head_style)],
                 [Paragraph(lines, block_cell_style)]],
                colWidths=[158]
            )
            mini.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (0,0), colors.HexColor(title_color)),
                ('BACKGROUND', (0,1), (0,1), colors.HexColor('#f8fafc')),
                ('BOX', (0,0), (-1,-1), 0.75, colors.HexColor(title_color)),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor(title_color)),
                ('TOPPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('LEFTPADDING', (0,0), (-1,-1), 6),
            ]))
            block_cells.append(mini)

        rows = [block_cells[i:i+NCOLS] for i in range(0, len(block_cells), NCOLS)]
        for row in rows:
            while len(row) < NCOLS:
                row.append("")
        block_grid = Table(rows, colWidths=[166]*NCOLS)
        block_grid.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        story.append(block_grid)
        story.append(Spacer(1, 12))

    sec2_title = "<b>🔍 SECTION 2: PENDING ARTICLES BREAKDOWN (SHORTAGE LIST)</b>" if report_type in ("SHORTAGE", "CONTRACT_SHORTLIST") else "<b>🔍 SECTION 2: ARTICLES COMPLETE BREAKDOWN</b>"
    story.append(Paragraph(sec2_title, h2_style))

    # Remarks column: only for Shortage / Contract Shortlist reports, showing
    # the manual remarks typed by the operator during DC entry for that
    # exact Call-Off + Article + Item Type combination.
    show_remarks = report_type in ("SHORTAGE", "CONTRACT_SHORTLIST")
    remarks_map = {}
    if show_remarks:
        df_remarks_raw = q("SELECT call_off_no, article, category, remark FROM inventory WHERE TRIM(remark)!=''")
        if not df_remarks_raw.empty:
            for (co, art, cat), grp in df_remarks_raw.groupby(["call_off_no", "article", "category"]):
                uniq = list(dict.fromkeys([str(x).strip() for x in grp["remark"] if str(x).strip()]))
                remarks_map[(co, art, cat)] = "; ".join(uniq)
    remark_cell_style = ParagraphStyle('RemarkCell', parent=styles['Normal'], fontSize=7, leading=9, textColor=colors.HexColor('#334155'))

    if show_remarks:
        art_data = [["Call-Off", "Contract #", "Art. No", "Item Type", "Ordered", "Received", "Remaining", "Remarks"]]
    else:
        art_data = [["Call-Off", "Contract #", "Art. No", "Item Type", "Ordered", "Received", "Remaining"]]

    for _, r in df_articles.iterrows():
        row = [
            str(r["Call-Off No"]), str(r["Contract #"]), str(r["Article"]), str(r["Item Type"]),
            round_and_format(r['Total Ordered']), round_and_format(r['Total Received']), round_and_format(r['Remaining Balance'])
        ]
        if show_remarks:
            rmk_text = remarks_map.get((r["Call-Off No"], r["Article"], r["Item Type"]), "-")
            row.append(Paragraph(rmk_text, remark_cell_style))
        art_data.append(row)

    if len(art_data) == 1:
        no_data_msg = "No pending / active items found for this Brand & Contract selection." if report_type == "CONTRACT_SHORTLIST" else "No pending / shortage items found for current filter selection."
        ncols = 8 if show_remarks else 7
        art_data.append([no_data_msg] + ["-"] * (ncols - 1))

    col_widths = [50, 60, 60, 90, 50, 50, 50, 102] if show_remarks else [55, 65, 65, 125, 65, 65, 65]
    t_art = Table(art_data, colWidths=col_widths)
    t_art.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (4,1), (6,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(t_art)
    story.append(Spacer(1, 25))
    
    sig_data = [["-------------------------\nReport Checked By\n(NABA Packaging Team)", "-------------------------\nAuthorized Signature\n(Kaleem Ullah Sharif)"]]
    t_sig = Table(sig_data, colWidths=[250, 250])
    t_sig.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER'), ('FONTSIZE', (0,0), (-1,-1), 9)]))
    story.append(t_sig)
    
    doc.build(story)
    buffer.seek(0)
    return buffer

# ═══════════════════════════════════════════════
# NEW ADDITION: Ditto DC generation — moved to module level so the exact
# same functions/layout are reusable from BOTH the DC Entry tab and the All
# Entries tab (previously duplicated only in All Entries). Incorporates all
# of the client's latest layout requests:
#  - Header order: Phone -> Email -> "DELIVERY CHALLAN" (moved down)
#  - DC #, Company PO (was "Token #"), and Call-Off number repositioned to
#    sit directly above the item table (right above the Remarks column)
#  - "PACK OF 00 TO 00 CARTONS" placeholder removed entirely
#  - Dates printed as DD-MM-YYYY
#  - "Prepared By: {logged-in user}" + a Receiver Name / Signature & Stamp
#    acknowledgement block at the absolute bottom of the page
# ═══════════════════════════════════════════════
DITTO_GLOBAL_CATEGORIES = {"Safety", "Washing Paper"}
DITTO_CATEGORY_SHORT_LABELS = {
    "Inlay Card / Bandrolle": "Inlay Cards",
    "Tag Card / Barcode Sticker": "Tag Cards",
    "Barcode Item": "Barcode Stickers",
    "Safety": "Safety Stickers",
    "Washing Paper": "Washing Papers",
    "Transparent Sticker": "Transparent Stickers",
    "Eco Friendly": "Eco Stickers",
}

def _vogue_logo_flowable(height=34):
    """Decodes the embedded VOGUE Printers logo PNG into a reportlab Image
    flowable, preserving its original aspect ratio, for the DC PDF header."""
    raw = base64.b64decode(VOGUE_LOGO_B64)
    src_w, src_h = 328, 91  # native pixel size of the supplied logo PNG
    width = height * (src_w / src_h)
    return Image(io.BytesIO(raw), width=width, height=height)

def _fmt_date_ddmmyyyy(d):
    """Prints dates as DD-MM-YYYY on all printouts, regardless of how the
    date is stored internally (kept as ISO YYYY-MM-DD in the DB itself)."""
    try:
        s = str(d).split(" ")[0]
        y, m, dd = s.split("-")
        if len(y) == 4:
            return f"{dd}-{m}-{y}"
        return s
    except Exception:
        return str(d)

def _dc_article_matrix(items):
    cats_present = sorted({i["category"] for i in items if i["category"] not in DITTO_GLOBAL_CATEGORIES})
    # Consolidated DC exports intentionally suppress the old article-by-article
    # matrix; category totals are shown instead.
    if items and all(i.get("summary_only") for i in items):
        return [], [], {}
    arts_present = sorted({i["article"] for i in items})
    matrix = {a: {c: 0.0 for c in cats_present} for a in arts_present}
    for i in items:
        if i["category"] in DITTO_GLOBAL_CATEGORIES:
            continue
        matrix[i["article"]][i["category"]] += float(i.get("qty") or 0)
    return cats_present, arts_present, matrix

def _dc_category_totals(items):
    totals = {}
    for i in items:
        totals[i["category"]] = totals.get(i["category"], 0) + float(i.get("qty") or 0)
    return totals

def _generate_ditto_dc_excel(dc_no, call_off_no, contract_no, token_no, destination, entry_date, items, filename_base, prepared_by, company_type="Vertex Packaging"):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename_base[:31]
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    def _centered(row, text, size=10, b=False):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        c = ws.cell(row=row, column=2, value=text)
        c.font = Font(bold=b, size=size)
        c.alignment = center

    ddmmyyyy = _fmt_date_ddmmyyyy(entry_date)

    # NEW ADDITION: dynamic letterhead — Vertex Packaging vs VOGUE Printers.
    lh = COMPANY_LETTERHEADS.get(company_type, COMPANY_LETTERHEADS["Vertex Packaging"])
    if company_type == "VOGUE Printers":
        try:
            from openpyxl.drawing.image import Image as XLImage
            xl_logo = XLImage(io.BytesIO(base64.b64decode(VOGUE_LOGO_B64)))
            xl_logo.height, xl_logo.width = 55, 198  # keeps native ~328:91 aspect ratio
            ws.row_dimensions[2].height = 42
            ws.add_image(xl_logo, "C2")
        except Exception:
            _centered(2, lh["name"], size=20, b=True)  # fallback if Pillow/image load fails
    else:
        _centered(2, lh["name"], size=20, b=True)
    _centered(3, lh["address"], size=10)
    _centered(4, lh["phone"].replace("&nbsp;", " "), size=10)
    _centered(5, lh["email"], size=10)
    _centered(6, "DELIVERY CHALLAN", size=14, b=True)

    ws["B8"] = f"Date: {ddmmyyyy}"
    ws["B9"] = f"Cont #{contract_no}"
    ws.merge_cells("B10:F10")
    ws["B10"] = "Customer Name:  Gul Ahmed Textile Mills Limited (Karachi)"

    # DC #, Company PO (Token), Call-Off — moved directly above the item
    # table (right above the Remarks column), as requested.
    ws["G8"] = f"DC No: {dc_no}"
    ws["D12"] = None
    ws["D12"] = f"Company PO # {token_no}" if token_no else "Company PO # —"
    ws["G9"] = f"Call-Off: {call_off_no}"
    ws["D12"] = None
    ws["G10"] = f"Destination: {destination}"
    ws["G11"] = f"PO: {token_no}" if token_no else "PO: —"
    for cell in ("G8", "G9", "G10", "G11"):
        ws[cell].font = bold
    ws["G11"] = f"PO: {token_no}" if token_no else "PO: -"

    headers = ["S.No", "Customer PO", "Item Type / Description (as per PO)", "UOM", "Quantity", "Remarks"]
    for i, h in enumerate(headers):
        ws.cell(row=13, column=2 + i, value=h).font = bold

    n_slots = max(7, len(items))
    for i in range(n_slots):
        r = 14 + i
        ws.cell(row=r, column=2, value=i + 1)
        ws.cell(row=r, column=5, value="Nos")
        if i < len(items):
            item = items[i]
            ws.cell(row=r, column=3, value=item.get("customer_po", ""))
            ws.cell(row=r, column=4, value=item.get("description", ""))
            ws.cell(row=r, column=6, value=item.get("qty", ""))
            ws.cell(row=r, column=7, value=item.get("remark", ""))

    # Per-category Total Sum lines (this DC's qty only)
    r = 14 + n_slots + 2
    cat_totals = _dc_category_totals(items)
    for cat, tot in cat_totals.items():
        ws.cell(row=r, column=4, value=f"{cat} Total: {round_and_format(tot)} Pcs").font = bold
        r += 1
    r += 1

    # Footer: Prepared By + Receiver acknowledgement block, at the bottom
    r += 3
    ws.cell(row=r, column=2, value=f"Prepared By: {prepared_by}").font = bold
    r += 3
    ws.cell(row=r, column=2, value="Receiver Name: ______________________________")
    r += 2
    ws.cell(row=r, column=2, value="Signature & Stamp: ______________________________")

    for col, w in {"B": 12, "C": 14, "D": 42, "E": 12, "F": 12, "G": 18}.items():
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _generate_ditto_dc_pdf(dc_no, call_off_no, contract_no, token_no, destination, entry_date, items, prepared_by, matrix_items=None, company_type="Vertex Packaging"):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=25, bottomMargin=25)
    story = []
    styles = getSampleStyleSheet()
    center_title = ParagraphStyle('DCTitle', parent=styles['Normal'], fontSize=22, leading=26, alignment=1, fontName='Helvetica-Bold')
    center_sub = ParagraphStyle('DCSub', parent=styles['Normal'], fontSize=10, leading=13, alignment=1)
    center_dc = ParagraphStyle('DCHdr', parent=styles['Normal'], fontSize=14, leading=18, alignment=1, fontName='Helvetica-Bold')

    ddmmyyyy = _fmt_date_ddmmyyyy(entry_date)

    # NEW ADDITION: dynamic letterhead — Vertex Packaging vs VOGUE Printers.
    # Backend logic/formulas are identical either way; only this header block
    # changes based on the company_type saved with the DC.
    lh = COMPANY_LETTERHEADS.get(company_type, COMPANY_LETTERHEADS["Vertex Packaging"])
    if company_type == "VOGUE Printers":
        # Client-supplied logo, centered at the top like the Vertex title.
        logo_img = _vogue_logo_flowable(height=40)
        logo_img.hAlign = 'CENTER'
        story.append(logo_img)
        story.append(Spacer(1, 4))
    else:
        story.append(Paragraph(lh["name"], center_title))
    story.append(Paragraph(lh["address"], center_sub))
    story.append(Paragraph(lh["phone"], center_sub))
    story.append(Paragraph(lh["email"], center_sub))
    story.append(Paragraph("DELIVERY CHALLAN", center_dc))
    story.append(Spacer(1, 10))

    # Keep delivery references in one balanced two-column header. The right
    # column is stacked opposite the left details and finishes above Remarks.
    info_data = [
        [f"Date: {ddmmyyyy}", f"DC No: {dc_no}"],
        [f"Sale Contract: {contract_no}", f"Company Token/PO: {token_no}" if token_no else "Company Token/PO: —"],
        ["Customer Name:  Gul Ahmed Textile Mills Limited (Karachi)", f"Call-Off No: {call_off_no}"],
        ["", f"Destination: {destination}"],
    ]
    # Right-side values align exactly with the 125-point Remarks column.
    # The saved company token is intentionally printed with the label "PO".
    info_data[1][1] = f"Call-Off: {call_off_no}"
    info_data[2][1] = f"Destination: {destination}"
    info_data[3][1] = f"PO: {token_no}" if token_no else "PO: —"
    t_info = Table(info_data, colWidths=[415, 125])
    t_info.setStyle(TableStyle([('FONTSIZE', (0, 0), (-1, -1), 10), ('BOTTOMPADDING', (0, 0), (-1, -1), 4)]))
    story.append(t_info)
    story.append(Spacer(1, 10))

    item_data = [["S.No", "Customer PO", "Item Type / Description (as per PO)", "UOM", "Quantity", "Remarks"]]
    for i, item in enumerate(items):
        item_data.append([i + 1, item.get("customer_po", ""), item.get("description", ""),
                           "Nos", round_and_format(item.get("qty") or 0), item.get("remark", "")])
    n_slots = max(7, len(items))
    for i in range(len(items), n_slots):
        item_data.append([i + 1, "", "", "Nos", "", ""])

    t_items = Table(item_data, colWidths=[35, 70, 210, 40, 60, 125])
    t_items.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
    ]))
    story.append(t_items)
    story.append(Spacer(1, 10))

    cat_totals = _dc_category_totals(items)
    for cat, tot in cat_totals.items():
        story.append(Paragraph(f"<b>{cat} Total: {round_and_format(tot)}</b>",
                                ParagraphStyle('CatTot', parent=styles['Normal'], fontSize=9, alignment=2)))
    story.append(Spacer(1, 14))

    # BUGFIX: the main item table above is intentionally CONSOLIDATED across
    # articles (one row per distinct accessory, article dropped from the
    # key — see the grouping comment in render_ditto_dc_section), so `items`
    # itself no longer carries a usable "article" value here. That's why
    # this Article-Wise Summary used to print a blank Article # column
    # whenever a DC spanned more than one article. Feeding it `matrix_items`
    # instead — the UN-consolidated per-article rows — restores the real
    # Article Numbers next to their Tag Card / Inlay Card breakdown.
    cats_present, arts_present, matrix = _dc_article_matrix(matrix_items if matrix_items is not None else items)
    if arts_present:
        # NEW ADDITION: when the article list is long, spread it across 2 or
        # 3 side-by-side column-blocks instead of one long vertical list, so
        # the whole DC reliably stays on a single printed page.
        n = len(arts_present)
        n_cols = 1 if n <= 8 else (2 if n <= 16 else 3)
        chunk_size = math.ceil(n / n_cols)
        chunks = [arts_present[i:i + chunk_size] for i in range(0, n, chunk_size)]

        mini_tables = []
        for chunk in chunks:
            data = [["Article #"] + [DITTO_CATEGORY_SHORT_LABELS.get(c, c) for c in cats_present]]
            for art in chunk:
                data.append([art] + [round_and_format(matrix[art][c]) for c in cats_present])
            col_w = [55] + [max(45, int(160 / max(len(cats_present), 1)))] * len(cats_present)
            t = Table(data, colWidths=col_w)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
            ]))
            mini_tables.append(t)

        story.append(Paragraph("<b>Article-Wise Summary</b>", ParagraphStyle('AWTitle', parent=styles['Normal'], fontSize=10, spaceAfter=4)))
        if len(mini_tables) == 1:
            story.append(mini_tables[0])
        else:
            container = Table([mini_tables])
            container.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(container)

    # Footer: Prepared By + Receiver acknowledgement, pinned near the bottom
    story.append(Spacer(1, 30))
    sign_data = [[f"Prepared By: {prepared_by}", ""],
                 ["Receiver Name: ______________________________", ""],
                 ["Signature & Stamp: ______________________________", ""]]
    t_sign = Table(sign_data, colWidths=[300, 240])
    t_sign.setStyle(TableStyle([('FONTSIZE', (0, 0), (-1, -1), 9), ('TOPPADDING', (0, 0), (-1, -1), 10)]))
    story.append(t_sign)

    doc.build(story)
    buffer.seek(0)
    return buffer

@st.cache_data(ttl=300, show_spinner=False)
def _build_ditto_dc_exports(dc_no, call_off_no, contract_no, token_no, destination,
                            entry_date, items_json, filename_base, prepared_by, matrix_items_json=None,
                            company_type="Vertex Packaging"):
    """Build each unchanged DC once; Streamlit reruns then reuse the bytes."""
    items = json.loads(items_json)
    matrix_items = json.loads(matrix_items_json) if matrix_items_json is not None else None
    pdf = _generate_ditto_dc_pdf(
        dc_no, call_off_no, contract_no, token_no, destination, entry_date, items, prepared_by,
        matrix_items=matrix_items, company_type=company_type)
    xlsx = _generate_ditto_dc_excel(
        dc_no, call_off_no, contract_no, token_no, destination, entry_date, items,
        filename_base, prepared_by, company_type=company_type)
    return pdf.getvalue(), xlsx.getvalue()

def render_ditto_dc_section(key_prefix, default_dc=None):
    """Reusable Ditto DC preview/export block — used identically from both
    the DC Entry tab and the All Entries tab."""
    ditto_dc_opts = q("SELECT DISTINCT dc_no FROM inventory WHERE TRIM(dc_no)!='' ORDER BY dc_no")["dc_no"].tolist()
    if not ditto_dc_opts:
        st.info("No DC entries available yet.")
        return
    default_idx = 0
    opts_with_blank = ["-- Select --"] + ditto_dc_opts
    if default_dc and default_dc in ditto_dc_opts:
        default_idx = opts_with_blank.index(default_dc)
    ditto_dc_sel = st.selectbox("Select DC No.", opts_with_blank, index=default_idx, key=f"{key_prefix}_ditto_dc_sel")

    if ditto_dc_sel == "-- Select --":
        return

    df_dc_lines_raw = q("""
        SELECT call_off_no, contract_no, po_no, article, category, qty, entry_date, remark,
               item_description, company_token, destination, style_type, company_type
        FROM inventory WHERE dc_no=? ORDER BY id
    """, [ditto_dc_sel])

    if df_dc_lines_raw.empty:
        st.info("No line items found for this DC.")
        return

    # BUGFIX (priority — cross-Call-Off data merging on the single-DC
    # printout): "DC No." is a free-typed text field, NOT validated unique
    # per Call-Off. If the same DC No. was ever saved under two different
    # Call-Offs (typo or reused DC No.), the query above — which only
    # filtered on dc_no — pulled BOTH Call-Offs' lines into one printout,
    # even though the header only ever displayed one Call-Off number. That
    # is exactly what produced the mixed Article-Wise Summary reported.
    # Fix: determine the dominant (most common) Call-Off for this DC No.,
    # then strictly filter every line — main table AND Article-Wise
    # Summary — to that Call-Off only. Anything else is excluded and
    # flagged below instead of silently blending in.
    coff_series = df_dc_lines_raw["call_off_no"].astype(str).str.strip()
    dc_calloff = coff_series.mode().iloc[0]
    df_dc_lines = df_dc_lines_raw[coff_series == dc_calloff].reset_index(drop=True)
    _stray_rows = df_dc_lines_raw[coff_series != dc_calloff]
    if not _stray_rows.empty:
        _other_coffs = sorted(_stray_rows["call_off_no"].astype(str).str.strip().unique().tolist())
        st.warning(
            f"⚠️ DC No. **{ditto_dc_sel}** is also saved under a different Call-Off "
            f"({', '.join(_other_coffs)}) in the database. This printout is strictly "
            f"limited to Call-Off **{dc_calloff}** — those other lines are excluded, "
            f"not merged in. This usually means the DC No. was accidentally reused; "
            f"consider correcting the DC No. on the other Call-Off's entries."
        )

    hdr = df_dc_lines.iloc[0]
    dc_token = hdr["company_token"] if str(hdr["company_token"]).strip() else ""
    dc_dest = hdr["destination"] if str(hdr["destination"]).strip() else ""
    # NEW ADDITION: which brand's letterhead this DC prints under. Old rows
    # saved before this feature default to Vertex Packaging (see migration).
    dc_company = hdr["company_type"] if str(hdr.get("company_type") or "").strip() in COMPANY_TYPES else "Vertex Packaging"

    # ═══════════════════════════════════════════════
    # BUGFIX: rows were duplicating per-article because (a) the fallback
    # description embeds the article number when no custom description was
    # saved, and (b) a second grouping pass keyed on that already-unique
    # text — which never matched across articles — then prepended the
    # category name AGAIN ("Inlay Card / Bandrolle: Inlay Card / Bandrolle
    # — Article 200991"). Fixed by consolidating ONCE, purely by
    # (category, description), dropping article from the key entirely so
    # the SAME accessory dispatched under different articles correctly
    # collapses into a single summed row — matching the client's reference
    # DC sample (one row per distinct item, no article-wise repetition).
    # ═══════════════════════════════════════════════
    def _is_inlay_or_bandroll(desc, cat):
        t = f"{desc} {cat}".lower()
        return "inlay card" in t or "band roll" in t

    raw_rows = df_dc_lines.to_dict("records")
    grouped = {}
    for r in raw_rows:
        desc = str(r["item_description"]).strip() or str(r["category"])
        key = (r["category"], desc)
        if key not in grouped:
            grouped[key] = {
                "customer_po": r["po_no"], "description": desc, "qty": 0.0,
                "remarks": [], "category": r["category"], "article": "",
                "style_totals": {"Normal": 0.0, "Topper": 0.0, "Split": 0.0},
                "is_inlay": _is_inlay_or_bandroll(desc, r["category"]),
            }
        grouped[key]["qty"] += float(r["qty"] or 0)
        rmk = str(r["remark"] or "").strip()
        if rmk and rmk not in grouped[key]["remarks"]:
            grouped[key]["remarks"].append(rmk)
        st_type = str(r.get("style_type") or "Normal").strip()
        if st_type not in grouped[key]["style_totals"]:
            st_type = "Normal"
        grouped[key]["style_totals"][st_type] += float(r["qty"] or 0)

    line_items = []
    for g in grouped.values():
        if g["is_inlay"]:
            s = g["style_totals"]
            remark = f"Normal: {int(s['Normal']):,} | Topper: {int(s['Topper']):,} | Split: {int(s['Split']):,}"
        else:
            remark = " | ".join(g["remarks"])
        line_items.append({
            "customer_po": g["customer_po"], "description": g["description"], "qty": g["qty"],
            "remark": remark, "category": g["category"], "article": g["article"],
        })


    # NEW: matrix_items keeps the RAW per-article rows (article number intact)
    # purely for the PDF's Article-Wise Summary — separate from `line_items`
    # above, which is deliberately consolidated across articles for the main
    # table. See the article-number bugfix note inside _generate_ditto_dc_pdf.
    matrix_items = [
        {"article": r["article"], "category": r["category"], "qty": float(r["qty"] or 0)}
        for r in raw_rows
    ]

    ditto_filename_base = f"DC-{ditto_dc_sel}_GulAhmed_{_fmt_date_ddmmyyyy(hdr['entry_date'])}_Cont-{hdr['contract_no']}"
    st.markdown(f"**📄 File Name:** `{ditto_filename_base}`")

    prepared_by = current_user["full_name"] if current_user else "—"

    items_json = json.dumps(line_items, sort_keys=True, default=str)
    matrix_items_json = json.dumps(matrix_items, sort_keys=True, default=str)
    pdf_bytes, xlsx_bytes = _build_ditto_dc_exports(
        ditto_dc_sel, hdr["call_off_no"], hdr["contract_no"], dc_token, dc_dest,
        str(hdr["entry_date"]), items_json, ditto_filename_base, prepared_by,
        matrix_items_json=matrix_items_json, company_type=dc_company)
    ditto_pdf_buf = io.BytesIO(pdf_bytes)
    ditto_xlsx_buf = io.BytesIO(xlsx_bytes)

    st.markdown("##### 👁️ Live Print Preview")
    b64_pdf = base64.b64encode(ditto_pdf_buf.getvalue()).decode()
    # BUGFIX (Chrome pop-up blocker): the preview iframe and the Print button
    # are now rendered together inside ONE components.html sandbox, so the
    # button can call .print() directly on the embedded iframe instead of
    # opening a new window/tab — which is what Chrome was blocking before.
    components.html(f"""
        <iframe id="dcPreview_{key_prefix}" src="data:application/pdf;base64,{b64_pdf}"
                width="100%" height="600" style="border:1px solid #334155;border-radius:8px;"></iframe>
        <button onclick="document.getElementById('dcPreview_{key_prefix}').contentWindow.print()"
                style="width:100%;padding:9px 0;margin-top:8px;background:#ef4444;color:white;border:none;
                       border-radius:6px;cursor:pointer;font-weight:600;font-size:14px;">
            🖨️ Print DC
        </button>
    """, height=660)

    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button(
            label="⬇️ Download PDF Report", data=ditto_pdf_buf,
            file_name=f"{ditto_filename_base}.pdf", mime="application/pdf",
            type="primary", key=f"{key_prefix}_download_pdf")
    with dl2:
        st.download_button(
            label="⬇️ Download Excel Report", data=ditto_xlsx_buf,
            file_name=f"{ditto_filename_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{key_prefix}_download_xlsx")

# ─────────────────────────────────────────────
# PAGE CONFIG & CSS
# ─────────────────────────────────────────────
st.set_page_config(page_title="Vertex Packaging | Inventory System", layout="wide", page_icon="📦")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif;}
.stApp{background-color:#1a202c !important;color:#e2e8f0 !important;}
.hdr{background:linear-gradient(135deg,#0f172a,#1e293b);padding:20px 30px;border-radius:12px;
     margin-bottom:20px;border-left:6px solid #38bdf8;box-shadow:0 4px 6px rgba(0,0,0,.2);}
.hdr h1{color:#f8fafc;font-size:1.6rem;font-weight:700;margin:0;}
.hdr .cl{color:#f59e0b;font-size:.88rem;font-weight:600;margin:4px 0 0;}
.hdr .sb{color:#38bdf8;font-size:.78rem;margin:2px 0 0;letter-spacing:.05em;font-weight:500;}
.sec{color:#38bdf8;font-weight:700;font-size:1rem;border-bottom:2px solid #1e293b;
     padding-bottom:6px;margin-bottom:16px;margin-top:10px;}
.kpi-row{display:flex;gap:10px;flex-wrap:wrap;margin:12px 0;}
.kpi{padding:10px 16px;border-radius:8px;font-size:.85rem;font-weight:600;
     box-shadow:0 1px 3px rgba(0,0,0,.1);color:#ffffff !important;}
.kb{background:#1e3a8a;border-left:4px solid #3b82f6;}
.kg{background:#064e3b;border-left:4px solid #10b981;}
.kr{background:#7f1d1d;border-left:4px solid #ef4444;}
.ka{background:#78350f;border-left:4px solid #f59e0b;}
.kp{background:#581c87;border-left:4px solid #8b5cf6;}
.ke {background:#0f766e;border-left:4px solid #14b8a6;}
.auto-box{background:#0f172a;border:1px solid #334155;border-radius:8px;
          padding:12px;font-size:.85rem;color:#cbd5e1;margin:6px 0;}
.footer{text-align:center;padding:15px;color:#94a3b8;font-size:.75rem;
        border-top:1px solid #334155;margin-top:30px;font-weight:500;}
p,span,label,th,td{color:#cbd5e1 !important;}
.stMarkdown div p strong{color:#ffffff !important;}
div[data-testid="stExpander"]{background-color:#0f172a !important;
     border:1px solid #334155 !important;border-radius:8px;}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hdr">
  <h1>📦 Vertex Packaging — Smart Inventory Tracker</h1>
  <p class="cl">🏢 CEO: Shahzad Bhai — Lahore</p>
</div>
""", unsafe_allow_html=True)

# BUGFIX: _init_schema() (creates tables + seeds the default admin account)
# used to only run lazily inside get_conn(). The login form below only calls
# q(), which never triggered it — so after a fresh reboot, a login attempt
# could run before the admin account was (re)seeded, wrongly showing
# "Invalid username or password" even on a correct first login. Calling it
# explicitly here guarantees schema + seeding always exist before anything
# else runs. st.cache_resource makes this a cheap no-op after the first call.
_init_schema()

# ═══════════════════════════════════════════════
# NEW ADDITION: Trial Version Notice (pure addition, no layout/CSS changed)
# ═══════════════════════════════════════════════
st.warning("⚠️ **Trial Version Notice:** This system is currently running on a **Trial Basis** for testing purposes. "
           "یہ نظام فی الحال ٹیسٹنگ کے مقاصد کے لیے ٹرائل بیس پر چل رہا ہے۔")

# ═══════════════════════════════════════════════
# NEW ADDITION: Website Visit Counter — counts once per new browser session
# (not on every rerun/click), stored in the site_stats table. Shown even
# before login, since a "visit" happens as soon as the page loads.
# ═══════════════════════════════════════════════
if "visit_counted" not in st.session_state:
    _vc_conn = get_conn()
    _vc_conn.execute("UPDATE site_stats SET stat_value = stat_value + 1 WHERE stat_key='total_visits'")
    _vc_conn.commit()
    _vc_conn.close()
    st.session_state["visit_counted"] = True

_total_visits = scalar("SELECT stat_value FROM site_stats WHERE stat_key='total_visits'")
st.caption(f"👁️ Total Visits: {int(_total_visits):,}")

# ═══════════════════════════════════════════════
# AUTHENTICATION & ROLE-BASED ACCESS CONTROL (NEW — Cloud version only)
# ═══════════════════════════════════════════════
if "auth_user" not in st.session_state:
    st.session_state["auth_user"] = None

if st.session_state["auth_user"] is None:
    st.markdown("### 🔐 Login")
    with st.form("login_form"):
        li_user = st.text_input("Username")
        li_pass = st.text_input("Password", type="password")
        li_submit = st.form_submit_button("Login", type="primary")
    if li_submit:
        urow = q("SELECT username, password_hash, role, full_name FROM app_users WHERE username=?", [li_user.strip()])
        if not urow.empty and _verify_password(li_pass, urow.iloc[0]["password_hash"]):
            st.session_state["auth_user"] = {
                "username": urow.iloc[0]["username"],
                "role": urow.iloc[0]["role"],
                "full_name": urow.iloc[0]["full_name"],
            }
            st.rerun()
        else:
            st.error("❌ Invalid username or password.")
    st.info("First time setup? Default login is **admin / admin123** — please change it immediately "
            "from the 👤 User Management tab after logging in.")
    st.stop()

current_user = st.session_state["auth_user"]
current_role = current_user["role"]

# NEW ADDITION: update this user's last_seen timestamp on every rerun, and
# build the Active Users list. Admins are ALWAYS excluded from this list —
# per the strict requirement, nobody (including other viewers) should be
# able to tell an Admin is online.
_now_ts = str(datetime.datetime.now())
_conn_ls = get_conn()
_conn_ls.execute("UPDATE app_users SET last_seen=? WHERE username=?", (_now_ts, current_user["username"]))
_conn_ls.commit()
_conn_ls.close()

_ACTIVE_WINDOW_MINUTES = 5
_df_active = q("SELECT username, role, full_name, last_seen FROM app_users "
               "WHERE last_seen IS NOT NULL AND last_seen != '' AND role != 'Admin'")
_active_list = []
_now_dt = datetime.datetime.now()
for _, _r in _df_active.iterrows():
    try:
        _ls = datetime.datetime.fromisoformat(_r["last_seen"])
        if (_now_dt - _ls).total_seconds() <= _ACTIVE_WINDOW_MINUTES * 60:
            _active_list.append(f"{_r['full_name']} ({_r['role']})")
    except Exception:
        pass

top_c1, top_c2 = st.columns([5, 1])
with top_c1:
    st.caption(f"👋 Logged in as **{current_user['full_name']}** ({current_user['username']}) — Role: **{current_role}**")
    if _active_list:
        st.caption("🟢 Active now: " + ", ".join(_active_list))
with top_c2:
    if st.button("🚪 Logout", key="logout_btn"):
        st.session_state["auth_user"] = None
        st.rerun()

# Role → allowed tab labels. Tabs the current role can't access still render
# in the tab bar (Streamlit limitation: tabs can't be created conditionally
# per-user without breaking layout) but show a 🔒 access-denied message
# instead of their real content.
TAB_ACCESS = {
    "🔍 Global Search":    ["Admin", "Data Entry", "CEO"],
    "➕ DC Entry":         ["Admin", "Data Entry", "CEO"],
    "📋 All Entries":      ["Admin", "Data Entry"],
    "📊 Master Ledger":    ["Admin", "Data Entry", "Viewer", "CEO"],
    "📤 Sheet Upload":     ["Admin"],
    "🚚 Bilty Management": ["Admin", "Data Entry", "CEO"],
    "👤 User Management":  ["Admin"],
    # NEW ADDITION: Rider only ever has access to this one tab — every other
    # tab above will render its 🔒 access-denied message for them via
    # _access_ok(), which is the existing, established pattern in this app
    # for restricting a role to a subset of tabs without breaking the
    # Streamlit tab-bar layout (tabs can't be created conditionally).
    "💸 Daily Expenses":   ["Admin", "CEO", "Rider", "Data Entry"],
}

# Roles allowed to actually SAVE a new DC Entry (vs. just viewing the tab /
# the live Bilty indicator). CEO can see everything in this tab but the
# Save button is hidden for them — view-only, as requested.
DC_ENTRY_WRITE_ROLES = ["Admin", "Data Entry"]

def _access_ok(tab_label):
    if current_role not in TAB_ACCESS[tab_label]:
        st.warning(f"🔒 Your role (**{current_role}**) does not have access to this tab. Contact an Admin if you need it.")
        return False
    return True

if "inline_edit_id" not in st.session_state:
    st.session_state["inline_edit_id"] = None

tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
    "🔍 Global Search","➕ DC Entry","📋 All Entries",
    "📊 Master Ledger","📤 Sheet Upload","🚚 Bilty Management","👤 User Management",
    "💸 Daily Expenses"
])

# ═══════════════════════════════════════════════
# TAB 1 — GLOBAL SEARCH (UNCHANGED)
# ═══════════════════════════════════════════════
with tab1:
    if _access_ok("🔍 Global Search"):
        st.markdown('<div class="sec">🔍 Global Search — Contract / PO / Article / DC / Token</div>', unsafe_allow_html=True)
        gs = st.text_input("🔎 Type to search:", placeholder="Contract #, PO, Article, DC, Token...", key="global_search_main")

        if gs and len(gs.strip()) >= 2:
            gq = f"%{gs.strip()}%"
            # PERFORMANCE: these used to be 6 separate round-trips to the DB
            # on every keystroke. Combined into 1 query (each branch wrapped
            # as its own subquery so per-branch LIMIT works portably on
            # SQLite, PostgreSQL and MySQL alike).
            sug_sql = """
                SELECT * FROM (SELECT 'PO No.'    AS t, po_no          AS val FROM sheet_orders WHERE po_no LIKE ? AND TRIM(po_no)!='' LIMIT 4) x1
                UNION ALL
                SELECT * FROM (SELECT 'Article'   AS t, article        AS val FROM sheet_orders WHERE article LIKE ? LIMIT 4) x2
                UNION ALL
                SELECT * FROM (SELECT 'DC No.'    AS t, dc_no          AS val FROM inventory WHERE dc_no LIKE ? AND TRIM(dc_no)!='' LIMIT 4) x3
                UNION ALL
                SELECT * FROM (SELECT 'Token'     AS t, company_token  AS val FROM inventory WHERE company_token LIKE ? AND TRIM(company_token)!='' LIMIT 4) x4
                UNION ALL
                SELECT * FROM (SELECT 'Call-Off'  AS t, call_off_no    AS val FROM sheet_orders WHERE call_off_no LIKE ? LIMIT 4) x5
                UNION ALL
                SELECT * FROM (SELECT 'Contract #' AS t, contract_no  AS val FROM inventory WHERE contract_no LIKE ? AND TRIM(contract_no)!='' LIMIT 4) x6
            """
            sug = q(sug_sql, [gq, gq, gq, gq, gq, gq]).dropna()
            sug = sug[sug["val"].astype(str).str.strip() != ""]

            if not sug.empty:
                st.markdown("**Quick select:**")
                cols_s = st.columns(min(5, len(sug)))
                for i, (_, sg) in enumerate(sug.iterrows()):
                    with cols_s[i % 5]:
                        if st.button(f"[{sg['t']}] {sg['val']}", key=f"sug_{i}_{sg['val']}"):
                            st.session_state["gs_sel"] = str(sg["val"])
                            st.rerun()

            active = st.session_state.get("gs_sel","") or gs.strip()
            aq = f"%{active}%"

            df_ord = q("""
                SELECT call_off_no,po_no,brand,article,category,SUM(order_qty) as ordered
                FROM sheet_orders
                WHERE po_no LIKE ? OR article LIKE ? OR call_off_no LIKE ? OR sale_contract LIKE ?
                GROUP BY call_off_no,po_no,brand,article,category
            """, [aq,aq,aq,aq])

            df_inv = q("""
                SELECT id,dc_no,company_token,contract_no,call_off_no,po_no,
                       article,category,qty,entry_date,remark
                FROM inventory
                WHERE po_no LIKE ? OR article LIKE ? OR dc_no LIKE ?
                   OR call_off_no LIKE ? OR company_token LIKE ? OR contract_no LIKE ?
                ORDER BY entry_date DESC
            """, [aq,aq,aq,aq,aq,aq])

            tot_o = df_ord["ordered"].sum() if not df_ord.empty else 0
            tot_r = df_inv["qty"].sum()      if not df_inv.empty else 0
            tot_b = tot_o - tot_r

            st.markdown(f"""<div class="kpi-row">
              <div class="kpi kb">📦 Ordered: {tot_o:,.0f}</div>
              <div class="kpi kg">✅ Received: {tot_r:,.0f}</div>
              <div class="kpi {'kr' if tot_b<0 else 'ka'}">⚖️ Balance: {tot_b:,.0f}</div>
              <div class="kpi kp">🔢 DC Entries: {len(df_inv)}</div>
            </div>""", unsafe_allow_html=True)

            if not df_ord.empty:
                df_it = df_ord.groupby("category")["ordered"].sum().reset_index()
                df_ir = df_inv.groupby("category")["qty"].sum().reset_index() if not df_inv.empty \
                        else pd.DataFrame(columns=["category","qty"])
                df_it = df_it.merge(df_ir, on="category", how="left")
                df_it["qty"]     = df_it["qty"].fillna(0)
                df_it["Balance"] = df_it["ordered"] - df_it["qty"]
                df_it.columns    = ["Item Type","Ordered","Received","Balance"]
                st.subheader("📊 Balance by Item Type")
                st.dataframe(df_it, width='stretch', hide_index=True)

            if not df_inv.empty:
                st.subheader(f"📋 DC Entry History ({len(df_inv)} records)")
                disp = df_inv.rename(columns={
                    "dc_no":"DC No.","company_token":"Company Token",
                    "contract_no":"Contract #","call_off_no":"Call-Off",
                    "po_no":"PO No.","article":"Article No.",
                    "category":"Item Type","qty":"Qty",
                    "entry_date":"Date","remark":"Remark"
                }).drop(columns=["id"])
                st.dataframe(disp, width='stretch', hide_index=True)
            else:
                st.info("No DC entries found.")

            # ═══════════════════════════════════════════════
            # NEW ADDITION: Bilty Linkage — pure addition, does not alter any
            # existing search logic above. Shows Bilty (Lahore→Karachi dispatch)
            # records matching the same Call-Off / Contract / Article search term,
            # so a Contract # or Article search also surfaces its dispatch history.
            # ═══════════════════════════════════════════════
            df_bilty_search = q("""
                SELECT bilty_date AS "Date", call_off_no AS "Call-Off", contract_no AS "Contract #",
                       article AS "Article", category AS "Item Type", qty AS "Qty",
                       cartons AS "Cartons", transport_mode AS "Transport"
                FROM bilty
                WHERE call_off_no LIKE ? OR contract_no LIKE ? OR article LIKE ?
                ORDER BY id DESC
            """, [aq, aq, aq])

            if not df_bilty_search.empty:
                st.subheader(f"🚚 Bilty Dispatch History ({len(df_bilty_search)} records)")
                st.dataframe(df_bilty_search, width='stretch', hide_index=True)

            if st.button("🔄 Clear Search", key="clear_gs"):
                st.session_state.pop("gs_sel", None)
                st.rerun()

# TAB 2 — DC ENTRY (UNCHANGED)
# ═══════════════════════════════════════════════
with tab2:
    if _access_ok("➕ DC Entry"):
        st.markdown('<div class="sec">➕ New DC Entry — Call-Off Triggered Auto-Load</div>', unsafe_allow_html=True)
    
        dc_main_cols = st.columns([3, 2])
    
        with dc_main_cols[0]:
            st.markdown("**Step 1 — Type Call-Off No. to auto-load Contract, PO & Articles**")
            c_sc1, c_sc2, info_col = st.columns([1.5, 1.5, 2])

            with c_sc1:
                f_coff = st.text_input("Call-Off No. *", placeholder="e.g. 288", key="dc_coff_input").strip()

            f_po, f_contract, brand = "", "", ""
            contracts_for_coff, po_for_sc, art_list = [], [], []

            if f_coff:
                # Retrieve the contracts and articles using safe helpers
                contracts_for_coff = get_contracts_for_calloff(f_coff)
                if contracts_for_coff:
                    with c_sc2:
                        if len(contracts_for_coff) == 1:
                            f_contract = contracts_for_coff[0]
                            st.text_input("Contract # (Auto-loaded)", value=f_contract, disabled=True, key="dc_cont_ro")
                        else:
                            f_contract = st.selectbox("Select Contract # *", contracts_for_coff, key="dc_cont_sel")
                
                    # PERFORMANCE FIX: cached (see get_calloff_brand_po) — was
                    # 2 fresh round-trips on every rerun.
                    brand, po_for_sc = get_calloff_brand_po(f_coff, f_contract)

                    # Keep an article selectable until every item category
                    # ordered for it has been saved.  A single saved Safety
                    # or Washing Paper row must not hide Tag/Inlay Card.
                    # PERFORMANCE FIX: now cached (see _get_pending_articles) —
                    # this was the heaviest uncached query in the tab.
                    art_list = _get_pending_articles(f_coff, f_contract)
                else:
                    # Fallback when there are no contracts
                    f_contract = ""  # Safeguard downstream code from NameError
                    brand, po_for_sc = get_calloff_brand_po(f_coff)

            with info_col:
                if f_coff and (contracts_for_coff or po_for_sc):
                    po_display = " | ".join(po_for_sc) if po_for_sc else "—"
                    brand_display = brand if brand else "—"
                    st.markdown(f"""
                    <div class="auto-box" style="background:#064e3b;border:1px solid #10b981;color:#fff; padding:8px; font-size:11px;">
                      ✅ <b>Call-Off Verified!</b><br>
                      Brand: <b>{brand_display}</b><br>
                      Contract #: <b>{f_contract or '—'}</b> | PO: <b>{po_display}</b>
                    </div>""", unsafe_allow_html=True)
                    if len(po_for_sc) == 1:
                        f_po = po_for_sc[0]
                    elif len(po_for_sc) > 1:
                        f_po = st.selectbox("Select PO No. *", po_for_sc, key="dc_po_sel")
                elif f_coff:
                    st.markdown("""
                    <div class="auto-box" style="background:#7f1d1d;border:1px solid #ef4444;color:#fff; padding:6px;">
                      ⚠️ <b>Call-Off Not Found!</b>
                    </div>""", unsafe_allow_html=True)

            st.markdown(f"**Step 2 — Select Article & Enter DC Details**")
            art_opts = ["-- Select Article --"] + art_list
            c1, c2, c3 = st.columns(3)

            with c1:
                f_art_sel = st.selectbox("Article No. *", art_opts, key="dc_art", disabled=(not bool(art_list)))
                f_art  = "" if f_art_sel == "-- Select Article --" else f_art_sel

                # NEW: Live Bilty (Lahore factory dispatch) indicator — shows as
                # soon as Call-Off + Article are both selected. This is informational
                # only and reads from the separate `bilty` ledger; it does not
                # touch the existing Ordered/Received calculations below.
                if f_coff and f_art:
                    bilty_done_art = get_bilty_qty(f_coff, f_art)
                    ordered_for_art = get_total_ordered_for_article(f_coff, f_art)
                    rb = int(math.floor(bilty_done_art + 0.5))
                    ro = int(math.floor(ordered_for_art + 0.5))
                    pct_txt = f" ({(rb/ro*100):.0f}%)" if ro > 0 else ""

                    # NEW ADDITION: item-wise breakdown for this article, sourced
                    # ONLY from the bilty table (historical dispatch records) —
                    # rendered inside this exact same blue box.
                    df_bilty_breakdown = get_bilty_breakdown(f_coff, f_art)
                    breakdown_html = ""
                    if not df_bilty_breakdown.empty:
                        rows_html = "".join(
                            f'<div style="padding-left:14px;">• {r["category"]}: <b>{int(math.floor(r["tot"] + 0.5)):,} Pcs</b></div>'
                            for _, r in df_bilty_breakdown.iterrows()
                        )
                        breakdown_html = f'<div style="margin-top:5px;font-size:11px;">{rows_html}</div>'

                    st.markdown(f"""
                    <div class="auto-box" style="background:#0c4a6e;border:1px solid #38bdf8;color:#fff;padding:7px;font-size:11.5px;margin-bottom:6px;">
                      🚚 <b>Total Bilty Done:</b> {rb:,} Pcs out of {ro:,} Pcs{pct_txt} <span style="opacity:.8;">(Lahore ➜ Karachi, Article {f_art})</span>
                      {breakdown_html}
                    </div>""", unsafe_allow_html=True)

                f_type = st.selectbox("Item Type *", ITEM_TYPES, key="dc_type")
            
                f_style = "—"
                if f_type == "Inlay Card / Bandrolle":
                    f_style = st.selectbox("Style Type *", STYLES_INLAY, key="dc_style_inlay")
                
                f_token = st.text_input("Company Token", placeholder="e.g. TOK-771", key="dc_token")
            with c2:
                f_dc   = st.text_input("DC No. *", key="dc_dcno")
                f_date = st.date_input("Entry Date *", value=date.today(), key="dc_date")
                # NEW ADDITION: multi-company DC letterhead selector. Backend
                # formulas/logic are identical for both brands — this only
                # decides which header/logo prints on the DC PDF.
                f_company = st.selectbox("🏢 Company Header (DC Type) *", COMPANY_TYPES, key="dc_company")
            with c3:
                # ═══════════════════════════════════════════════
                # NEW ADDITION: Dual-Pack (Jersey/Molton) split entry.
                # Auto-detected from variant-tagged sheet_orders rows
                # (populated by the new Packaging Detail parser in Sheet
                # Upload). When a Contract has BOTH Jersey and Molton
                # articles for the selected Item Type, two separate
                # article+qty inputs appear instead of the normal single
                # Quantity field below — matching how your actual DC sheets
                # (e.g. DC-4670, Cont #232603624) list Jersey and Molton as
                # separate line items. When not dual-pack, nothing changes.
                # ═══════════════════════════════════════════════
                is_dual_pack = False
                dp_jersey_articles, dp_molton_articles = [], []
                if f_coff and f_contract:
                    dp_jersey_articles, dp_molton_articles = get_dual_pack_articles(f_coff, f_contract, f_type)
                    is_dual_pack = bool(dp_jersey_articles) and bool(dp_molton_articles)

                if is_dual_pack:
                    st.caption("🧵 Dual-Pack Contract — enter Jersey & Molton separately:")
                    dp_j_art = st.selectbox("Jersey Article", dp_jersey_articles, key="dc_dp_j_art")
                    autofill_item_description("dc_desc_jersey", "dc_desc_jersey_source", dp_j_art, f_type)
                    f_desc_jersey = st.text_input(
                        "Jersey Item Description (as per PO)",
                        key="dc_desc_jersey",
                        placeholder="e.g. INLAY CARD FITTED 36.5X40.5 CM- DIXX JERSEY")
                    f_qty_jersey = st.number_input("Jersey Qty (Pcs)", min_value=0.0, step=1.0, format="%g", key="dc_qty_jersey")
                    dp_m_art = st.selectbox("Molton Article", dp_molton_articles, key="dc_dp_m_art")
                    autofill_item_description("dc_desc_molton", "dc_desc_molton_source", dp_m_art, f_type)
                    f_desc_molton = st.text_input(
                        "Molton Item Description (as per PO)",
                        key="dc_desc_molton",
                        placeholder="e.g. INLAY CARD FITTED 19.29X14.37 CM- DIXX MOLTON")
                    f_qty_molton = st.number_input("Molton Qty (Pcs)", min_value=0.0, step=1.0, format="%g", key="dc_qty_molton")
                    f_qty = 0.0  # existing single-flow field unused in dual-pack mode
                else:
                    autofill_item_description("dc_desc", "dc_desc_source", f_art, f_type)

                    # NEW ADDITION: Combo Article quality picker. Once TWO or
                    # more distinct descriptions have ever been saved for this
                    # exact Article+Item Type (e.g. one "...MOLTEN+LYCRA" line
                    # and one plain "...MOLTEN" line), they're offered here as
                    # a dropdown instead of forcing a retype/copy-paste. Pick
                    # one to reuse it verbatim, or pick "Type New" to enter a
                    # third variant — either way it still saves into the same
                    # history for next time.
                    past_descs = get_cached_item_descriptions_list(f_art, f_type) if f_art else []
                    NEW_DESC_OPTION = "✏️ Type New / Custom Description"
                    if len(past_descs) >= 2:
                        picker_choice = st.selectbox(
                            "🔗 Combo Article — Select Quality / Description",
                            past_descs + [NEW_DESC_OPTION],
                            key="dc_desc_picker")
                        if picker_choice != NEW_DESC_OPTION:
                            st.session_state["dc_desc"] = picker_choice

                    f_desc = st.text_input(
                        "Item Description (as per PO)",
                        key="dc_desc",
                        placeholder="e.g. SAFETY STICKER TRANSPARENT (5X1.5 CM) - BH")
                    # HOTFIX: Combo Article detection — one article number
                    # carrying multiple distinct qualities (e.g. "MOLTEN+LYCRA"),
                    # each needing its own separate Inlay Card / Tag Card entry
                    # under that same article. Detected from a "+" in the typed
                    # description — how these are written on the source PO/DC
                    # sheets — and "clearly highlighted" per spec so it's obvious
                    # this entry is being treated as a combo before you save.
                    is_combo_article = "+" in str(f_desc).upper()
                    if is_combo_article:
                        st.caption("🔗 Combo Article detected (combined quality) — over-access check relaxed for this entry.")
                    f_qty = st.number_input("Quantity (Pcs) *", min_value=0.0, step=1.0, format="%g", key="dc_qty")
                    f_qty_jersey = f_qty_molton = 0.0
                    f_desc_jersey = f_desc_molton = ""
                    dp_j_art = dp_m_art = ""

                f_remark = st.text_area("Remark / Notes", height=90, key="dc_remark")
                f_destination = st.text_input("Destination", key="dc_destination", placeholder="e.g. SOHRAB/HSU")

        with dc_main_cols[1]:
            # HOTFIX (per client decision): the Live Contract Status Counter
            # panel — heading, grid, and the "Check Contract Status" button —
            # has been removed entirely, not just hidden behind a click.
            #
            # max_allowed is still computed here, silently, with NO UI: the
            # Save button's over-delivery validation further down depends on
            # it, and removing it would silently disable that safety check.
            # This is a single cached lookup (get_ordered_for_article_contract
            # / get_received_for_article_contract), not the 7-category grid
            # that was actually causing the lag — so keeping it does not
            # reintroduce any of the slowness that was removed.
            max_allowed = 0
            if f_coff and f_contract:
                max_allowed = get_ordered_for_article_contract(f_coff, f_contract, f_type, f_art) - get_received_for_article_contract(f_coff, f_contract, f_type, f_art)

        if f_art and f_type:
            o_qty = get_ordered_qty(f_art, f_type, coff=f_coff or None, po=f_po or None)
            r_qty = get_received_qty(f_art, f_type, coff=f_coff or None, po=f_po or None)
        
            rounded_o_qty = int(math.floor(o_qty + 0.5))
            rounded_r_qty = int(math.floor(r_qty + 0.5))
            pend  = rounded_o_qty - rounded_r_qty
            after = pend - (int(f_qty) if f_qty > 0 else 0)
            clr   = "#ef4444" if after < 0 else ("#10b981" if after == 0 else "#f59e0b")
            sign  = "⚠️ Over" if after < 0 else ("✅ Done" if after == 0 else "⏳ Remaining")
            bp1,bp2,bp3,bp4 = st.columns(4)
            bp1.markdown(f"<div class='kpi kb'>📦 Ordered<br><b>{rounded_o_qty:,}</b></div>", unsafe_allow_html=True)
            bp2.markdown(f"<div class='kpi kg'>✅ Received<br><b>{rounded_r_qty:,}</b></div>", unsafe_allow_html=True)
            bp3.markdown(f"<div class='kpi ka'>⏳ Pending<br><b>{pend:,}</b></div>", unsafe_allow_html=True)
            bp4.markdown(f"<div class='kpi' style='background:#0f172a;color:{clr}'>{sign}<br><b>{after:+,}</b></div>", unsafe_allow_html=True)

        st.markdown("---")
        if current_role not in DC_ENTRY_WRITE_ROLES:
            st.info(f"🔒 Your role (**{current_role}**) has view-only access to DC Entry — you can see live stock/Bilty status above, "
                    "but cannot add new entries. Contact an Admin or Data Entry user if a new DC needs to be recorded.")
        elif is_dual_pack and st.button("💾 Save Dual-Pack Entry", type="primary", key="dc_save_dp"):
            # NEW ADDITION: Dual-Pack save — writes up to two separate
            # inventory rows (one per variant), each validated against its
            # OWN article's remaining balance, so over-delivery blocking
            # still applies correctly per article. Existing single-flow
            # save logic below is completely untouched.
            s_dc, s_po, s_coff, s_cont = str(f_dc).strip(), str(f_po).strip(), str(f_coff).strip(), str(f_contract).strip()
            if not s_dc or not s_po or not s_coff or (f_qty_jersey <= 0 and f_qty_molton <= 0):
                st.error("⚠️ DC No., Call-Off, PO and at least one of Jersey/Molton Quantity are required.")
            else:
                errors = []
                if f_qty_jersey > 0:
                    max_j = get_ordered_qty(dp_j_art, f_type, coff=s_coff) - get_received_qty(dp_j_art, f_type, coff=s_coff)
                    if int(f_qty_jersey) > int(math.floor(max_j + 0.5)) and max_j >= 0:
                        errors.append(f"Jersey: max allowed is {int(math.floor(max_j + 0.5)):,} Pcs for article {dp_j_art}")
                if f_qty_molton > 0:
                    max_m = get_ordered_qty(dp_m_art, f_type, coff=s_coff) - get_received_qty(dp_m_art, f_type, coff=s_coff)
                    if int(f_qty_molton) > int(math.floor(max_m + 0.5)) and max_m >= 0:
                        errors.append(f"Molton: max allowed is {int(math.floor(max_m + 0.5)):,} Pcs for article {dp_m_art}")

                if errors:
                    st.error("🚨 ALERT! Over-delivery blocked. " + " | ".join(errors))
                else:
                    conn = get_conn()
                    saved_parts = []
                    if f_qty_jersey > 0:
                        conn.execute("""
                            INSERT INTO inventory
                            (call_off_no,contract_no,dc_no,po_no,article,category,qty,entry_date,remark,company_token,style_type,item_description,destination,company_type)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """, (s_coff, s_cont, s_dc, s_po, str(dp_j_art), f_type, float(f_qty_jersey),
                              str(f_date), f"[Jersey] {str(f_remark).strip()}".strip(), str(f_token).strip(), f_style, str(f_desc_jersey).strip(), str(f_destination).strip(), f_company))
                        saved_parts.append(f"Jersey {f_qty_jersey:,.0f} Pcs (Art. {dp_j_art})")
                    if f_qty_molton > 0:
                        conn.execute("""
                            INSERT INTO inventory
                            (call_off_no,contract_no,dc_no,po_no,article,category,qty,entry_date,remark,company_token,style_type,item_description,destination,company_type)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """, (s_coff, s_cont, s_dc, s_po, str(dp_m_art), f_type, float(f_qty_molton),
                              str(f_date), f"[Molton] {str(f_remark).strip()}".strip(), str(f_token).strip(), f_style, str(f_desc_molton).strip(), str(f_destination).strip(), f_company))
                        saved_parts.append(f"Molton {f_qty_molton:,.0f} Pcs (Art. {dp_m_art})")
                    conn.commit()
                    conn.close()
                    save_cached_item_description(dp_j_art, f_type, f_desc_jersey)
                    save_cached_item_description(dp_m_art, f_type, f_desc_molton)
                    _clear_dc_entry_caches()
                    st.success(f"✅ Dual-Pack Entry saved — DC {s_dc} | " + " + ".join(saved_parts))
                    st.session_state["last_saved_dc"] = s_dc
                    st.rerun()
        elif not is_dual_pack and st.button("💾 Save Entry", type="primary", key="dc_save"):
            s_dc   = str(f_dc).strip()
            s_po   = str(f_po).strip()
            s_coff = str(f_coff).strip()
            s_cont = str(f_contract).strip()
            s_art  = str(f_art).strip()
        
            rounded_max_allowed = int(math.floor(max_allowed + 0.5))
        
            if not s_dc or not s_po or not s_coff or f_qty <= 0 or not s_art:
                st.error("⚠️ DC No., Call-Off, PO, Article and Quantity are required.")
            elif int(f_qty) > rounded_max_allowed and max_allowed >= 0 and not is_combo_article:
                st.error(f"🚨 ALERT! Over-delivery blocked. Max remaining allowed for {f_type} is exactly {rounded_max_allowed:,} Pcs. You cannot enter {int(f_qty):,} Pcs.")
            else:
                conn = get_conn()
                conn.execute("""
                    INSERT INTO inventory
                    (call_off_no,contract_no,dc_no,po_no,article,category,qty,entry_date,remark,company_token,style_type,item_description,destination,company_type)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (s_coff, s_cont, s_dc, s_po, s_art, f_type, float(f_qty),
                      str(f_date), str(f_remark).strip(), str(f_token).strip(), f_style, str(f_desc).strip(), str(f_destination).strip(), f_company))
                conn.commit()
                conn.close()
                save_cached_item_description(s_art, f_type, f_desc)
                _clear_dc_entry_caches()
                st.success(f"✅ Entry saved — DC {s_dc} | {f_qty:,.0f} pcs of {f_type}")
                st.session_state["last_saved_dc"] = s_dc
                st.rerun()

        # ═══════════════════════════════════════════════
        # NEW ADDITION: Live Print Preview + Export, directly inside the DC
        # Entry tab (avoids Chrome's pop-up blocker entirely — the preview
        # iframe and Print button share one sandbox, no new window/tab is
        # ever opened). Defaults to the DC you just saved.
        # ═══════════════════════════════════════════════
        st.markdown("---")
        st.markdown('<div class="sec">🖨️ Print Preview & Export (this DC)</div>', unsafe_allow_html=True)
        render_ditto_dc_section(key_prefix="dcentry", default_dc=st.session_state.get("last_saved_dc"))

# ═══════════════════════════════════════════════
# TAB 3 — ALL ENTRIES (UNCHANGED)
# ═══════════════════════════════════════════════
with tab3:
    if _access_ok("📋 All Entries"):
        st.markdown('<div class="sec">📋 Registered DC Entries Registry</div>', unsafe_allow_html=True)

        # PERFORMANCE FIX: previously this loaded the ENTIRE inventory table into
        # a DataFrame on every rerun and did the filtering in pandas — slow and
        # growing slower with every new DC entry. Now a cheap COUNT checks if any
        # data exists, filter dropdown options come from lightweight indexed
        # DISTINCT queries, and the actual filtering happens in SQL (using the
        # indexes created in get_conn) so only the matching rows are ever loaded.
        total_inv_count = scalar("SELECT COUNT(*) FROM inventory")

        if total_inv_count == 0:
            st.info("No DC entries available yet.")
        else:
            st.markdown("##### 🛠️ Advanced Multi-Filters")
            f_cols = st.columns(6)

            # PERFORMANCE: on a remote DB, 6 separate round-trips (one per
            # dropdown) are much slower than 1. This combines them into a
            # single UNION ALL query, cached briefly so repeated reruns
            # (e.g. while typing elsewhere on the page) don't re-hit the DB.
            @st.cache_data(ttl=15, show_spinner=False)
            def _tab3_filter_options():
                cols = ["call_off_no", "po_no", "article", "dc_no", "company_token", "category"]
                union_sql = " UNION ALL ".join(
                    f"SELECT '{c}' AS col_name, {c} AS val FROM inventory WHERE TRIM({c})!=''" for c in cols
                )
                df = q(union_sql)
                return {c: sorted(df.loc[df["col_name"] == c, "val"].dropna().unique().tolist()) for c in cols}

            _opts = _tab3_filter_options()

            with f_cols[0]:
                sel_coff = st.selectbox("Filter Call-Off", ["All"] + _opts["call_off_no"], key="f3_coff")
            with f_cols[1]:
                sel_po = st.selectbox("Filter PO No.", ["All"] + _opts["po_no"], key="f3_po")
            with f_cols[2]:
                sel_art = st.selectbox("Filter Article", ["All"] + _opts["article"], key="f3_art")
            with f_cols[3]:
                sel_dc = st.selectbox("Filter DC No.", ["All"] + _opts["dc_no"], key="f3_dc")
            with f_cols[4]:
                sel_tok = st.selectbox("Filter Token", ["All"] + _opts["company_token"], key="f3_tok")
            with f_cols[5]:
                sel_cat = st.selectbox("Filter Item Type", ["All"] + _opts["category"], key="f3_cat")


            where_sql, params = "WHERE 1=1", []
            if sel_coff != "All": where_sql += " AND call_off_no=?";    params.append(sel_coff)
            if sel_po != "All":   where_sql += " AND po_no=?";          params.append(sel_po)
            if sel_art != "All":  where_sql += " AND article=?";        params.append(sel_art)
            if sel_dc != "All":   where_sql += " AND dc_no=?";          params.append(sel_dc)
            if sel_tok != "All":  where_sql += " AND company_token=?";  params.append(sel_tok)
            if sel_cat != "All":  where_sql += " AND category=?";       params.append(sel_cat)

            df_inv = q(f"""
                SELECT id,call_off_no,contract_no,dc_no,company_token,po_no,
                       article,category,style_type,qty,entry_date,remark
                FROM inventory {where_sql}
                ORDER BY entry_date DESC, id DESC
            """, params)

            tot_r3 = df_inv["qty"].sum()
            tot_o3 = scalar("SELECT SUM(order_qty) FROM sheet_orders")
            net3   = tot_o3 - tot_r3
        
            st.markdown(f"""<div class="kpi-row">
              <div class="kpi kb">📦 Total Ordered: {tot_o3:,.0f}</div>
              <div class="kpi kg">✅ Current Filtered Received: {tot_r3:,.0f}</div>
              <div class="kpi {'kr' if net3<0 else 'ka'}">⚖️ Net Balance: {net3:,.0f}</div>
              <div class="kpi kp">🔢 Displayed Entries: {len(df_inv)}</div>
            </div>""", unsafe_allow_html=True)

            df_show = df_inv.copy()
            df_show.columns = ["ID","Call-Off","Contract #","DC No.","Token","PO No.",
                               "Article","Item Type","Style","Qty","Date","Remark"]
            st.dataframe(df_show, width='stretch', hide_index=True)

            st.markdown("---")
            st.markdown("### 🛠️ Edit / Delete Entries")

            for _, row in df_inv.iterrows():
                with st.expander(f"📦 ID:{row['id']} | DC:{row['dc_no']} | Token:{row['company_token'] or '—'} | Article:{row['article']} | {row['qty']:.0f} Pcs"):
                    if st.session_state["inline_edit_id"] == row["id"]:
                        st.markdown(f"#### ✏️ Editing Record ID: {row['id']}")
                        ec1,ec2,ec3 = st.columns(3)
                        with ec1:
                            e_coff  = st.text_input("Call-Off No.", value=str(row["call_off_no"] or ""), key=f"e_coff_{row['id']}")
                            e_cont  = st.text_input("Contract #",   value=str(row["contract_no"]  or ""), key=f"e_cont_{row['id']}")
                            e_po    = st.text_input("PO No.",        value=str(row["po_no"]        or ""), key=f"e_po_{row['id']}")
                        with ec2:
                            e_art   = st.text_input("Article No.",  value=str(row["article"]       or ""), key=f"e_art_{row['id']}")
                            cat_cur = str(row["category"] or "")
                            e_type  = st.selectbox("Item Type", ITEM_TYPES, index=ITEM_TYPES.index(cat_cur) if cat_cur in ITEM_TYPES else 0, key=f"e_type_{row['id']}")
                            e_dc    = st.text_input("DC No.",       value=str(row["dc_no"]         or ""), key=f"e_dc_{row['id']}")
                        with ec3:
                            e_tok   = st.text_input("Company Token",value=str(row["company_token"] or ""), key=f"e_tok_{row['id']}")
                            try:    dv = datetime.strptime(str(row["entry_date"]),"%Y-%m-%d").date()
                            except: dv = date.today()
                            e_date  = st.date_input("Date", value=dv, key=f"e_date_{row['id']}")
                            e_qty   = st.number_input("Quantity", min_value=0.0, step=1.0, value=float(row["qty"] or 0), format="%g", key=f"e_qty_{row['id']}")
                            e_rem   = st.text_area("Remark", value=str(row["remark"] or ""), height=70, key=f"e_rem_{row['id']}")

                        eb1,eb2,_ = st.columns([1.5,1.5,5])
                        with eb1:
                            if st.button("💾 Update & Save", key=f"save_inline_{row['id']}", type="primary"):
                                conn = get_conn()
                                conn.execute("""
                                    UPDATE inventory SET call_off_no=?,contract_no=?,po_no=?,
                                    article=?,category=?,dc_no=?,entry_date=?,qty=?,
                                    remark=?,company_token=? WHERE id=?
                                """, (e_coff.strip(), e_cont.strip(), e_po.strip(), e_art.strip(), e_type, e_dc.strip(), str(e_date), float(e_qty), e_rem.strip(), e_tok.strip(), row["id"]))
                                conn.commit(); conn.close()

                                # Clear inline-edit UI state so the widgets reload fresh
                                # values from the database on rerun (fixes stale/ghost
                                # quantity showing after Update & Save).
                                st.session_state["inline_edit_id"] = None
                                for _k in (f"e_coff_{row['id']}", f"e_cont_{row['id']}", f"e_po_{row['id']}",
                                           f"e_art_{row['id']}", f"e_type_{row['id']}", f"e_dc_{row['id']}",
                                           f"e_tok_{row['id']}", f"e_date_{row['id']}", f"e_qty_{row['id']}",
                                           f"e_rem_{row['id']}"):
                                    if _k in st.session_state:
                                        del st.session_state[_k]

                                st.success("✅ Updated successfully!")
                                st.rerun()
                        with eb2:
                            if st.button("❌ Cancel", key=f"cancel_inline_{row['id']}"):
                                st.session_state["inline_edit_id"] = None
                                st.rerun()
                    else:
                        ca,cb,_ = st.columns([1.5,1.5,5])
                        with ca:
                            if st.button("✏️ Edit", key=f"edit_{row['id']}", type="secondary"):
                                st.session_state["inline_edit_id"] = row["id"]
                                st.rerun()
                        with cb:
                            if st.button("🗑️ Delete", key=f"del_{row['id']}", type="primary"):
                                st.session_state[f"cdel_{row['id']}"] = True
                        if st.session_state.get(f"cdel_{row['id']}"):
                            st.warning(f"Delete DC **{row['dc_no']}** — {row['qty']:.0f} pcs? Balance will revert.")
                            cy,cn,_ = st.columns([1,1,6])
                            if cy.button("✅ Confirm", key=f"cy_{row['id']}"):
                                conn = get_conn()
                                conn.execute("DELETE FROM inventory WHERE id=?", (row["id"],))
                                conn.commit(); conn.close()
                                st.session_state.pop(f"cdel_{row['id']}", None)
                                st.success("Deleted."); st.rerun()
                            if cn.button("❌ No", key=f"cn_{row['id']}"):
                                st.session_state.pop(f"cdel_{row['id']}", None)
                                st.rerun()

            st.markdown("---")
            st.markdown("### 🗑️ Bulk Delete — Entire DC")
            st.caption("Deletes every line item saved under one DC No. in a single action — for correcting a DC that needs to be fully redone, instead of deleting each line one by one above.")
            bd_c1, bd_c2 = st.columns([3, 2])
            with bd_c1:
                bulk_del_dc = st.text_input("Enter DC Number for Full Deletion", key="bulk_del_dc_input", placeholder="e.g. 4753")
            with bd_c2:
                st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
                bulk_del_clicked = st.button("🗑️ Delete Entire DC", key="bulk_del_dc_btn", type="primary")

            if bulk_del_clicked:
                dc_to_delete = bulk_del_dc.strip()
                if not dc_to_delete:
                    st.error("⚠️ Please enter a DC Number first.")
                else:
                    _bd_count = scalar("SELECT COUNT(*) FROM inventory WHERE dc_no=?", [dc_to_delete])
                    if _bd_count == 0:
                        st.error(f"⚠️ No entries found for DC No. **{dc_to_delete}**.")
                    else:
                        st.session_state["bulk_del_pending_dc"] = dc_to_delete
                        st.session_state["bulk_del_pending_count"] = _bd_count

            # Safety confirmation popup — nothing is deleted until this is
            # explicitly confirmed.
            if st.session_state.get("bulk_del_pending_dc"):
                _pdc = st.session_state["bulk_del_pending_dc"]
                _pct = st.session_state["bulk_del_pending_count"]
                st.warning(f"⚠️ This will permanently delete **{_pct}** line item(s) under DC No. **{_pdc}**. This cannot be undone. Are you sure?")
                bd_cy, bd_cn, _ = st.columns([1.5, 1.5, 5])
                with bd_cy:
                    if st.button("✅ Yes, Delete Entire DC", key="bulk_del_confirm_yes", type="primary"):
                        conn = get_conn()
                        conn.execute("DELETE FROM inventory WHERE dc_no=?", (_pdc,))
                        conn.commit(); conn.close()
                        st.session_state.pop("bulk_del_pending_dc", None)
                        st.session_state.pop("bulk_del_pending_count", None)
                        st.session_state["inline_edit_id"] = None
                        st.success(f"✅ Deleted all {_pct} line item(s) under DC No. {_pdc}.")
                        st.rerun()
                with bd_cn:
                    if st.button("❌ Cancel", key="bulk_del_confirm_no"):
                        st.session_state.pop("bulk_del_pending_dc", None)
                        st.session_state.pop("bulk_del_pending_count", None)
                        st.rerun()

            st.markdown("---")
            st.markdown("### ⚠️ Total System Reset")
            if st.checkbox("Yes, I want to delete EVERYTHING", key="confirm_reset"):
                if st.button("🚨 FULL SOFTWARE RESET", type="primary", key="btn_reset"):
                    conn = get_conn()
                    conn.execute("DELETE FROM inventory")
                    conn.execute("DELETE FROM sheet_orders")
                    conn.commit(); conn.close()
                    st.session_state["inline_edit_id"] = None
                    st.success("System reset complete!")
                    st.rerun()

            st.markdown("---")
            st.markdown('<div class="sec">🖨️ NEW: Generate Ditto DC</div>', unsafe_allow_html=True)
            st.caption("موجودہ DC انٹریز سے، بالکل اصل ڈی سی لے آؤٹ جیسی فائل بنائیں — لائیو پریویو اور پرنٹ بٹن کے ساتھ۔")
            render_ditto_dc_section(key_prefix="allentries")

# ═══════════════════════════════════════════════
# TAB 4 — MASTER LEDGER (UPDATED LOGIC ADDEED!)
# ═══════════════════════════════════════════════
with tab4:
    if _access_ok("📊 Master Ledger"):
        st.markdown('<div class="sec">📊 Master Ledger (Order vs Delivery Status)</div>', unsafe_allow_html=True)

        df_ledger_raw = q("""
            SELECT
                so.call_off_no        AS "Call-Off No",
                so.sale_contract      AS "Contract #",
                so.brand              AS "Brand",
                so.article            AS "Article",
                so.category           AS "Item Type",
                SUM(so.order_qty)     AS "Total Ordered",
                MAX(COALESCE(inv.total_received, 0))                        AS "Total Received",
                SUM(so.order_qty) - MAX(COALESCE(inv.total_received, 0))   AS "Remaining Balance"
            FROM sheet_orders so
            LEFT JOIN (
                SELECT call_off_no, article, category, SUM(qty) AS total_received
                FROM inventory
                GROUP BY call_off_no, article, category
            ) inv ON so.call_off_no = inv.call_off_no
                  AND so.article    = inv.article
                  AND so.category   = inv.category
            GROUP BY so.call_off_no, so.sale_contract, so.brand, so.article, so.category
            ORDER BY so.call_off_no DESC, so.article ASC
        """)

        if df_ledger_raw.empty:
            st.info("Master Ledger is empty. Upload a contract sheet in Tab 5 first.")
        else:
            st.markdown("##### 🛠️ Dynamic Ledger Filters")
            fl_cols = st.columns(5)
        
            with fl_cols[0]:
                l_opt_coff = ["All"] + sorted(df_ledger_raw["Call-Off No"].unique().tolist())
                l_sel_coff = st.selectbox("Ledger Call-Off", l_opt_coff, key="l_coff")
            with fl_cols[1]:
                l_opt_cont = ["All"] + sorted(df_ledger_raw["Contract #"].unique().tolist())
                l_sel_cont = st.selectbox("Ledger Contract", l_opt_cont, key="l_cont")
            with fl_cols[2]:
                l_opt_br = ["All"] + sorted(df_ledger_raw["Brand"].unique().tolist())
                l_sel_br = st.selectbox("Ledger Brand", l_opt_br, key="l_brand")
            with fl_cols[3]:
                l_opt_art = ["All"] + sorted(df_ledger_raw["Article"].unique().tolist())
                l_sel_art = st.selectbox("Ledger Article", l_opt_art, key="l_art")
            with fl_cols[4]:
                l_opt_cat = ["All"] + sorted(df_ledger_raw["Item Type"].unique().tolist())
                l_sel_cat = st.selectbox("Ledger Item Type", l_opt_cat, key="l_cat")

            df_ledger = df_ledger_raw.copy()
            if l_sel_coff != "All": df_ledger = df_ledger[df_ledger["Call-Off No"] == l_sel_coff]
            if l_sel_cont != "All": df_ledger = df_ledger[df_ledger["Contract #"] == l_sel_cont]
            if l_sel_br != "All":   df_ledger = df_ledger[df_ledger["Brand"] == l_sel_br]
            if l_sel_art != "All":  df_ledger = df_ledger[df_ledger["Article"] == l_sel_art]
            if l_sel_cat != "All":  df_ledger = df_ledger[df_ledger["Item Type"] == l_sel_cat]

            tot_o = df_ledger["Total Ordered"].sum()
            tot_r = df_ledger["Total Received"].sum()
            tot_b = df_ledger["Remaining Balance"].sum()

            st.markdown(f"""<div class="kpi-row">
              <div class="kpi kb">📦 Filtered Ordered: {round_and_format(tot_o)}</div>
              <div class="kpi kg">✅ Filtered Received: {round_and_format(tot_r)}</div>
              <div class="kpi {'kr' if tot_b<0 else 'ka'}">⚖️ Filtered Balance: {round_and_format(tot_b)}</div>
            </div>""", unsafe_allow_html=True)

            st.markdown("##### 🏷️ Item-Wise Remaining Status (Filtered Global Summary)")
            item_summary = df_ledger.groupby("Item Type")["Remaining Balance"].sum().to_dict()
        
            color_classes = {
                "Inlay Card / Bandrolle": "kb",
                "Tag Card / Barcode Sticker": "kg",
                "Barcode Item": "kr",
                "Safety": "ka",
                "Washing Paper": "kp",
                "Transparent Sticker": "ke",
                "Eco Friendly": "kb"
            }
        
            item_cols = st.columns(3)
            for idx, it_name in enumerate(ITEM_TYPES):
                rem_val = item_summary.get(it_name, 0)
                cls = color_classes.get(it_name, "kb")
                with item_cols[idx % 3]:
                    st.markdown(f'<div class="kpi {cls}" style="margin-bottom: 6px;">⏳ Rem. {it_name}<br><b>{round_and_format(rem_val)}</b> Pcs</div>', unsafe_allow_html=True)

            # ═══════════════════════════════════════════════
            # 🆕 BUGFIX: ARTICLE-WISE BREAKDOWN — the block above only ever
            # showed a flat GLOBAL total per category (e.g. total Inlay
            # Cards across every article combined). This renders one boxed
            # block PER ARTICLE with its own nested, non-zero category
            # balances underneath — in a dynamic 1/2/3-column layout
            # depending on how many articles are pending.
            # ═══════════════════════════════════════════════
            st.markdown("---")
            st.markdown("##### 🎯 Article-Wise Remaining Breakdown (Detailed)")
            article_blocks = build_article_blocks(df_ledger)
            if not article_blocks:
                st.caption("✅ No pending balance for any article in the current filter selection.")
            else:
                n_articles = len(article_blocks)
                n_cols = 3 if n_articles >= 3 else (2 if n_articles == 2 else 1)
                art_block_cols = st.columns(n_cols)
                for idx, (coff, article, cats) in enumerate(article_blocks):
                    cat_lines = "".join(
                        f'<div style="font-size:12px; margin-top:2px;">📦 {cat}: <b>{bal:,}</b> Pcs</div>'
                        for cat, bal in cats.items()
                    )
                    coff_txt = f' <span style="opacity:.8; font-weight:400;">| Call-Off: {coff}</span>' if coff else ""
                    with art_block_cols[idx % n_cols]:
                        st.markdown(f'''
                        <div class="kpi kb" style="text-align:left; margin-bottom:8px; padding:10px; border-radius:6px;">
                            <span style="font-size:13px; font-weight:700;">🎯 Article: {article}{coff_txt}</span>
                            {cat_lines}
                        </div>
                        ''', unsafe_allow_html=True)

            # ═══════════════════════════════════════════════
            # 🆕 NEW ADDITION: CONTRACT-WISE SHORTFALL SUMMARY
            # ═══════════════════════════════════════════════
            st.markdown("---")
            st.markdown("### 📌 CONTRACT-WISE SHORTFALL SUMMARY")
            st.info("💡 نیچے ہر سیلز کنٹریکٹ کے حساب سے الگ الگ بقایا (Shortfall) بریک ڈاؤن دکھایا گیا ہے:")
        
            # گوبل لیجر سے اس کال آف کے تمام یونیک کنٹریکٹس نکالیں
            unique_contracts_in_ledger = sorted(df_ledger["Contract #"].unique().tolist())
        
            for contract_no in unique_contracts_in_ledger:
                # اس مخصوص کنٹریکٹ کا ڈیٹا فلٹر کریں
                df_contract_sub = df_ledger[df_ledger["Contract #"] == contract_no]
                contract_item_summary = df_contract_sub.groupby("Item Type")["Remaining Balance"].sum().to_dict()
            
                # صرف وہ آئٹمز چیک کریں جن کا شارٹ فال 0 سے زیادہ ہے
                has_shortage = any(v > 0 for v in contract_item_summary.values())
            
                with st.expander(f"📄 Contract: {contract_no} | {'🚨 Shortage Pending' if has_shortage else '✅ Fully Cleared'}", expanded=True):
                    sub_cols = st.columns(3)
                    col_counter = 0
                    for it_name in ITEM_TYPES:
                        rem_val_sub = contract_item_summary.get(it_name, 0)
                    
                        # اگر بقایا 0 سے زیادہ ہے تو ڈبے کا رنگ لال (kr) ہوگا ورنہ ہرا (kg)
                        sub_cls = "kr" if rem_val_sub > 0 else "kg"
                    
                        with sub_cols[col_counter % 3]:
                            st.markdown(f'''
                            <div class="kpi {sub_cls}" style="margin-bottom: 6px; padding: 8px; border-radius: 6px;">
                                <span style="font-size: 11px; font-weight: 600; display:block;">📦 {it_name}</span>
                                <span style="font-size: 13px; font-weight: 700;">Rem: {round_and_format(rem_val_sub)} Pcs</span>
                            </div>
                            ''', unsafe_allow_html=True)
                        col_counter += 1
            # ═══════════════════════════════════════════════

            st.markdown("---")
            hide_zero = st.checkbox("Hide rows with zero Remaining Balance", key="hide_zero_bal")
            if hide_zero:
                df_ledger = df_ledger[df_ledger["Remaining Balance"].apply(round_bal) != 0]

            df_fmt = df_ledger.copy()
            df_fmt["Total Ordered"]     = df_fmt["Total Ordered"].apply(round_and_format)
            df_fmt["Total Received"]    = df_fmt["Total Received"].apply(round_and_format)
            df_fmt["Remaining Balance"] = df_fmt["Remaining Balance"].apply(lambda x: "" if x == 0 else round_and_format(x))

            st.dataframe(df_fmt, width='stretch', hide_index=True)

            st.markdown("##### 🖨️ Export options:")
            btn_c1, btn_c2, btn_c3 = st.columns([2.5, 3, 3.2])
        
            with btn_c1:
                pdf_buf_master = generate_ledger_pdf(item_summary, df_ledger, l_sel_coff, l_sel_cont, l_sel_art, report_type="MASTER")
                st.download_button(
                    label="🖨️ Print Full Master Ledger (PDF)",
                    data=pdf_buf_master,
                    file_name=f"Master_Ledger_CO_{l_sel_coff}_CN_{l_sel_cont}.pdf",
                    mime="application/pdf",
                    type="primary",
                    key="btn_print_master"
                )
            
            with btn_c2:
                df_shortage_only = df_ledger[df_ledger["Remaining Balance"].apply(round_bal) > 0]
                pdf_buf_shortage = generate_ledger_pdf(item_summary, df_shortage_only, l_sel_coff, l_sel_cont, l_sel_art, report_type="SHORTAGE")
                st.download_button(
                    label="🚨 Print Shortage / Remaining Only (PDF)",
                    data=pdf_buf_shortage,
                    file_name=f"Shortage_Report_CO_{l_sel_coff}_CN_{l_sel_cont}.pdf",
                    mime="application/pdf",
                    type="secondary",
                    key="btn_print_shortage"
                )

            with btn_c3:
                # Contract-wise Shortlist: only ACTIVE/PENDING items for the currently
                # selected Brand + Contract filters. Strict rounding removes any
                # plus/minus float micro-garbage so true-zero articles never appear.
                df_contract_shortlist = df_ledger[df_ledger["Remaining Balance"].apply(round_bal) > 0]
                pdf_buf_contract = generate_ledger_pdf(item_summary, df_contract_shortlist, l_sel_coff, l_sel_cont, l_sel_art, report_type="CONTRACT_SHORTLIST")
                safe_brand = (l_sel_br if l_sel_br != "All" else "AllBrands")
                safe_cont  = (l_sel_cont if l_sel_cont != "All" else "AllContracts")
                st.download_button(
                    label="📋 Print Contract Shortlist (PDF)",
                    data=pdf_buf_contract,
                    file_name=f"Contract_Shortlist_{safe_brand}_{safe_cont}.pdf",
                    mime="application/pdf",
                    type="secondary",
                    key="btn_print_contract_shortlist"
                )

# ═══════════════════════════════════════════════
# TAB 5 — SHEET UPLOAD (UNCHANGED)
# ═══════════════════════════════════════════════
with tab5:
    if _access_ok("📤 Sheet Upload"):
        st.markdown('<div class="sec">📤 Upload Sale Contract Sheet</div>', unsafe_allow_html=True)
        u1,_ = st.columns([1,2])
        with u1:
            u_coff = st.text_input("Call-Off No. * (e.g. 288):", key="u_coff")
            u_file = st.file_uploader("Select Excel or CSV File", type=["csv","xlsx"], key="u_file")

        if u_file and u_coff:
            if st.button("💾 Save Sheet Data", type="primary", key="u_save"):
                try:
                    raw = pd.read_csv(u_file, header=None) if u_file.name.endswith("csv") else pd.read_excel(u_file, header=None)
                    conn = get_conn()
                    conn.execute("DELETE FROM sheet_orders WHERE call_off_no=?", (u_coff.strip(),))
                    saved = 0
                    current_headers = {}

                    for idx, row in raw.iterrows():
                        row_str = [str(x).strip().lower().replace(" ", "") for x in row.values]
                    
                        if any("article" in s or "cont#" in s or "contract" in s for s in row_str):
                            current_headers = {str(row.iloc[i]).strip().lower().replace(" ", ""): i for i, s in enumerate(row_str) if str(row.iloc[i]).strip() != "nan"}
                            continue
                    
                        if not current_headers:
                            continue
                        
                        def get_val_by_keys(*keys):
                            for k in keys:
                                k_lbl = k.lower().replace(" ", "")
                                if k_lbl in current_headers:
                                    v = str(row.iloc[current_headers[k_lbl]]).strip()
                                    if v and v not in ["nan", "None", "-", "0", "0.0"]: return v
                            return ""
                    
                        art  = get_val_by_keys("article no.", "article no", "article", "art no", "item code")
                        cont = get_val_by_keys("cont #", "contract #", "sale contract", "contract")
                        po   = get_val_by_keys("po #", "po#", "po no", "purchase order")
                        br   = get_val_by_keys("brand", "customer/brand", "customer", "client")
                    
                        if not art or "total" in art.lower() or "total" in cont.lower():
                            continue
                        
                        for ct in ITEM_TYPES:
                            search_names = [ct.lower().replace(" ", "")]
                            if ct == "Inlay Card / Bandrolle":
                                search_names = ["inlaycard/bandrolle", "inlaycard", "bandrolle", "inlay", "bandroll"]
                            elif ct == "Tag Card / Barcode Sticker":
                                search_names = ["tagcard/barcodesticker", "tagcard", "barcodesticker", "tag", "tagcardsticker"]
                            elif ct == "Barcode Item":
                                search_names = ["barcodeitem", "barcode item", "barcode", "barcodepure", "barcodeonly"]
                            elif ct == "Transparent Sticker":
                                search_names = ["transparentsticker", "pricesticker", "roundsticker", "transparent", "sticker", "price", "round"]
                            elif ct == "Eco Friendly":
                                search_names = ["ecofriendly", "eco-friendly", "eco"]
                        
                            raw_qty = ""
                            for name in search_names:
                                cleaned_name = name.lower().replace(" ", "")
                                if cleaned_name in current_headers:
                                    raw_qty = str(row.iloc[current_headers[cleaned_name]]).strip()
                                    break
                                
                            try:
                                oq = float(raw_qty.replace(",","")) if raw_qty not in ["", "nan", "-", "None"] else 0
                            except:
                                oq = 0
                            
                            if oq > 0:
                                conn.execute(
                                    "INSERT INTO sheet_orders (call_off_no,po_no,sale_contract,brand,article,category,order_qty) VALUES (?,?,?,?,?,?,?)",
                                    (u_coff.strip(), po, cont, br, art, ct, oq))
                                saved += 1
                            
                    conn.commit(); conn.close()
                    if saved > 0:
                        st.success(f"🎉 Call-Off {u_coff.strip()} — All multi-row items ({saved} configs) auto-scanned and saved!")
                        st.rerun()
                    else:
                        st.warning("⚠️ No valid rows found. Please check column titles.")
                except Exception as e:
                    st.error(f"Error parsing sheet: {e}")

        st.markdown("---")
        st.markdown('<div class="sec">🗂️ Active Sheet Repository</div>', unsafe_allow_html=True)

        value_df = q("""
            SELECT call_off_no AS "Call-Off Sheet",
                   COUNT(*) AS "Config Rows",
                   SUM(order_qty) AS "Total Qty"
            FROM sheet_orders GROUP BY call_off_no ORDER BY call_off_no
        """)

        if value_df.empty:
            st.info("No sheets loaded yet.")
        else:
            value_df["Total Qty"] = value_df["Total Qty"].apply(lambda x: round_and_format(float(str(x).replace(",",""))))
            st.dataframe(value_df, width='stretch', hide_index=True)

            st.markdown("### 🗑️ Delete a Sheet")
            sheet_opts = q("SELECT DISTINCT call_off_no FROM sheet_orders ORDER BY call_off_no")["call_off_no"].tolist()
            to_drop = st.selectbox("Select sheet to delete:", ["-- Select --"] + sheet_opts, key="drop_sel")
            if to_drop != "-- Select --":
                st.warning(f"⚠️ This will permanently delete all order targets for **{to_drop}**.")
                if st.checkbox(f"Confirm delete '{to_drop}'", key="drop_chk"):
                    if st.button("🗑️ Erase Sheet", type="primary", key="drop_btn"):
                        conn = get_conn()
                        conn.execute("DELETE FROM sheet_orders WHERE call_off_no=?", (to_drop,))
                        conn.commit(); conn.close()
                        st.success(f"✅ Sheet '{to_drop}' deleted!"); st.rerun()

        # ═══════════════════════════════════════════════
        # NEW ADDITION: "Packaging Detail" Call-Off Sheet Parser
        # ═══════════════════════════════════════════════
        # Pure addition — a second, independent upload path for the
        # "Packaging Detail" sheet format (columns: Call Off, Cont #, PO #,
        # Article no., Brand, Article description, Inlay Card, Tag Card,
        # Barcode Item, Price sticker, Eco friendly, Washing Paper, Safety).
        # No hard-coded Call-Off/Contract number — everything is auto-detected
        # straight from the sheet, so it works the same for 290, 291, 292...
        # Dual-pack (Jersey/Molton) variants are auto-tagged from the Article
        # description text and stored in the new `variant` column, so the DC
        # Entry tab can offer split quantity fields when both exist for a PO.
        st.markdown("---")
        st.markdown('<div class="sec">📦 NEW: Packaging Detail Call-Off Sheet (Auto-Detect)</div>', unsafe_allow_html=True)
        st.caption("یہ نیا، علیحدہ اپلوڈ راستہ ہے — کوئی کانٹریکٹ نمبر ہارڈ کوڈڈ نہیں، شیٹ سے خود بخود ڈیٹیکٹ ہوگا۔ پرانا اپلوڈ اوپر ویسے ہی کام کرتا رہے گا۔")

        pd_file = st.file_uploader("Select Packaging Detail Excel File (.xlsx)", type=["xlsx"], key="pd_file")

        # Universal category mapping — maps the Packaging Detail sheet's own
        # columns onto the EXACT existing ITEM_TYPES strings (not a new
        # category scheme) so 290-uploaded data flows straight through the
        # unchanged DC Entry / Master Ledger / PDF reports with zero
        # disruption. "Price sticker" has no exact existing counterpart, so
        # it's routed into "Barcode Item" as the closest generic sticker
        # bucket — adjust PD_CATEGORY_MAP below if a different mapping is
        # wanted.
        PD_CATEGORY_MAP = {
            "Inlay Card":    "Inlay Card / Bandrolle",
            "Tag Card":      "Tag Card / Barcode Sticker",
            "Barcode Item":  "Barcode Item",
            "Price sticker": "Barcode Item",
            "Eco friendly":  "Eco Friendly",
            "Washing Paper": "Washing Paper",
            "Safety":        "Safety",
        }

        def _pd_col_idx(header, name_options):
            for i, h in enumerate(header):
                for opt in name_options:
                    if opt.lower() in (h or "").lower():
                        return i
            return None

        def parse_packaging_detail_sheet(file_obj):
            import openpyxl
            import re
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.worksheets[0]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            header = [str(h).strip() if h else "" for h in rows[0]]

            idx = {
                "call_off": _pd_col_idx(header, ["Call Off"]),
                "cont":     _pd_col_idx(header, ["Cont #"]),
                "po":       _pd_col_idx(header, ["PO #"]),
                "article":  _pd_col_idx(header, ["Article no"]),
                "brand":    _pd_col_idx(header, ["Brand"]),
                "desc":     _pd_col_idx(header, ["Article description"]),
            }
            for src_col in PD_CATEGORY_MAP:
                idx[src_col] = _pd_col_idx(header, [src_col])

            out = []
            for r in rows[1:]:
                if r is None or all(v is None for v in r):
                    continue

                def _s(key):
                    i = idx.get(key)
                    return str(r[i]).strip() if i is not None and r[i] is not None else ""

                call_off, cont, po, article, brand, desc = (
                    _s("call_off"), _s("cont"), _s("po"), _s("article"), _s("brand"), _s("desc"))
                if not call_off or not article:
                    continue

                combined_up = (brand + " " + desc).upper()
                if re.search(r'\bMOLTON\b|\bMO\b', combined_up):
                    variant = "Molton"
                elif re.search(r'\bJERSEY\b|\bJER\b', combined_up):
                    variant = "Jersey"
                else:
                    variant = ""

                for src_col, uni_cat in PD_CATEGORY_MAP.items():
                    ci = idx.get(src_col)
                    if ci is None:
                        continue
                    val = r[ci]
                    if val is None or str(val).strip() in ("-", ""):
                        continue
                    try:
                        qty = float(val)
                    except (ValueError, TypeError):
                        continue
                    if qty <= 0:
                        continue
                    out.append({
                        "call_off_no": call_off, "po_no": po, "sale_contract": cont,
                        "brand": brand, "article": article, "category": uni_cat,
                        "order_qty": qty, "variant": variant,
                    })
            return out

        if pd_file:
            try:
                pd_rows = parse_packaging_detail_sheet(pd_file)
            except Exception as e:
                pd_rows = []
                st.error(f"⚠️ Could not parse this file: {e}")

            if pd_rows:
                df_preview = pd.DataFrame(pd_rows)
                detected_calloffs = sorted(df_preview["call_off_no"].unique().tolist())
                detected_contracts = sorted(df_preview["sale_contract"].unique().tolist())
                detected_pos = sorted(df_preview["po_no"].unique().tolist())
                st.success(f"✅ Auto-detected — Call-Off(s): {', '.join(detected_calloffs)} | "
                           f"Contract(s): {', '.join(detected_contracts)} | {len(detected_pos)} PO(s) | {len(df_preview)} rows")
                dual_pack_contracts = sorted(
                    df_preview[df_preview["variant"] != ""].groupby("sale_contract")["variant"].nunique()
                    .loc[lambda s: s > 1].index.tolist()
                )
                if dual_pack_contracts:
                    st.info(f"🧵 Dual-pack (Jersey + Molton) detected for Contract(s): {', '.join(dual_pack_contracts)} — "
                             "DC Entry will show split quantity fields for these.")
                st.dataframe(df_preview, width='stretch', hide_index=True)

                if st.button("💾 Insert into System", type="primary", key="pd_insert_btn"):
                    conn = get_conn()
                    for co in detected_calloffs:
                        conn.execute("DELETE FROM sheet_orders WHERE call_off_no=?", (co,))
                    for row in pd_rows:
                        conn.execute("""
                            INSERT INTO sheet_orders (call_off_no,po_no,sale_contract,brand,article,category,order_qty,variant)
                            VALUES (?,?,?,?,?,?,?,?)
                        """, (row["call_off_no"], row["po_no"], row["sale_contract"], row["brand"],
                              row["article"], row["category"], row["order_qty"], row["variant"]))
                    conn.commit()
                    conn.close()
                    st.success(f"✅ {len(pd_rows)} rows inserted for Call-Off(s): {', '.join(detected_calloffs)}")
                    st.rerun()
            else:
                st.warning("No usable rows found — please check the sheet matches the expected 'Packaging Detail' column layout.")


# ═══════════════════════════════════════════════
# TAB 6 — 🆕 BILTY MANAGEMENT (Lahore ➜ Karachi Dispatch)
# ═══════════════════════════════════════════════
with tab6:
    if _access_ok("🚚 Bilty Management"):
        st.markdown('<div class="sec">🚚 Bilty Management — Lahore Factory Dispatch</div>', unsafe_allow_html=True)
        st.caption("یہ ایک نیا علیحدہ بلٹی لیجر ہے۔ اس کا پرانے Master Ledger (Ordered vs Received) کے حساب کتاب پر کوئی اثر نہیں پڑتا — صرف بلٹی/ڈسپیچ کا ریکارڈ رکھتا ہے۔")

        # NEW: Date moved to the TOP as a single global header/filter for the
        # whole day's dispatch, instead of being buried per-row at the bottom.
        st.markdown("""
        <div class="auto-box" style="background:#0c4a6e;border:1px solid #38bdf8;color:#fff;padding:6px 10px;margin-bottom:8px;font-size:12px;">
          🛣️ <b>Route:</b> Lahore ➜ Karachi
        </div>""", unsafe_allow_html=True)
        bilty_date_val = st.date_input("📅 Bilty Date (applies to this whole dispatch)", value=date.today(), key="bilty_date")

        b_c1, b_c2 = st.columns([2, 2])

        with b_c1:
            b_coff_opts = get_calloff_list()
            b_sel_coff = st.selectbox("Select Call-Off Sheet *", ["-- Select --"] + b_coff_opts, key="bilty_coff")

        b_sel_cont = "-- Select --"
        if b_sel_coff != "-- Select --":
            cont_opts = q(
                "SELECT DISTINCT sale_contract FROM sheet_orders WHERE call_off_no=? AND TRIM(sale_contract)!='' ORDER BY sale_contract",
                [b_sel_coff]
            )["sale_contract"].tolist()
            with b_c2:
                b_sel_cont = st.selectbox("Select Sales Contract *", ["-- Select --"] + cont_opts, key="bilty_cont")

        if b_sel_coff == "-- Select --" or b_sel_cont == "-- Select --":
            st.info("👆 Please select a Call-Off Sheet and Sales Contract to load its articles.")
        else:
            # FIX: quantity now comes straight from the Call-Off Sheet (sum of
            # order_qty per Article + Category) instead of a manual fixed number.
            # No typing needed — the checkbox label shows the exact sheet quantity,
            # and ticking it adds that exact quantity to the grand total.
            df_arts = q("""
                SELECT article, category, SUM(order_qty) AS sheet_qty
                FROM sheet_orders
                WHERE call_off_no=? AND sale_contract=?
                GROUP BY article, category
                ORDER BY article, category
            """, [b_sel_coff, b_sel_cont])

            if df_arts.empty:
                st.info("No articles found for this Call-Off / Contract combination.")
            else:
                st.markdown("##### ✅ Tick items packed into this Bilty")
                st.caption("مقدار خودکار طور پر کال-آف شیٹ سے لی جا رہی ہے۔ چاہیں تو مکمل مقدار کے لیے صرف ٹک کریں، یا جزوی/کم مقدار کے لیے نیچے دیے گئے خانے میں خود لکھ دیں۔")
                grand_total = 0.0
                ticked_items = []

                for art, grp in df_arts.groupby("article"):
                    with st.expander(f"📦 Article: {art}", expanded=True):
                        tcols = st.columns(3)
                        for i, (_, r_row) in enumerate(grp.iterrows()):
                            cat = r_row["category"]
                            cat_qty = float(r_row["sheet_qty"] or 0)
                            cb_key = f"bilty_cb_{b_sel_coff}_{b_sel_cont}_{art}_{cat}"
                            # NEW ADDITION: manual quantity override, sits right
                            # next to the existing checkbox. Defaults to the full
                            # sheet quantity (same as before if left untouched) —
                            # editable down for a partial/custom dispatch amount.
                            mq_key = f"bilty_manualqty_{b_sel_coff}_{b_sel_cont}_{art}_{cat}"
                            with tcols[i % 3]:
                                is_checked = st.checkbox(f"{cat} ({cat_qty:,.0f} Pcs)", key=cb_key)
                                manual_qty = st.number_input(
                                    f"Qty for {cat} (override)", min_value=0.0,
                                    max_value=cat_qty if cat_qty > 0 else None,
                                    value=cat_qty, step=1.0, key=mq_key,
                                    disabled=not is_checked, label_visibility="collapsed"
                                )
                            if is_checked:
                                grand_total += manual_qty
                                ticked_items.append((art, cat, manual_qty))

                st.markdown(f"""<div class="kpi-row">
                  <div class="kpi kp" style="font-size:1rem;">🧮 Live Grand Total: {grand_total:,.0f} Pcs ({len(ticked_items)} item(s) ticked)</div>
                </div>""", unsafe_allow_html=True)

                st.markdown("---")
                bc1, bc2, bc3 = st.columns(3)
                with bc1:
                    cartons_n = st.number_input("Number of Cartons *", min_value=0, step=1, key="bilty_cartons")
                with bc2:
                    transport_mode_sel = st.selectbox("Transport Mode *", ["By Air", "By Train", "By Road (Vehicle)"], key="bilty_transport")
                with bc3:
                    vehicle_no = ""
                    if transport_mode_sel == "By Road (Vehicle)":
                        vehicle_no = st.text_input("Vehicle No.", key="bilty_vehicle_no", placeholder="e.g. LES-1234")
                transport_mode = f"{transport_mode_sel} ({vehicle_no})" if vehicle_no else transport_mode_sel

                if st.button("🚚 Save Bilty Record", type="primary", key="bilty_save"):
                    if not ticked_items:
                        st.error("⚠️ Please tick at least one item before saving.")
                    elif cartons_n <= 0:
                        st.error("⚠️ Please enter the number of cartons.")
                    else:
                        conn = get_conn()
                        for art, cat, cat_qty in ticked_items:
                            conn.execute("""
                                INSERT INTO bilty (call_off_no, contract_no, article, category, qty, cartons, transport_mode, bilty_date, created_at)
                                VALUES (?,?,?,?,?,?,?,?,?)
                            """, (b_sel_coff, b_sel_cont, art, cat, float(cat_qty), int(cartons_n),
                                  transport_mode, str(bilty_date_val), str(datetime.datetime.now())))
                        conn.commit()
                        conn.close()
                        st.success(f"✅ Bilty saved — {len(ticked_items)} item(s), {grand_total:,.0f} Pcs total, {cartons_n} Carton(s) via {transport_mode}.")
                        st.rerun()

                st.markdown("---")
                st.markdown("##### 📜 Bilty History for this Contract")
                df_bilty_hist_full = q("""
                    SELECT id, bilty_date AS "Date", article AS "Article", category AS "Item Type",
                           qty AS "Qty", cartons AS "Cartons", transport_mode AS "Transport"
                    FROM bilty WHERE call_off_no=? AND contract_no=?
                    ORDER BY id DESC
                """, [b_sel_coff, b_sel_cont])
                if df_bilty_hist_full.empty:
                    st.info("No Bilty records saved yet for this contract.")
                else:
                    st.dataframe(df_bilty_hist_full.drop(columns=["id"]), width='stretch', hide_index=True)

                    # ═══════════════════════════════════════════════
                    # NEW ADDITION: Edit / Delete a Bilty record — CEO/Admin only.
                    # ═══════════════════════════════════════════════
                    if current_role in ("Admin", "CEO"):
                        with st.expander("✏️ Edit / 🗑️ Delete a Bilty Entry"):
                            row_opts = {
                                f"#{r['id']} — {r['Date']} | {r['Article']} | {r['Item Type']} | {r['Qty']:,.0f} Pcs": r["id"]
                                for _, r in df_bilty_hist_full.iterrows()
                            }
                            sel_label = st.selectbox("Select entry", ["-- Select --"] + list(row_opts.keys()), key="bilty_edit_sel")
                            if sel_label != "-- Select --":
                                sel_id = row_opts[sel_label]
                                sel_row = df_bilty_hist_full[df_bilty_hist_full["id"] == sel_id].iloc[0]

                                st.markdown("**✏️ Edit Entry**")
                                with st.form(f"bilty_edit_form_{sel_id}"):
                                    ec1, ec2, ec3 = st.columns(3)
                                    with ec1:
                                        e_date = st.date_input("Bilty Date", value=pd.to_datetime(sel_row["Date"]).date(), key=f"e_date_{sel_id}")
                                        e_article = st.text_input("Article", value=str(sel_row["Article"]), key=f"e_art_{sel_id}")
                                    with ec2:
                                        e_category = st.text_input("Item Type / Category", value=str(sel_row["Item Type"]), key=f"e_cat_{sel_id}")
                                        e_qty = st.number_input("Qty", min_value=0.0, value=float(sel_row["Qty"]), step=1.0, key=f"e_qty_{sel_id}")
                                    with ec3:
                                        e_cartons = st.number_input("Cartons", min_value=0, value=int(sel_row["Cartons"]), step=1, key=f"e_cart_{sel_id}")
                                        e_transport = st.text_input("Transport Mode", value=str(sel_row["Transport"]), key=f"e_trans_{sel_id}")

                                    if st.form_submit_button("💾 Update Entry", type="primary"):
                                        conn = get_conn()
                                        conn.execute("""
                                            UPDATE bilty SET bilty_date=?, article=?, category=?, qty=?, cartons=?, transport_mode=?
                                            WHERE id=?
                                        """, (str(e_date), e_article.strip(), e_category.strip(), float(e_qty), int(e_cartons), e_transport.strip(), int(sel_id)))
                                        conn.commit()
                                        conn.close()
                                        st.success(f"✅ Bilty entry #{sel_id} updated.")
                                        st.rerun()

                                st.markdown("**🗑️ Delete Entry**")
                                confirm_bilty_del = st.checkbox(f"Confirm permanent delete of entry #{sel_id}", key=f"bilty_del_chk_{sel_id}")
                                if confirm_bilty_del and st.button("🗑️ Delete This Bilty Entry", key=f"bilty_del_btn_{sel_id}"):
                                    conn = get_conn()
                                    conn.execute("DELETE FROM bilty WHERE id=?", (int(sel_id),))
                                    conn.commit()
                                    conn.close()
                                    st.success(f"✅ Bilty entry #{sel_id} deleted.")
                                    st.rerun()

        # ═══════════════════════════════════════════════
        # NEW ADDITION: Dual Summary (Contract-Wise + Item/Category-Wise) for
        # the selected Bilty Date, plus Excel/PDF export of this summary.
        # ═══════════════════════════════════════════════
        st.markdown("---")
        st.markdown(f'<div class="sec">📊 Daily Dispatch Summary — {_fmt_date_ddmmyyyy(str(bilty_date_val))}</div>', unsafe_allow_html=True)

        df_day = q("""
            SELECT call_off_no, contract_no, article, category, qty, cartons, transport_mode
            FROM bilty WHERE bilty_date=?
        """, [str(bilty_date_val)])

        if df_day.empty:
            st.info("اس تاریخ کے لیے ابھی کوئی بلٹی ریکارڈ نہیں۔")
        else:
            sum_c1, sum_c2 = st.columns(2)
            with sum_c1:
                st.markdown("**📑 Contract-Wise Sum**")
                df_contract_sum = df_day.groupby("contract_no").agg(
                    Total_Qty=("qty", "sum"), Total_Cartons=("cartons", "sum")
                ).reset_index().rename(columns={"contract_no": "Contract #", "Total_Qty": "Total Qty", "Total_Cartons": "Total Cartons"})
                st.dataframe(df_contract_sum, width='stretch', hide_index=True)
            with sum_c2:
                st.markdown("**🏷️ Item/Category-Wise Sum**")
                df_category_sum = df_day.groupby("category").agg(Total_Qty=("qty", "sum")).reset_index().rename(
                    columns={"category": "Item Type / Category", "Total_Qty": "Total Qty"})
                st.dataframe(df_category_sum, width='stretch', hide_index=True)

            def _generate_bilty_summary_excel(day_str, df_raw, df_contract, df_category):
                import openpyxl
                from openpyxl.styles import Font
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"Bilty Summary {day_str}"[:31]
                bold = Font(bold=True)
                ws["A1"] = f"Bilty Dispatch Summary — {_fmt_date_ddmmyyyy(day_str)}"
                ws["A1"].font = Font(bold=True, size=14)
                ws["A2"] = "Route: Lahore ➜ Karachi"

                r = 4
                ws.cell(row=r, column=1, value="Contract-Wise Sum").font = bold
                r += 1
                for ci, col in enumerate(df_contract.columns):
                    ws.cell(row=r, column=1 + ci, value=col).font = bold
                r += 1
                for _, row in df_contract.iterrows():
                    for ci, col in enumerate(df_contract.columns):
                        ws.cell(row=r, column=1 + ci, value=row[col])
                    r += 1

                r += 2
                ws.cell(row=r, column=1, value="Item/Category-Wise Sum").font = bold
                r += 1
                for ci, col in enumerate(df_category.columns):
                    ws.cell(row=r, column=1 + ci, value=col).font = bold
                r += 1
                for _, row in df_category.iterrows():
                    for ci, col in enumerate(df_category.columns):
                        ws.cell(row=r, column=1 + ci, value=row[col])
                    r += 1

                r += 2
                ws.cell(row=r, column=1, value="Detailed Dispatch Log").font = bold
                r += 1
                detail_cols = ["call_off_no", "contract_no", "article", "category", "qty", "cartons", "transport_mode"]
                for ci, col in enumerate(detail_cols):
                    ws.cell(row=r, column=1 + ci, value=col).font = bold
                r += 1
                for _, row in df_raw.iterrows():
                    for ci, col in enumerate(detail_cols):
                        ws.cell(row=r, column=1 + ci, value=row[col])
                    r += 1

                for col in "ABCDEFG":
                    ws.column_dimensions[col].width = 18
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf

            def _generate_bilty_summary_pdf(day_str, df_contract, df_category):
                buffer = io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
                story = []
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle('BiltyTitle', parent=styles['Heading1'], fontSize=15, alignment=1)
                story.append(Paragraph(f"<b>VERTEX PACKAGING — Bilty Dispatch Summary</b>", title_style))
                story.append(Paragraph(f"Date: {_fmt_date_ddmmyyyy(day_str)} | Route: Lahore ➜ Karachi",
                                        ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, alignment=1)))
                story.append(Spacer(1, 14))

                story.append(Paragraph("<b>Contract-Wise Sum</b>", styles['Heading2']))
                t1_data = [list(df_contract.columns)] + df_contract.values.tolist()
                t1 = Table(t1_data, colWidths=[170, 170, 170])
                t1.setStyle(TableStyle([('BACKGROUND', (0,0),(-1,0), colors.HexColor('#f1f5f9')), ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'), ('GRID',(0,0),(-1,-1),0.5,colors.grey), ('FONTSIZE',(0,0),(-1,-1),9)]))
                story.append(t1)
                story.append(Spacer(1, 16))

                story.append(Paragraph("<b>Item/Category-Wise Sum</b>", styles['Heading2']))
                t2_data = [list(df_category.columns)] + df_category.values.tolist()
                t2 = Table(t2_data, colWidths=[255, 255])
                t2.setStyle(TableStyle([('BACKGROUND', (0,0),(-1,0), colors.HexColor('#f1f5f9')), ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'), ('GRID',(0,0),(-1,-1),0.5,colors.grey), ('FONTSIZE',(0,0),(-1,-1),9)]))
                story.append(t2)

                doc.build(story)
                buffer.seek(0)
                return buffer

            day_str = str(bilty_date_val)
            bexp1, bexp2 = st.columns(2)
            with bexp1:
                st.download_button(
                    label="⬇️ Export to Excel",
                    data=_generate_bilty_summary_excel(day_str, df_day, df_contract_sum, df_category_sum),
                    file_name=f"Bilty_Summary_{_fmt_date_ddmmyyyy(day_str)}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="bilty_export_excel")
            with bexp2:
                st.download_button(
                    label="⬇️ Export to PDF",
                    data=_generate_bilty_summary_pdf(day_str, df_contract_sum, df_category_sum),
                    file_name=f"Bilty_Summary_{_fmt_date_ddmmyyyy(day_str)}.pdf",
                    mime="application/pdf",
                    type="primary", key="bilty_export_pdf")

# ═══════════════════════════════════════════════
# TAB 7 — 🆕 USER MANAGEMENT (Admin only)
# ═══════════════════════════════════════════════
with tab7:
    if _access_ok("👤 User Management"):
        st.markdown('<div class="sec">👤 User Management — Admin Only</div>', unsafe_allow_html=True)

        # NEW ADDITION: Performance Diagnostics — every DB call slower than
        # SLOW_QUERY_THRESHOLD_S (2s) gets logged here with its elapsed time
        # and exact SQL, live, as the app is used. This is the tool for
        # pinpointing exactly which function/query is causing a slow save —
        # reproduce the slow action once, then check this panel immediately
        # after; the same entries are also printed to the Streamlit Cloud
        # app logs (Manage app → logs) if you'd rather grep there.
        with st.expander("🩺 Performance Diagnostics — Slow Query Log", expanded=False):
            if not _SLOW_QUERY_LOG:
                st.caption("No slow queries recorded yet this session (threshold: "
                           f"{SLOW_QUERY_THRESHOLD_S:.0f}s). Reproduce the slow action, "
                           "then reopen this panel.")
            else:
                df_slow = pd.DataFrame(list(reversed(_SLOW_QUERY_LOG)))
                st.dataframe(df_slow, width='stretch', hide_index=True)
                st.caption(f"Showing the {len(df_slow)} most recent slow/failed queries this app session "
                           "(most recent first). ERROR rows show a connection-level failure; RETRY rows "
                           "show how long the automatic retry took after that failure.")
            if st.button("🗑️ Clear Log", key="clear_slow_log"):
                _SLOW_QUERY_LOG.clear()
                st.rerun()

        st.markdown("##### ➕ Create New User")
        with st.form("create_user_form", clear_on_submit=True):
            nu_c1, nu_c2 = st.columns(2)
            with nu_c1:
                nu_username = st.text_input("Username *")
                nu_fullname = st.text_input("Full Name *")
            with nu_c2:
                nu_password = st.text_input("Password *", type="password")
                nu_role = st.selectbox("Role *", ["Admin", "Data Entry", "Viewer", "CEO", "Rider"])
            nu_submit = st.form_submit_button("➕ Create User", type="primary")

        if nu_submit:
            if not nu_username.strip() or not nu_password or not nu_fullname.strip():
                st.error("⚠️ Username, Full Name and Password are all required.")
            else:
                existing = q("SELECT id FROM app_users WHERE username=?", [nu_username.strip()])
                if not existing.empty:
                    st.error(f"⚠️ Username '{nu_username.strip()}' already exists.")
                else:
                    conn = get_conn()
                    conn.execute("""
                        INSERT INTO app_users (username,password_hash,role,full_name,created_at)
                        VALUES (?,?,?,?,?)
                    """, (nu_username.strip(), _hash_password(nu_password), nu_role,
                          nu_fullname.strip(), str(datetime.datetime.now())))
                    conn.commit()
                    conn.close()
                    st.success(f"✅ User '{nu_username.strip()}' created with role '{nu_role}'.")
                    st.rerun()

        st.markdown("---")
        st.markdown("##### 🔑 Reset a User's Password")
        all_usernames = q("SELECT username FROM app_users ORDER BY username")["username"].tolist()
        rp_c1, rp_c2, rp_c3 = st.columns([2, 2, 1])
        with rp_c1:
            rp_user = st.selectbox("Select User", all_usernames, key="rp_user")
        with rp_c2:
            rp_newpass = st.text_input("New Password", type="password", key="rp_newpass")
        with rp_c3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🔑 Reset Password", key="rp_btn"):
                if not rp_newpass:
                    st.error("⚠️ Enter a new password first.")
                else:
                    conn = get_conn()
                    conn.execute("UPDATE app_users SET password_hash=? WHERE username=?",
                                 (_hash_password(rp_newpass), rp_user))
                    conn.commit()
                    conn.close()
                    st.success(f"✅ Password reset for '{rp_user}'.")

        st.markdown("---")
        st.markdown("##### 📋 All Users")
        df_users = q("SELECT username AS Username, full_name AS \"Full Name\", role AS Role, created_at AS \"Created At\" FROM app_users ORDER BY username")
        st.dataframe(df_users, width='stretch', hide_index=True)

        st.markdown("##### 🗑️ Delete a User")
        del_candidates = [u for u in all_usernames if u != current_user["username"]]
        if del_candidates:
            del_user = st.selectbox("Select user to delete", ["-- Select --"] + del_candidates, key="del_user_sel")
            if del_user != "-- Select --":
                if st.checkbox(f"Confirm delete '{del_user}'", key="del_user_chk"):
                    if st.button("🗑️ Delete User", key="del_user_btn"):
                        conn = get_conn()
                        conn.execute("DELETE FROM app_users WHERE username=?", (del_user,))
                        conn.commit()
                        conn.close()
                        st.success(f"✅ User '{del_user}' deleted.")
                        st.rerun()
        else:
            st.caption("No other users to delete.")

# ═══════════════════════════════════════════════
# TAB 8 — 🆕 DAILY EXPENSES / STAFF EXPENSE
# ═══════════════════════════════════════════════
with tab8:
    if _access_ok("💸 Daily Expenses"):
        st.markdown('<div class="sec">💸 Daily Expenses — Staff / Rider Expense Ledger</div>', unsafe_allow_html=True)

        # SECURITY UPDATE: Rider AND Data Entry are both "restricted" users —
        # they may add their own expenses, but must not see, filter, edit or
        # delete anyone else's. Only Admin/CEO get the full ledger view.
        is_restricted = current_role in ("Rider", "Data Entry")

        # ───────────── ENTRY FORM ─────────────
        st.markdown("##### ➕ New Expense Entry")
        with st.form("expense_entry_form", clear_on_submit=True):
            exp_c1, exp_c2 = st.columns(2)
            with exp_c1:
                exp_date = st.date_input("Date", value=date.today(), key="exp_date")
                if is_restricted:
                    target_user = current_user["username"]
                    st.text_input("Entering Expense For (You)", value=target_user, disabled=True, key="exp_target_ro")
                else:
                    _riders_df = q("SELECT username FROM app_users WHERE role='Rider' ORDER BY username")
                    rider_opts = _riders_df["username"].tolist() if not _riders_df.empty else []
                    if current_user["username"] not in rider_opts:
                        rider_opts = [current_user["username"]] + rider_opts
                    target_user = st.selectbox("Rider / Staff *", rider_opts, key="exp_target_user")
                exp_category = st.selectbox("Category *", EXPENSE_CATEGORIES, key="exp_category")
            with exp_c2:
                exp_amount = st.number_input("Amount (PKR) *", min_value=0.0, step=50.0, key="exp_amount")
                exp_ref = st.text_input("Reference / Bill No. (Optional)", key="exp_ref")
                exp_remarks = st.text_area("Remarks", key="exp_remarks", height=80)

            st.markdown("**📸 Bill / Receipt Photo**")
            exp_photo_mode = st.radio(
                "Photo source", ["📷 Take Photo", "🖼️ Upload from Gallery", "— Skip —"],
                horizontal=True, key="exp_photo_mode")
            bill_img_file = None
            if exp_photo_mode == "📷 Take Photo":
                bill_img_file = st.camera_input("📷 بل یا پرچی کی تصویر کھینچیں (Take Bill Photo)", key="exp_camera")
            elif exp_photo_mode == "🖼️ Upload from Gallery":
                bill_img_file = st.file_uploader("🖼️ Upload Bill/Receipt Image", type=["png", "jpg", "jpeg"], key="exp_upload")

            # Convert image file to bytes for database storage
            img_bytes = bill_img_file.getvalue() if bill_img_file is not None else None

            exp_submit = st.form_submit_button("💾 Save Expense", type="primary")

        if exp_submit:
            if is_restricted is False and not target_user:
                st.error("⚠️ No Rider/Staff selected. Create a Rider account first from 👤 User Management.")
            elif exp_amount <= 0:
                st.error("⚠️ Amount must be greater than 0.")
            else:
                conn = get_conn()
                conn.execute("""
                    INSERT INTO rider_expenses (date, user_id, entered_by, category, amount, reference_no, remarks, bill_image)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (str(exp_date), target_user, current_user["username"], exp_category, float(exp_amount),
                      str(exp_ref).strip(), str(exp_remarks).strip(), img_bytes))
                conn.commit()
                conn.close()
                st.success(f"✅ Expense of PKR {exp_amount:,.0f} saved for {target_user}.")
                st.rerun()

        st.markdown("---")

        # ───────────── LEDGER ─────────────
        if is_restricted:
            st.markdown("##### 📋 My Expense History")
            df_exp = q("""SELECT id, date, category, amount, reference_no, remarks, bill_image
                          FROM rider_expenses WHERE user_id=? ORDER BY date DESC, id DESC""",
                       [current_user["username"]])
            filt_rider = "All"
        else:
            st.markdown("##### 📋 Expense Ledger — CEO / Admin View")
            all_riders = q("SELECT DISTINCT user_id FROM rider_expenses ORDER BY user_id")["user_id"].tolist()
            filt_rider = st.selectbox("Filter by Rider/Staff", ["All"] + all_riders, key="exp_filter_rider")
            if filt_rider == "All":
                df_exp = q("""SELECT id, date, user_id, category, amount, reference_no, remarks, bill_image
                              FROM rider_expenses ORDER BY date DESC, id DESC""")
            else:
                df_exp = q("""SELECT id, date, user_id, category, amount, reference_no, remarks, bill_image
                              FROM rider_expenses WHERE user_id=? ORDER BY date DESC, id DESC""", [filt_rider])

        if df_exp.empty:
            st.info("No expense entries yet.")
        else:
            total_spent = float(df_exp["amount"].sum())
            adv_mask = df_exp["category"].astype(str).str.contains("Advance Salary", na=False)
            total_advance = float(df_exp.loc[adv_mask, "amount"].sum())
            total_other = total_spent - total_advance
            kc1, kc2, kc3 = st.columns(3)
            kc1.markdown(f"<div class='kpi kb'>💰 Total Spent<br><b>PKR {total_spent:,.0f}</b></div>", unsafe_allow_html=True)
            kc2.markdown(f"<div class='kpi ka'>🧾 Operational Spend<br><b>PKR {total_other:,.0f}</b></div>", unsafe_allow_html=True)
            kc3.markdown(f"<div class='kpi kp'>💵 Salary Advances<br><b>PKR {total_advance:,.0f}</b></div>", unsafe_allow_html=True)
            st.markdown("---")

            for _, row in df_exp.iterrows():
                header_bits = [str(row["date"]), row["category"], f"PKR {row['amount']:,.0f}"]
                if not is_restricted:
                    header_bits.insert(1, f"👤 {row['user_id']}")
                ref_txt = f" | Ref: {row['reference_no']}" if str(row.get("reference_no") or "").strip() else ""
                row_c1, row_c2 = st.columns([5, 1])
                with row_c1:
                    st.markdown(f"**{' | '.join(header_bits)}**{ref_txt}")
                    if str(row.get("remarks") or "").strip():
                        st.caption(row["remarks"])
                with row_c2:
                    has_img = row.get("bill_image") is not None
                    if has_img:
                        with st.popover("👁️ View Receipt"):
                            try:
                                st.image(bytes(row["bill_image"]), width=300)
                            except Exception:
                                st.caption("⚠️ Could not load this receipt image.")
                    else:
                        st.caption("— no receipt —")
                if not is_restricted:
                    del_c1, del_c2 = st.columns([3, 1])
                    with del_c1:
                        confirm_del = st.checkbox(f"Confirm delete entry #{row['id']}", key=f"exp_del_chk_{row['id']}")
                    with del_c2:
                        if confirm_del and st.button("🗑️ Delete", key=f"exp_del_btn_{row['id']}"):
                            conn = get_conn()
                            conn.execute("DELETE FROM rider_expenses WHERE id=?", (int(row["id"]),))
                            conn.commit()
                            conn.close()
                            st.success("✅ Entry deleted.")
                            st.rerun()
                st.markdown("---")

            if not is_restricted:
                def _generate_expense_excel(df):
                    import openpyxl
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Staff Expenses"
                    export_cols = [c for c in df.columns if c != "bill_image"]
                    bold = openpyxl.styles.Font(bold=True)
                    for ci, col in enumerate(export_cols):
                        ws.cell(row=1, column=1 + ci, value=col).font = bold
                    for ri, (_, r) in enumerate(df.iterrows(), start=2):
                        for ci, col in enumerate(export_cols):
                            ws.cell(row=ri, column=1 + ci, value=r[col])
                    for col_letter in "ABCDEFGH":
                        ws.column_dimensions[col_letter].width = 18
                    buf = io.BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    return buf

                st.download_button(
                    label="⬇️ Export Ledger to Excel",
                    data=_generate_expense_excel(df_exp),
                    file_name=f"Staff_Expenses_{filt_rider if filt_rider != 'All' else 'AllRiders'}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="exp_export_excel"
                )

st.markdown("""
<div class="footer">
  🏭 NABA TECH BY KALEEM ULLAH SHARIF &nbsp;|&nbsp;
  Customer: Vertex (Shahzad Bhai) Lahore &nbsp;|&nbsp; v6.4 Cloud
</div>
""", unsafe_allow_html=True)
